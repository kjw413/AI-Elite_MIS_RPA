"""Read-only reconciliation of Gwangju production quantities across MIS workbooks."""

from __future__ import annotations

from datetime import date, timedelta
from pathlib import Path

from openpyxl import load_workbook

import energy_builder as eb


PRODUCTION_DB = Path(r"E:\DB_MIS\DB_생산실적.xlsx")
PRODUCTION_RAW = Path(r"E:\DB_MIS\RawDB_생산실적.xlsx")
ENERGY_DB = Path(r"E:\DB_MIS\DB_에너지.xlsx")
START = date(2026, 8, 1)
END = date(2026, 8, 17)


def load_energy_actuals(path: Path = ENERGY_DB) -> dict[str, float]:
    workbook = load_workbook(path, read_only=True, data_only=False, keep_links=False)
    try:
        sheet = workbook["광주"]
        headers = [sheet.cell(1, col).value for col in range(1, sheet.max_column + 1)]
        mix_col = headers.index("믹스생산량[kg]")
        return {
            eb._date_key(row[0]): float(row[mix_col] or 0)
            for row in sheet.iter_rows(min_row=2, values_only=True)
            if eb._date_key(row[0])
        }
    finally:
        workbook.close()


def reconcile() -> list[dict[str, float | str | None]]:
    db_actuals = eb._load_production_actuals(PRODUCTION_DB)
    raw_actuals = eb._load_raw_production_actuals(PRODUCTION_RAW)
    merged = dict(db_actuals)
    merged.update(raw_actuals)
    energy_actuals = load_energy_actuals()

    rows: list[dict[str, float | str | None]] = []
    day = START
    while day <= END:
        day_key = eb._date_key(day)
        db_value = db_actuals.get(("F30", day_key))
        raw_value = raw_actuals.get(("F30", day_key))
        energy_value = energy_actuals.get(day_key)
        rows.append(
            {
                "date": day.isoformat(),
                "db_production_kg": db_value,
                "raw_override_kg": raw_value,
                "merged_kg": merged.get(("F30", day_key)),
                "db_energy_kg": energy_value,
                "gap_kg": None if db_value is None or energy_value is None else db_value - energy_value,
            }
        )
        day += timedelta(days=1)
    return rows


if __name__ == "__main__":
    rows = reconcile()
    print("date,db_production_kg,raw_override_kg,merged_kg,db_energy_kg,gap_kg")
    for row in rows:
        print(
            f"{row['date']},{row['db_production_kg']},{row['raw_override_kg']},"
            f"{row['merged_kg']},{row['db_energy_kg']},{row['gap_kg']}"
        )
    verified = [row for row in rows if row["db_production_kg"] is not None]
    db_total = sum(float(row["db_production_kg"] or 0) for row in verified)
    energy_total = sum(float(row["db_energy_kg"] or 0) for row in verified)
    print(f"verified_days={len(verified)}")
    print(f"db_total_kg={db_total:.0f}")
    print(f"energy_total_kg={energy_total:.0f}")
    print(f"gap_total_kg={db_total - energy_total:.0f}")
