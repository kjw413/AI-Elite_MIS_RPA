# 에너지 수집 원본/웹 입력 파일의 스키마 정의 · 적재 · 읽기 · 가공
"""
MIS '원단위 실적입력(일단위)' 화면에서 수집한 에너지 실적을 별도 원본 파일에
영속 저장하고, 생산량·원단위를 보완한 `DB_에너지.xlsx` 만 BEMS 웹앱에 전달한다.

    [MIS 원단위 실적입력(일단위)]
      냉동전력/공압기/전력량/전력비/전력단가/연료량/연료비/연료단가/
      용수량/폐수량/원수COD/배출수COD
                     │  utility_daily_rpa (수집)
                     ▼
          RawDB_에너지.xlsx     행 = 일자, 열 = MIS 원본 항목
                     │  build_energy_dataset (가공)
                     ▼
              DB_에너지.xlsx     생산량·원단위 포함
              시트 = 공장명          → BEMS 웹앱이 startup 에 직접 적재

행=일자 방향은 `DB_생산실적.xlsx` 의 `daily` 시트, `DB_재공품.xlsx` 와 통일된 형태다.

== 이력 ==
- 2026-07 수집 화면이 '유틸리티 일자별 사용량 추이' → '원단위 실적입력(일단위)' 로
  바뀌며 단가·비용·COD 가 추가됐다.
- 과거의 `DB_에너지.xlsx`는 행=항목·열=날짜인 전치형이었으나 2026-07-30 폐지했다.
  으로 재가공했으나, BEMS 파서가 tidy 형태를 그대로 읽을 수 있어(전치 감지 분기를 타지
  않고 머리글 부분매칭만으로 컬럼이 잡힘) 2026-07-30 폐지했다.
- 수집 원본은 생산실적 유무와 관계없이 먼저 보존한다. 가공 단계에서
  `RawDB_생산실적.xlsx`의 공장별 생산량을 우선 동기화하고 DB 파일로 과거 기간을
  보완한다. 광주는 품목별 믹스 환산 후 합산하며, 생산량이 없으면 최종 파일 가공만
  중단해 빈 원단위를 만들지 않는다.
- 냉동·공압·전력·연료·용수 원단위는 Python에서 계산하지 않고 RawDB 수식으로
  관리한다. 빈 수식을 자동 보완한 뒤 Excel에서 전체 재계산·저장해 유틸리티 실적
  변경을 바로 반영한다.
"""

from __future__ import annotations

import logging
from collections.abc import Sequence
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path

import openpyxl
from openpyxl.formula.translate import Translator
from openpyxl.utils import get_column_letter

from _common import atomic_save_workbook, sampled_db_path
from factories import (
    FACTORY_KR_TO_CODE,
    FACTORY_PHYSICAL_DISPLAY_ORDER,
)

log = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# 경로
# ---------------------------------------------------------------------------
# MIS 좌표 수집 결과를 가공과 무관하게 보존하는 원본 파일.
DEFAULT_RAW_PATH = Path(
    sampled_db_path("RawDB_에너지.xlsx", "ENERGY_RAW_XLSX")
)

# 가공 산출물 = 웹앱 입력 파일. env 이름은 BEMS `v5_common.PATH_ENERGY_SOURCE`
# 와 동일한 `ENERGY_SOURCE_XLSX` 를 쓴다 — 한 변수로 양쪽이 같은 파일을 가리킨다.
DEFAULT_OUTPUT_PATH = Path(sampled_db_path("DB_에너지.xlsx", "ENERGY_SOURCE_XLSX"))
DEFAULT_PRODUCTION_PATH = Path(
    sampled_db_path("DB_생산실적.xlsx", "PRODUCTION_DW_XLSX")
)
DEFAULT_PRODUCTION_RAW_PATH = Path(
    sampled_db_path("RawDB_생산실적.xlsx", "PRODUCTION_RAW_XLSX")
)
DEFAULT_WIP_PATH = Path(
    sampled_db_path("DB_재공품.xlsx", "WIP_SUMMARY_XLSX")
)

FACTORY_SHEETS: tuple[str, ...] = FACTORY_PHYSICAL_DISPLAY_ORDER


# ---------------------------------------------------------------------------
# 스키마 — RawDB_에너지 열 정의의 단일 출처
# ---------------------------------------------------------------------------
@dataclass(frozen=True)
class EnergyField:
    key: str        # 코드 내부 식별자
    label: str      # RawDB_에너지 열 머리글 (엑셀에 그대로 쓰이는 문자열)
    source: str     # 'unit_input' = MIS 원단위 실적입력 화면에서 수집
                    # 'production' = 최신 생산 RawDB 우선 믹스 환산 kg 합계
                    # 'formula'    = RawDB 엑셀 수식으로 계산하는 원단위


# 전 항목 정의. 라벨 문자열은 BEMS 파서가 **부분매칭**으로 컬럼을 찾는 키이므로
# (`excel_parser.kor_to_eng`, `daily_energy_sync_service._KOR_SUBSTR_MAP`)
# 바꿀 때는 양쪽 매핑을 함께 확인해야 한다.
FIELDS: tuple[EnergyField, ...] = (
    EnergyField("freezing_power_kwh",   "냉동전력량[kWh]",            "unit_input"),
    EnergyField("air_compressor_kwh",   "공압기[kWh]",                "unit_input"),
    EnergyField("total_power_kwh",      "전력량[kWh]",                "unit_input"),
    EnergyField("fuel_nm3",             "연료량[N㎥]",                "unit_input"),
    EnergyField("water_ton",            "용수량[ton]",                "unit_input"),
    EnergyField("wastewater_ton",       "폐수량[ton]",                "unit_input"),
    EnergyField("mix_prod_kg",          "믹스생산량[kg]",             "production"),
    EnergyField("freezing_per_ton_kwh", "냉동원단위[kWh/mix-ton]",    "formula"),
    EnergyField(
        "air_compressor_per_ton_kwh", "공압기원단위[kWh/mix-ton]", "formula",
    ),
    EnergyField("power_per_ton_kwh",    "전력원단위[kWh/mix-ton]",    "formula"),
    EnergyField("fuel_per_ton_nm3",     "연료원단위[N㎥/mix-ton]",    "formula"),
    EnergyField("water_per_ton_ton",    "용수원단위[ton/mix-ton]",    "formula"),
    EnergyField("power_cost_krw",       "전력비[원]",                 "unit_input"),
    EnergyField("power_price_krw_kwh",  "전력단가[원/kWh]",           "unit_input"),
    EnergyField("fuel_cost_krw",        "연료비[원]",                 "unit_input"),
    EnergyField("fuel_price_krw_nm3",   "연료단가[원/N㎥]",           "unit_input"),
    EnergyField("influent_cod_ppm",     "원수COD[ppm]",               "unit_input"),
    EnergyField("effluent_cod_ppm",     "배출수COD[ppm]",             "unit_input"),
)

FIELD_BY_KEY: dict[str, EnergyField] = {f.key: f for f in FIELDS}
COLLECTED_FIELD_KEYS: frozenset[str] = frozenset(
    field.key for field in FIELDS if field.source == "unit_input"
)

# RawDB_에너지 열 순서 — MIS '원단위 실적입력' 화면 열 순서를 그대로 따르고,
# legacy 항목(믹스/원단위)을 뒤에 붙인다. 사람이 원본과 눈으로 대조하기 쉽게.
RAW_COLUMN_KEYS: tuple[str, ...] = (
    "freezing_power_kwh", "air_compressor_kwh",
    "total_power_kwh", "power_cost_krw", "power_price_krw_kwh",
    "fuel_nm3", "fuel_cost_krw", "fuel_price_krw_nm3",
    "water_ton", "wastewater_ton",
    "influent_cod_ppm", "effluent_cod_ppm",
    "mix_prod_kg",
    "freezing_per_ton_kwh", "air_compressor_per_ton_kwh",
    "power_per_ton_kwh", "fuel_per_ton_nm3", "water_per_ton_ton",
)
RAW_DATE_HEADER = "날짜"

# 원단위는 사용량 × 1,000 / 믹스생산량[kg] 수식으로 관리한다. 기존 RawDB에
# 사용자가 만든 수식이 있으면 행 참조만 옮겨 복제하고, 템플릿이 없을 때만 아래
# 열 매핑으로 같은 수식을 생성한다.
UNIT_FORMULA_SOURCE_KEYS: dict[str, str] = {
    "freezing_per_ton_kwh": "freezing_power_kwh",
    "air_compressor_per_ton_kwh": "air_compressor_kwh",
    "power_per_ton_kwh": "total_power_kwh",
    "fuel_per_ton_nm3": "fuel_nm3",
    "water_per_ton_ton": "water_ton",
}

_PRODUCTION_CACHE_PATH: Path | None = None
_PRODUCTION_CACHE_MTIME_NS: int | None = None
_PRODUCTION_CACHE_INCLUDE_GWANGJU_WIP: bool | None = None
_PRODUCTION_CACHE: dict[tuple[str, str], float] | None = None
_PRODUCTION_RAW_CACHE_PATH: Path | None = None
_PRODUCTION_RAW_CACHE_MTIME_NS: int | None = None
_PRODUCTION_RAW_CACHE_INCLUDE_GWANGJU_WIP: bool | None = None
_PRODUCTION_RAW_CACHE: dict[tuple[str, str], float] | None = None
_WIP_CACHE_PATH: Path | None = None
_WIP_CACHE_MTIME_NS: int | None = None
_WIP_CACHE: dict[tuple[str, str], float] | None = None


# 광주(F30) 품목별 믹스 환산계수. BEMS Next의
# production_correction_service.py와 같은 기준을 사용한다. 129998/129999는 MIS
# 생산실적에 재공품 성격으로 기록되는 품목이고, 260xxx는 재공품 품목이다.
GWANGJU_FACTORY_CODE = FACTORY_KR_TO_CODE["광주"]
GWANGJU_SKIM_MILK_MIX_FACTOR = 10.91954
GWANGJU_WIP_MIX_CONVERSION: dict[str, float] = {
    "260014": GWANGJU_SKIM_MILK_MIX_FACTOR,
    "260016": 1.0,
    "260039": 1.0,
    "260042": 4.0,
    "260047": 1.0,
    "260351": 1.0,
    "260352": 1.0,
}
GWANGJU_PRODUCTION_RECORDED_WIP_MIX_CONVERSION: dict[str, float] = {
    "129998": GWANGJU_SKIM_MILK_MIX_FACTOR,
    "129999": 1.0,
}
GWANGJU_MIX_CONVERSION: dict[str, float] = {
    **GWANGJU_WIP_MIX_CONVERSION,
    **GWANGJU_PRODUCTION_RECORDED_WIP_MIX_CONVERSION,
}


def _raw_column(field_key: str) -> int:
    return RAW_COLUMN_KEYS.index(field_key) + 2


def _date_key(value) -> str | None:
    """셀 값을 'YY-MM-DD' 비교 키로 정규화한다 (date/datetime/문자열 혼재 대응)."""
    if value is None:
        return None
    if isinstance(value, (datetime, date)):
        return value.strftime("%y-%m-%d")
    text = str(value).strip().lstrip("'")
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%y-%m-%d", "%Y/%m/%d", "%y/%m/%d"):
        try:
            return datetime.strptime(text, fmt).strftime("%y-%m-%d")
        except ValueError:
            continue
    return text


def _as_date(value) -> date | None:
    """셀 값을 date 로 변환한다 (기간 필터 비교용). 날짜가 아니면 None."""
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip().lstrip("'") if value is not None else ""
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%y-%m-%d", "%Y/%m/%d", "%y/%m/%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _item_code(value) -> str:
    """엑셀 숫자/문자 품목코드를 비교 가능한 문자열로 정규화한다."""
    if value is None:
        return ""
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value).strip()
    if text.endswith(".0"):
        try:
            number = float(text)
        except ValueError:
            pass
        else:
            if number.is_integer():
                return str(int(number))
    return text


def _mix_equivalent_qty(factory: str, item_code, actual_qty: float) -> float:
    """광주만 품목별 계수를 적용하고 다른 공장은 기존 수량을 그대로 돌려준다."""
    amount = float(actual_qty)
    if str(factory).strip() != GWANGJU_FACTORY_CODE:
        return amount
    factor = GWANGJU_MIX_CONVERSION.get(_item_code(item_code), 1.0)
    return amount * factor


def _aggregate_production_rows(
    rows, *, include_gwangju_wip: bool = True,
) -> dict[tuple[str, str], float]:
    """품목별 환산 후 공장·일자 합계를 만든다.

    광주 daily의 동일한 (일자, 품목, 수량) 행이 여러 시트/경로에서 반복된 경우에는
    한 번만 반영한다. 환산 전 수량과 환산 후 수량을 함께 더하지 않고, 각 원천 행은
    환산된 값 하나로만 합계에 기여한다. DB_재공품을 별도 합산하는 운영 경로에서는
    광주 260xxx를 여기서 제외해 중복을 막는다. 비광주는 기존 단순 합산을 유지한다.
    """
    actuals: dict[tuple[str, str], float] = {}
    seen_gwangju_rows: set[tuple[str, str, float]] = set()
    for day_value, factory_value, item_value, quantity_value in rows:
        day_key = _date_key(day_value)
        factory = str(factory_value or "").strip()
        if not day_key or not factory or quantity_value is None:
            continue
        try:
            amount = float(quantity_value)
        except (TypeError, ValueError):
            continue

        item_code = _item_code(item_value)
        if (factory == GWANGJU_FACTORY_CODE and not include_gwangju_wip
                and item_code in GWANGJU_WIP_MIX_CONVERSION):
            continue
        if factory == GWANGJU_FACTORY_CODE:
            duplicate_key = (day_key, item_code, amount)
            if duplicate_key in seen_gwangju_rows:
                log.warning(
                    "광주 생산실적 중복 행 제외: %s / %s / %s",
                    day_key, item_code or "(품목코드 없음)", amount,
                )
                continue
            seen_gwangju_rows.add(duplicate_key)

        key = (factory, day_key)
        actuals[key] = actuals.get(key, 0.0) + _mix_equivalent_qty(
            factory, item_code, amount
        )
    return actuals


def _fill_internal_zero_days(actuals: dict[tuple[str, str], float]) -> None:
    """공장별 생산실적 범위 안의 미등장 날짜를 무생산(0kg)으로 명시한다."""
    ranges: dict[str, tuple[date, date]] = {}
    for factory, day_key in actuals:
        day = datetime.strptime(day_key, "%y-%m-%d").date()
        current = ranges.get(factory)
        ranges[factory] = (
            min(current[0], day) if current else day,
            max(current[1], day) if current else day,
        )
    for factory, (start, end) in ranges.items():
        day = start
        while day <= end:
            actuals.setdefault((factory, _date_key(day)), 0.0)
            day += timedelta(days=1)


def _load_production_actuals(
    production_path: Path, *, include_gwangju_wip: bool = True,
) -> dict[tuple[str, str], float]:
    """DB_생산실적 daily를 공장·일자별 믹스 환산 kg 합계로 읽는다.

    광주는 품목별 환산 후 합산하고, 다른 공장은 actual_qty를 그대로 합산한다.
    """
    global _PRODUCTION_CACHE_PATH, _PRODUCTION_CACHE_MTIME_NS
    global _PRODUCTION_CACHE_INCLUDE_GWANGJU_WIP, _PRODUCTION_CACHE

    production_path = production_path.resolve()
    if not production_path.exists():
        raise FileNotFoundError(f"생산실적 파일이 없습니다: {production_path}")

    mtime_ns = production_path.stat().st_mtime_ns
    if (
        _PRODUCTION_CACHE is not None
        and _PRODUCTION_CACHE_PATH == production_path
        and _PRODUCTION_CACHE_MTIME_NS == mtime_ns
        and _PRODUCTION_CACHE_INCLUDE_GWANGJU_WIP == include_gwangju_wip
    ):
        return _PRODUCTION_CACHE

    wb = openpyxl.load_workbook(
        production_path, read_only=True, data_only=True, keep_links=False
    )
    try:
        if "daily" not in wb.sheetnames:
            raise ValueError(f"생산실적 파일에 daily 시트가 없습니다: {production_path}")
        rows = wb["daily"].iter_rows(values_only=True)
        try:
            header = next(rows)
        except StopIteration as exc:
            raise ValueError(f"생산실적 daily 시트가 비어 있습니다: {production_path}") from exc

        columns = {str(value).strip(): idx for idx, value in enumerate(header) if value}
        required = {"date", "factory", "actual_qty"}
        missing = required - set(columns)
        if missing:
            raise ValueError(
                f"생산실적 daily 필수 컬럼 누락: {sorted(missing)} ({production_path})"
            )

        item_col = columns.get("item_code")
        production_rows = []
        for row in rows:
            factory = str(row[columns["factory"]] or "").strip()
            if factory == GWANGJU_FACTORY_CODE and item_col is None:
                raise ValueError(
                    "광주 생산량 환산에 필요한 item_code 컬럼이 없습니다: "
                    f"{production_path}"
                )
            production_rows.append((
                row[columns["date"]],
                factory,
                row[item_col] if item_col is not None else None,
                row[columns["actual_qty"]],
            ))
        actuals = _aggregate_production_rows(
            production_rows, include_gwangju_wip=include_gwangju_wip
        )
    finally:
        wb.close()

    _fill_internal_zero_days(actuals)
    _PRODUCTION_CACHE_PATH = production_path
    _PRODUCTION_CACHE_MTIME_NS = mtime_ns
    _PRODUCTION_CACHE_INCLUDE_GWANGJU_WIP = include_gwangju_wip
    _PRODUCTION_CACHE = actuals
    log.info("생산실적 daily 로드: %s (%s개 공장·일자)", production_path, len(actuals))
    return actuals


def _load_raw_production_actuals(
    production_raw_path: Path, *, include_gwangju_wip: bool = True,
) -> dict[tuple[str, str], float]:
    """최신 RawDB_생산실적 월 구간을 공장·일자별 믹스 환산 kg로 읽는다.

    DB daily 로더와 동일하게 광주만 품목별 환산하고 다른 공장은 단순 합산한다.
    """
    global _PRODUCTION_RAW_CACHE_PATH
    global _PRODUCTION_RAW_CACHE_MTIME_NS
    global _PRODUCTION_RAW_CACHE_INCLUDE_GWANGJU_WIP, _PRODUCTION_RAW_CACHE

    production_raw_path = production_raw_path.resolve()
    if not production_raw_path.exists():
        raise FileNotFoundError(f"생산 RawDB 파일이 없습니다: {production_raw_path}")

    mtime_ns = production_raw_path.stat().st_mtime_ns
    if (
        _PRODUCTION_RAW_CACHE is not None
        and _PRODUCTION_RAW_CACHE_PATH == production_raw_path
        and _PRODUCTION_RAW_CACHE_MTIME_NS == mtime_ns
        and _PRODUCTION_RAW_CACHE_INCLUDE_GWANGJU_WIP == include_gwangju_wip
    ):
        return _PRODUCTION_RAW_CACHE

    import production_builder

    daily = production_builder.consolidate_raw_file(production_raw_path)
    production_rows = (
        (
            (row.date, row.factory, row.item_code, row.actual_qty)
            for row in daily.itertuples(index=False)
        )
        if not daily.empty
        else ()
    )
    actuals = _aggregate_production_rows(
        production_rows, include_gwangju_wip=include_gwangju_wip
    )

    # MIS 생산 화면이 무생산일 열을 생략해도, 각 시트의 기간 마커 안 날짜는 0kg로
    # 간주한다. 최신 종료일 이후 날짜는 만들지 않아 미수집 상태와 무생산을 구분한다.
    wb = openpyxl.load_workbook(
        production_raw_path, read_only=True, data_only=True, keep_links=False
    )
    try:
        for sheet_name in wb.sheetnames:
            meta = production_builder.parse_meta_from_sheet(sheet_name)
            if meta is None:
                continue
            marker = [wb[sheet_name].cell(1, col).value for col in range(1, 4)]
            if str(marker[0]).strip() != production_builder.PERIOD_MARKER:
                continue
            start = marker[1].date() if isinstance(marker[1], datetime) else marker[1]
            end = marker[2].date() if isinstance(marker[2], datetime) else marker[2]
            if not isinstance(start, date) or not isinstance(end, date):
                start = datetime.fromisoformat(str(marker[1])).date()
                end = datetime.fromisoformat(str(marker[2])).date()
            factories = ("F10A", "F10B") if meta.factory == "F10" else (meta.factory,)
            day = min(start, end)
            last = max(start, end)
            while day <= last:
                for factory in factories:
                    actuals.setdefault((factory, _date_key(day)), 0.0)
                day += timedelta(days=1)
    finally:
        wb.close()

    _PRODUCTION_RAW_CACHE_PATH = production_raw_path
    _PRODUCTION_RAW_CACHE_MTIME_NS = mtime_ns
    _PRODUCTION_RAW_CACHE_INCLUDE_GWANGJU_WIP = include_gwangju_wip
    _PRODUCTION_RAW_CACHE = actuals
    log.info(
        "최신 생산 RawDB 로드: %s (%s개 공장·일자)",
        production_raw_path, len(actuals),
    )
    return actuals


def _load_gwangju_wip_actuals(
    wip_path: Path,
) -> dict[tuple[str, str], float]:
    """DB_재공품 광주 시트를 일자별 믹스 환산 kg 합계로 읽는다.

    재공품 가공 단계가 job_number 반영 여부를 적용해 만든 DB를 권위 원천으로
    사용한다. 품목 열이 없는 기준은 0으로 보고, 날짜 중복 행은 마지막 행만 사용한다.
    """
    global _WIP_CACHE_PATH, _WIP_CACHE_MTIME_NS, _WIP_CACHE

    wip_path = wip_path.resolve()
    if not wip_path.exists():
        raise FileNotFoundError(f"재공품 파일이 없습니다: {wip_path}")

    mtime_ns = wip_path.stat().st_mtime_ns
    if (
        _WIP_CACHE is not None
        and _WIP_CACHE_PATH == wip_path
        and _WIP_CACHE_MTIME_NS == mtime_ns
    ):
        return _WIP_CACHE

    wb = openpyxl.load_workbook(
        wip_path, read_only=True, data_only=True, keep_links=False
    )
    try:
        sheet_name = next(
            (name for name in ("광주", "F30", "Sheet1") if name in wb.sheetnames),
            None,
        )
        if sheet_name is None:
            raise ValueError(f"재공품 파일에 광주 시트가 없습니다: {wip_path}")

        rows = wb[sheet_name].iter_rows(values_only=True)
        try:
            header = next(rows)
        except StopIteration as exc:
            raise ValueError(f"재공품 광주 시트가 비어 있습니다: {wip_path}") from exc

        item_columns = {
            idx: code
            for idx, value in enumerate(header)
            if (code := _item_code(value)) in GWANGJU_WIP_MIX_CONVERSION
        }
        if not item_columns:
            raise ValueError(
                "재공품 광주 시트에 환산 대상 품목 열이 없습니다: "
                f"{wip_path}"
            )

        actuals: dict[tuple[str, str], float] = {}
        for row in rows:
            if not row:
                continue
            day_key = _date_key(row[0])
            if not day_key:
                continue
            total = 0.0
            for idx, item_code in item_columns.items():
                value = row[idx] if idx < len(row) else None
                if value in (None, ""):
                    continue
                try:
                    amount = float(value)
                except (TypeError, ValueError):
                    log.warning(
                        "광주 재공품 숫자 변환 실패 제외: %s / %s / %r",
                        day_key, item_code, value,
                    )
                    continue
                total += amount * GWANGJU_WIP_MIX_CONVERSION[item_code]

            key = (GWANGJU_FACTORY_CODE, day_key)
            if key in actuals:
                log.warning("광주 재공품 날짜 중복: %s (마지막 행 사용)", day_key)
            actuals[key] = total
    finally:
        wb.close()

    _WIP_CACHE_PATH = wip_path
    _WIP_CACHE_MTIME_NS = mtime_ns
    _WIP_CACHE = actuals
    log.info("광주 재공품 DB 로드: %s (%s일)", wip_path, len(actuals))
    return actuals


# ---------------------------------------------------------------------------
# RawDB_에너지 (행=일자, 열=항목) 읽기/쓰기
# ---------------------------------------------------------------------------
def _ensure_raw_sheet(wb, sheet_name: str):
    """RawDB 시트를 확보하고 머리글 행을 보정한다."""
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(sheet_name)

    headers = [RAW_DATE_HEADER] + [FIELD_BY_KEY[k].label for k in RAW_COLUMN_KEYS]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        if cell.value != header:
            cell.value = header
    return ws


def _nearest_formula_cell(ws, column: int, target_row: int):
    """대상 행에서 가장 가까운 기존 수식 셀을 찾는다."""
    for distance in range(1, max(target_row, ws.max_row) + 1):
        for row_idx in (target_row - distance, target_row + distance):
            if row_idx < 2 or row_idx > ws.max_row:
                continue
            cell = ws.cell(row=row_idx, column=column)
            if isinstance(cell.value, str) and cell.value.startswith("="):
                return cell
    return None


def _default_unit_formula(unit_key: str, row_idx: int) -> str:
    usage_col = get_column_letter(_raw_column(UNIT_FORMULA_SOURCE_KEYS[unit_key]))
    mix_col = get_column_letter(_raw_column("mix_prod_kg"))
    return f"={usage_col}{row_idx}*1000/${mix_col}{row_idx}"


def _ensure_unit_formulas(ws, row_indices: set[int]) -> int:
    """수집으로 추가·갱신된 날짜 행의 빈 원단위 셀에 수식을 채운다."""
    filled = 0
    for row_idx in sorted(row_indices):
        if _date_key(ws.cell(row=row_idx, column=1).value) is None:
            continue
        for unit_key in UNIT_FORMULA_SOURCE_KEYS:
            target = ws.cell(row=row_idx, column=_raw_column(unit_key))
            if target.value not in (None, ""):
                continue

            template = _nearest_formula_cell(ws, target.column, row_idx)
            if template is not None:
                try:
                    target.value = Translator(
                        template.value, origin=template.coordinate
                    ).translate_formula(target.coordinate)
                except Exception:
                    log.warning(
                        "[%s] %s 수식 복제 실패, 기본 수식 사용",
                        ws.title, target.coordinate, exc_info=True,
                    )
                    target.value = _default_unit_formula(unit_key, row_idx)
            else:
                target.value = _default_unit_formula(unit_key, row_idx)
            filled += 1
    return filled


def _recalculate_with_excel(raw_path: Path) -> bool:
    """Excel COM으로 수식 전체를 재계산하고 계산 캐시까지 저장한다."""
    try:
        import pythoncom
        import win32com.client
    except ImportError:
        log.warning("pywin32가 없어 RawDB 수식 재계산을 건너뜁니다: %s", raw_path)
        return False

    excel = workbook = None
    pythoncom.CoInitialize()
    try:
        excel = win32com.client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        workbook = excel.Workbooks.Open(
            str(raw_path.resolve()), UpdateLinks=0, ReadOnly=False
        )
        excel.CalculateFullRebuild()
        workbook.Save()
        log.info("RawDB 원단위 수식 재계산 완료: %s", raw_path)
        return True
    except Exception:
        # openpyxl 저장은 이미 끝났으므로 수식 자체는 보존된다. 다음 RPA/수동 Excel
        # 열기에서 재계산할 수 있도록 워크북 계산 속성도 함께 지정해 둔다.
        log.exception("RawDB 수식 Excel 재계산 실패: %s", raw_path)
        return False
    finally:
        if workbook is not None:
            try:
                workbook.Close(SaveChanges=False)
            except Exception:
                pass
        if excel is not None:
            try:
                excel.Quit()
            except Exception:
                pass
        pythoncom.CoUninitialize()


def merge_production_actuals(
    production_path: Path | None = None,
    production_raw_path: Path | None = None,
    wip_path: Path | None = None,
) -> dict[tuple[str, str], float]:
    """믹스생산량의 권위 값 {(공장코드, 'YY-MM-DD'): kg} 을 만든다.

    생산실적 두 입력 모두 광주 품목별 환산을 끝낸 공장·일자 합계다. 기본은
    `DB_생산실적.xlsx`(가공 완료본)이고, 그 위에 `RawDB_생산실적.xlsx`(최신 수집분)를
    공장·일자 단위로 덮어써 같은 실적이 두 번 합산되지 않게 한다. 수집과 가공이
    분리돼 있어도 두 경로가 같은 규칙을 쓴다. 운영 기본 경로에서는 광주 260xxx를
    생산실적에서 제외하고 `DB_재공품.xlsx` 환산 합계를 더해 중복을 막는다.

    production_path만 지정하면(테스트 등) 최신 RawDB와 기본 재공품 DB는 섞지 않는다.
    wip_path를 명시하면 해당 재공품 DB를 합산한다.
    """
    wip_source = (
        Path(wip_path) if wip_path is not None
        else DEFAULT_WIP_PATH if production_path is None
        else None
    )
    include_gwangju_wip = wip_source is None
    actuals = dict(_load_production_actuals(
        Path(production_path or DEFAULT_PRODUCTION_PATH),
        include_gwangju_wip=include_gwangju_wip,
    ))
    raw_source = (
        Path(production_raw_path) if production_raw_path is not None
        else DEFAULT_PRODUCTION_RAW_PATH if production_path is None
        else None
    )
    if raw_source is not None:
        actuals.update(_load_raw_production_actuals(
            raw_source, include_gwangju_wip=include_gwangju_wip
        ))
    if wip_source is not None:
        for key, amount in _load_gwangju_wip_actuals(wip_source).items():
            actuals[key] = actuals.get(key, 0.0) + amount
    return actuals


def resolve_sheet_names(spec: str | None) -> list[str]:
    """'김해,F50' → ['김해', '경산']. 공장명·공장코드를 모두 받는다.

    MIS ORG 코드(F1A 등)가 아니라 `factories.FACTORY_MASTER` 의 f_code 를 쓴다 —
    이 모듈은 MIS 화면을 건드리지 않고 엑셀만 다루기 때문이다.
    """
    if not spec:
        return list(FACTORY_SHEETS)

    by_code = {code: name for name, code in FACTORY_KR_TO_CODE.items()}
    selected: set[str] = set()
    unknown: list[str] = []
    for token in spec.split(","):
        token = token.strip()
        if not token:
            continue
        if token in FACTORY_KR_TO_CODE:
            selected.add(token)
        elif token.upper() in by_code:
            selected.add(by_code[token.upper()])
        else:
            unknown.append(token)

    if unknown:
        raise SystemExit(
            f"알 수 없는 사업장: {unknown}\n"
            "  공장명 또는 공장코드를 쓰세요 — "
            + ", ".join(f"{n}={c}" for n, c in FACTORY_KR_TO_CODE.items())
        )
    if not selected:
        raise SystemExit("--factories 에 사업장이 하나도 지정되지 않았습니다.")
    return [name for name in FACTORY_SHEETS if name in selected]


def resync_production(
    raw_path: Path | None = None, *,
    date_from: date | None = None,
    date_to: date | None = None,
    sheet_names: Sequence[str] | None = None,
    production_path: Path | None = None,
    production_raw_path: Path | None = None,
    wip_path: Path | None = None,
    recalculate: bool = True,
    dry_run: bool = False,
) -> tuple[dict[str, dict[str, int]], list[tuple[str, date, float | None, float]]]:
    """MIS 접속 없이 `DB_에너지.xlsx` 의 믹스생산량만 생산실적으로 다시 맞춘다.

    수집(MIS 클릭)과 가공(엑셀 재집계)을 분리하기 위한 진입점. 생산실적이 늦게
    입력돼 수집 시점에 0 이 박혔거나, 구 '유틸리티 일자별' 화면 값이 남아 있는
    행을 사후에 정정할 때 쓴다. 사용량·단가·비용 열은 건드리지 않는다.

    Args:
        date_from/date_to: 대상 날짜 범위 (None 이면 무제한)
        sheet_names: 대상 시트. None 이면 전체
        dry_run: 파일을 저장하지 않고 변경 예정만 돌려준다

    Returns:
        (시트별 통계, 변경 목록[(시트, 날짜, 이전값, 새값)])
    """
    raw_path = Path(raw_path or DEFAULT_OUTPUT_PATH)
    if not raw_path.exists():
        raise FileNotFoundError(f"에너지 파일이 없습니다: {raw_path}")

    actuals = merge_production_actuals(
        production_path, production_raw_path, wip_path
    )
    targets = set(sheet_names) if sheet_names else None
    mix_col = _raw_column("mix_prod_kg")

    wb = openpyxl.load_workbook(raw_path)   # 원단위 수식 보존 (data_only=False)
    stats: dict[str, dict[str, int]] = {}
    changes: list[tuple[str, date, float | None, float]] = []
    formula_updates = 0
    try:
        for sheet_name in wb.sheetnames:
            factory_code = FACTORY_KR_TO_CODE.get(sheet_name)
            if factory_code is None:
                continue
            if targets is not None and sheet_name not in targets:
                continue

            ws = wb[sheet_name]
            updated = unchanged = missing = 0
            touched_rows: set[int] = set()
            for row_idx in range(2, ws.max_row + 1):
                day = _as_date(ws.cell(row=row_idx, column=1).value)
                if day is None:
                    continue
                if date_from and day < date_from:
                    continue
                if date_to and day > date_to:
                    continue

                want = actuals.get((factory_code, _date_key(day)))
                if want is None:
                    # 생산실적이 아직 없는 날 — 기존 값을 지우지 않고 남겨 둔다
                    missing += 1
                    continue

                cell = ws.cell(row=row_idx, column=mix_col)
                current = cell.value
                current_num = float(current) if isinstance(current, (int, float)) else None
                if current_num is not None and abs(current_num - want) <= 0.5:
                    unchanged += 1
                else:
                    changes.append((sheet_name, day, current_num, want))
                    if not dry_run:
                        cell.value = want
                    updated += 1
                touched_rows.add(row_idx)

            if not dry_run and touched_rows:
                formula_cnt = _ensure_unit_formulas(ws, touched_rows)
                formula_updates += formula_cnt
                if formula_cnt:
                    log.info(f"  [{sheet_name}] 원단위 수식 {formula_cnt}셀 보완")

            stats[sheet_name] = {
                "updated": updated, "unchanged": unchanged, "missing": missing,
            }
            log.info(f"  [{sheet_name}] 갱신 {updated} / 동일 {unchanged} / "
                     f"생산실적 없음 {missing}")

        if dry_run:
            return stats, changes

        if not changes and formula_updates == 0:
            log.info("변경할 값이 없어 파일을 저장하지 않습니다.")
            return stats, changes

        wb.calculation.calcMode = "auto"
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
        wb.calculation.calcOnSave = True
        atomic_save_workbook(wb, str(raw_path))
    finally:
        wb.close()

    if recalculate and not _recalculate_with_excel(raw_path):
        raise RuntimeError(
            "에너지 파일 저장은 완료됐지만 Excel 수식 재계산에 실패했습니다. "
            "파일 잠금과 Excel 설치 상태를 확인하세요."
        )
    log.info(
        f"DB_에너지 생산량 재집계 완료: {raw_path}  "
        f"({len(changes)}셀 변경 / 수식 {formula_updates}셀 보완)"
    )
    return stats, changes


def write_raw(records: dict[str, dict[date, dict[str, float]]],
              raw_path: Path | None = None, *,
              production_path: Path | None = None,
              production_raw_path: Path | None = None,
              wip_path: Path | None = None,
              sync_production: bool = True,
              recalculate: bool = True,
              ensure_formulas: bool = True) -> Path:
    """에너지 데이터를 지정 파일에 upsert 한다 (날짜 1건 = 1행).

    Args:
        records: { 시트명: { date: { field_key: value } } }
                 값이 None 인 항목은 기록하지 않는다 (기존 값 보존).
        production_path: 과거 생산량을 보완할 DB_생산실적.xlsx 경로.
        production_raw_path: 최신 월을 우선 반영할 RawDB_생산실적.xlsx 경로.
        wip_path: 광주 재공품을 환산·합산할 DB_재공품.xlsx 경로.
        sync_production: True이면 월 전체 생산량 합계를 N열에 다시 쓴다. 광주는
                         품목별 믹스 환산을 먼저 적용한다.
        recalculate: True이면 저장 후 Excel COM으로 수식과 계산 캐시를 갱신한다.
        ensure_formulas: True이면 적재 행의 빈 원단위 수식을 보완한다.
    """
    raw_path = Path(raw_path or DEFAULT_OUTPUT_PATH)
    raw_path.parent.mkdir(parents=True, exist_ok=True)

    production_actuals: dict[tuple[str, str], float] | None = None
    if sync_production:
        production_actuals = merge_production_actuals(
            production_path, production_raw_path, wip_path
        )

        missing_production: list[str] = []
        for sheet_name, by_date in records.items():
            factory_code = FACTORY_KR_TO_CODE.get(sheet_name)
            if factory_code is None:
                missing_production.append(f"{sheet_name}(공장코드 없음)")
                continue
            for day in by_date:
                day_key = _date_key(day)
                if (factory_code, day_key) not in production_actuals:
                    missing_production.append(f"{sheet_name}/{day_key}")
        if missing_production:
            preview = ", ".join(missing_production[:10])
            suffix = (
                "" if len(missing_production) <= 10
                else f" 외 {len(missing_production) - 10}건"
            )
            raise ValueError(
                "생산실적이 없어 RawDB 원단위를 만들 수 없습니다. "
                f"생산 RPA를 먼저 실행하세요: {preview}{suffix}"
            )

    if raw_path.exists():
        wb = openpyxl.load_workbook(raw_path)
    else:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

    total_new = total_updated = 0
    for sheet_name, by_date in records.items():
        if not by_date:
            continue
        ws = _ensure_raw_sheet(wb, sheet_name)

        # 기존 날짜 행 인덱싱
        row_of_date: dict[str, int] = {}
        last_row = 1
        for row_idx in range(2, ws.max_row + 1):
            key = _date_key(ws.cell(row=row_idx, column=1).value)
            if key:
                row_of_date[key] = row_idx
                last_row = row_idx

        new_cnt = upd_cnt = 0
        touched_rows: set[int] = set()
        for day, values in sorted(by_date.items()):
            key = _date_key(day)
            row_idx = row_of_date.get(key)
            if row_idx is None:
                last_row += 1
                row_idx = last_row
                date_cell = ws.cell(row=row_idx, column=1, value=day)
                date_cell.number_format = "YYYY-MM-DD"
                row_of_date[key] = row_idx
                new_cnt += 1
            else:
                upd_cnt += 1
            touched_rows.add(row_idx)
            for col_idx, field_key in enumerate(RAW_COLUMN_KEYS, start=2):
                value = values.get(field_key)
                if value is not None:
                    ws.cell(row=row_idx, column=col_idx, value=value)

            if production_actuals is not None:
                factory_code = FACTORY_KR_TO_CODE[sheet_name]
                ws.cell(row_idx, _raw_column("mix_prod_kg")).value = (
                    production_actuals[(factory_code, key)]
                )

        if ensure_formulas:
            formula_cnt = _ensure_unit_formulas(ws, touched_rows)
            if formula_cnt:
                log.info(
                    f"  [에너지/{sheet_name}] 원단위 수식 {formula_cnt}셀 자동 보완"
                )

        total_new += new_cnt
        total_updated += upd_cnt
        log.info(f"  [에너지/{sheet_name}] 신규 {new_cnt}일 / 갱신 {upd_cnt}일")

    # 시트 순서를 공장 순서로 정렬 (수집 순서와 무관하게 항상 동일)
    order = {name: i for i, name in enumerate(FACTORY_SHEETS)}
    wb._sheets.sort(key=lambda s: order.get(s.title, len(order)))

    # Excel 또는 BEMS가 파일을 열 때도 전체 재계산하도록 표시한다. openpyxl은
    # 수식 계산값 캐시를 만들지 못하므로 운영 실행에서는 저장 후 Excel COM도 호출한다.
    wb.calculation.calcMode = "auto"
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcOnSave = True

    try:
        atomic_save_workbook(wb, str(raw_path))
    finally:
        wb.close()
    if recalculate and not _recalculate_with_excel(raw_path):
        raise RuntimeError(
            "에너지 파일 저장은 완료됐지만 Excel 수식 재계산에 실패했습니다. "
            "파일 잠금과 Excel 설치 상태를 확인하세요."
        )
    log.info(f"에너지 파일 저장: {raw_path}  (신규 {total_new} / 갱신 {total_updated})")
    return raw_path


def write_collected_raw(
    records: dict[str, dict[date, dict[str, float]]],
    collection_path: Path | None = None,
) -> Path:
    """MIS 수집 항목만 별도 원본 파일에 즉시 영속 저장한다.

    생산량 동기화·원단위 수식·Excel COM 재계산은 하지 않는다. 따라서 생산실적
    가공이 실패하거나 프로세스가 종료돼도 좌표 수집 결과는 다시 사용할 수 있다.
    """
    filtered: dict[str, dict[date, dict[str, float]]] = {}
    for sheet_name, by_date in records.items():
        for day, values in by_date.items():
            collected_values = {
                key: value
                for key, value in values.items()
                if key in COLLECTED_FIELD_KEYS
            }
            if collected_values:
                filtered.setdefault(sheet_name, {})[day] = collected_values

    if not filtered:
        raise ValueError("영속 저장할 MIS 유틸리티 수집 항목이 없습니다.")

    return write_raw(
        filtered,
        Path(collection_path or DEFAULT_RAW_PATH),
        sync_production=False,
        recalculate=False,
        ensure_formulas=False,
    )


def read_collected_raw(
    collection_path: Path | None = None, *,
    date_from: date | None = None,
    date_to: date | None = None,
    sheet_names: Sequence[str] | None = None,
) -> dict[str, dict[date, dict[str, float]]]:
    """영속 수집 파일에서 지정 기간·사업장의 MIS 원본 항목만 읽는다."""
    path = Path(collection_path or DEFAULT_RAW_PATH)
    if not path.exists():
        return {}

    existing = read_raw(path)
    targets = set(sheet_names) if sheet_names else None
    filtered: dict[str, dict[date, dict[str, float]]] = {}
    for sheet_name, by_date in existing.items():
        if targets is not None and sheet_name not in targets:
            continue
        for day, values in by_date.items():
            if date_from and day < date_from:
                continue
            if date_to and day > date_to:
                continue
            collected_values = {
                key: value
                for key, value in values.items()
                if key in COLLECTED_FIELD_KEYS
            }
            if collected_values:
                filtered.setdefault(sheet_name, {})[day] = collected_values
    return filtered


def process_collected_raw(
    collection_path: Path | None = None,
    raw_path: Path | None = None, *,
    date_from: date | None = None,
    date_to: date | None = None,
    sheet_names: Sequence[str] | None = None,
    production_path: Path | None = None,
    production_raw_path: Path | None = None,
    wip_path: Path | None = None,
    recalculate: bool = True,
    dry_run: bool = False,
) -> tuple[int, dict[str, dict[str, int]], list[tuple[str, date, float | None, float]]]:
    """영속 수집 원본을 웹 입력 파일에 반영하고 생산량·원단위를 가공한다."""
    target_path = Path(raw_path or DEFAULT_OUTPUT_PATH)
    records = read_collected_raw(
        collection_path,
        date_from=date_from,
        date_to=date_to,
        sheet_names=sheet_names,
    )
    collected_days = sum(len(by_date) for by_date in records.values())

    if dry_run:
        if records:
            actuals = merge_production_actuals(
                production_path, production_raw_path, wip_path
            )
            missing = []
            for sheet_name, by_date in records.items():
                factory_code = FACTORY_KR_TO_CODE.get(sheet_name)
                for day in by_date:
                    key = _date_key(day)
                    if factory_code is None or (factory_code, key) not in actuals:
                        missing.append(f"{sheet_name}/{key}")
            if missing:
                preview = ", ".join(missing[:10])
                suffix = "" if len(missing) <= 10 else f" 외 {len(missing) - 10}건"
                raise ValueError(
                    "생산실적이 없어 에너지 원단위를 만들 수 없습니다: "
                    f"{preview}{suffix}"
                )
        if not target_path.exists():
            return collected_days, {}, []
        stats, changes = resync_production(
            target_path,
            date_from=date_from,
            date_to=date_to,
            sheet_names=sheet_names,
            production_path=production_path,
            production_raw_path=production_raw_path,
            wip_path=wip_path,
            recalculate=False,
            dry_run=True,
        )
        return collected_days, stats, changes

    if records:
        write_raw(
            records,
            target_path,
            production_path=production_path,
            production_raw_path=production_raw_path,
            wip_path=wip_path,
            sync_production=True,
            recalculate=recalculate,
            ensure_formulas=True,
        )

    if not target_path.exists():
        raise FileNotFoundError(
            "에너지 수집 원본과 기존 가공 파일이 모두 없습니다: "
            f"{collection_path or DEFAULT_RAW_PATH} / {target_path}"
        )

    stats, changes = resync_production(
        target_path,
        date_from=date_from,
        date_to=date_to,
        sheet_names=sheet_names,
        production_path=production_path,
        production_raw_path=production_raw_path,
        wip_path=wip_path,
        recalculate=recalculate,
        dry_run=False,
    )
    return collected_days, stats, changes


def read_raw(raw_path: Path | None = None) -> dict[str, dict[date, dict[str, float]]]:
    """에너지 파일을 { 시트명: { date: { field_key: value } } } 로 읽는다."""
    raw_path = Path(raw_path or DEFAULT_OUTPUT_PATH)
    if not raw_path.exists():
        raise FileNotFoundError(f"에너지 파일이 없습니다: {raw_path}")

    wb = openpyxl.load_workbook(raw_path, read_only=True, data_only=True)
    try:
        out: dict[str, dict[date, dict[str, float]]] = {}
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = ws.iter_rows(values_only=True)
            try:
                header = next(rows)
            except StopIteration:
                continue

            # 머리글 → field_key 매핑 (열 순서가 바뀌어도 라벨로 찾는다)
            label_to_key = {FIELD_BY_KEY[k].label: k for k in RAW_COLUMN_KEYS}
            col_keys: list[str | None] = []
            for cell in header:
                col_keys.append(label_to_key.get(str(cell).strip()) if cell else None)

            by_date: dict[date, dict[str, float]] = {}
            for row in rows:
                if not row or row[0] is None:
                    continue
                day = row[0]
                if isinstance(day, datetime):
                    day = day.date()
                if not isinstance(day, date):
                    continue
                values = {
                    key: row[i]
                    for i, key in enumerate(col_keys)
                    if key and i < len(row) and row[i] is not None
                }
                if values:
                    by_date[day] = values
            if by_date:
                out[sheet_name] = by_date
        return out
    finally:
        wb.close()
