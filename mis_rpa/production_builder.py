# 이 파일은 사내 DW(Data Warehouse) 화면에서 복사·붙여넣기한 월별 생산실적
# 엑셀들을 한 폴더에서 일괄 파싱하여 단일 tidy 데이터셋으로 통합합니다.
#
# == 분류 모델 (2개의 독립 차원) ==
#   category1 = 보관유형 (냉동/냉장/상온)
#   category2 = 제품유형 (IC/MY/FM/SN)
#   * 두 차원은 독립 — 예: 김해의 멸균 유음료는 (category1=상온, category2=MY)
#
# == 입력 폴더 구조 ==
#   {src_folder}/
#     ├── 남양주_냉동.xlsx                # category2 자동 추론 (냉동→IC)
#     ├── 남양주_상온.xlsx                # 자동 추론 (상온→SN 기본값)
#     ├── 김해_냉동.xlsx
#     ├── 김해_냉장.xlsx                # category2 는 Item 명 키워드로 분류 (FM/MY)
#     ├── 김해_상온_MY.xlsx             # 파일명에 category2 명시 (멸균 유음료 → MY)
#     ├── 광주_냉장_발효유.xlsx          # 파일명 명시 (모두 FM)
#     ├── 논산_냉장_유음료.xlsx          # 파일명 명시 (모두 MY)
#     └── ...
#
# 파일명 규칙:  {factory}_{category1}[_{category2}].xlsx
#   factory   ∈ {남양주, 김해, 광주, 논산}
#   category1 ∈ {냉동, 냉장, 상온}
#   category2 ∈ {IC, MY, FM, SN, 또는 한글 별칭(아이스크림/유음료/발효유/스낵)}
#             옵션 — 비우면 (factory, category1) 기본 규칙으로 추론
#
# == 시트 구조 (DW 화면 그대로 paste) ==
#   시트명: YYYY-MM (예: 2024-01)
#   컬럼:   Item Code | Item 명 | 물품대 | 누계 계획 | 누계 실적 | 누계 진척률 |
#           01일 | 02일 | ... | 31일
#   * 빈 시트 / "소계·합계" 행은 자동 skip
#
# == 출력 스키마 (DB tidy long) ==
#   date | item_code | item_name | factory | category1 | category2 |
#   planned_qty | actual_qty
#     - actual_qty   : 일별 실적 (01일, 02일, ...)
#     - planned_qty  : 월간 누계 계획 (해당 월의 모든 일자에 동일 값으로 반복)
#                      → 월단위 집계 시 .first() 또는 .max() 로 1회만 취함
from __future__ import annotations

import logging
import os
import re
import tempfile
import zipfile
from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import Iterable

import pandas as pd

from mis_rpa.config import sampled_db_path
from mis_rpa.factories import (
    FACTORY_CODE_TO_KR as DOMAIN_FACTORY_CODE_TO_KR,
    FACTORY_KR_TO_CODE as DOMAIN_FACTORY_KR_TO_CODE,
    FACTORY_PHYSICAL_DISPLAY_ORDER,
    NAMYANGJU_F10A_CODE,
    NAMYANGJU_F10B_CODE,
    NAMYANGJU_PARENT_CODE,
    PRODUCTION_PARENT_FACTORY_CODES,
)

logger = logging.getLogger(__name__)


# ── 상수 ───────────────────────────────────────────────────────
# 신구조(재공품과 통일): Raw 1개 파일(공장×보관유형 시트) → DB 파일
#   (품목군별 wide[= RawDB 와 동일한 공장×보관유형(×제품유형) 시트] + 제품마스터 + 계획 + daily)
# 회사는 "공장 전체"가 아니라 "품목군" 단위로 실적을 관리한다 → DB wide 시트도 RawDB 시트와 동일.
# (품목군 정의/배정은 아래 ProductGroup / load_product_groups / assign_product_group 참고)
DEFAULT_RAW_PATH = sampled_db_path("RawDB_생산실적.xlsx", "PRODUCTION_RAW_XLSX")
DEFAULT_OUTPUT_PATH = sampled_db_path("DB_생산실적.xlsx", "PRODUCTION_DW_XLSX")

# 구버전 폴더 기반 통합용 (legacy: build_dataset(src_folder=...) / migrate_f10_legacy.py 호환)
DEFAULT_SRC_FOLDER = sampled_db_path("Raw_생산실적", "PRODUCTION_RAW_DIR")


def _is_valid_xlsx(path: Path) -> bool:
    """Return True only for a complete xlsx zip package."""
    if not path.exists() or path.stat().st_size <= 0:
        return False
    try:
        with zipfile.ZipFile(path) as zf:
            names = set(zf.namelist())
            if "[Content_Types].xml" not in names or "xl/workbook.xml" not in names:
                return False
            return zf.testzip() is None
    except Exception:
        return False


def _latest_valid_backup(output_path: Path) -> Path | None:
    pattern = f"{output_path.stem}.bak.*{output_path.suffix}"
    backups = sorted(
        output_path.parent.glob(pattern),
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    for backup in backups:
        if _is_valid_xlsx(backup):
            return backup
    return None


def _existing_db_source(output_path: Path) -> Path:
    """Pick a readable DB workbook, falling back to the latest valid backup."""
    if not output_path.exists() or _is_valid_xlsx(output_path):
        return output_path

    backup = _latest_valid_backup(output_path)
    if backup is not None:
        logger.warning(
            f"기존 DB 파일 손상 감지: {output_path}. 최신 정상 백업으로 이력 병합: {backup}"
        )
        return backup

    logger.warning(
        f"기존 DB 파일 손상 감지: {output_path}. 정상 백업이 없어 기존 이력 없이 재생성합니다."
    )
    return output_path


def _write_xlsx_atomically(out_path: Path, write_func) -> Path:
    """Write a complete workbook to a temp file, then replace the target."""
    out_path.parent.mkdir(parents=True, exist_ok=True)
    tmp = tempfile.NamedTemporaryFile(
        prefix=f".{out_path.stem}.",
        suffix=out_path.suffix,
        dir=out_path.parent,
        delete=False,
    )
    tmp_path = Path(tmp.name)
    tmp.close()

    try:
        write_func(tmp_path)
        if not _is_valid_xlsx(tmp_path):
            raise ValueError(f"생성된 임시 xlsx 파일이 유효하지 않습니다: {tmp_path}")
        os.replace(tmp_path, out_path)
    except PermissionError as exc:
        raise PermissionError(
            f"{out_path} 교체 실패. Excel에서 파일을 열어둔 경우 닫고 다시 실행하세요. "
            f"완성된 임시 파일: {tmp_path}"
        ) from exc
    except Exception:
        try:
            tmp_path.unlink(missing_ok=True)
        except Exception:
            logger.warning(f"임시 xlsx 파일 삭제 실패: {tmp_path}", exc_info=True)
        raise

    return out_path

VALID_FACTORIES = set(PRODUCTION_PARENT_FACTORY_CODES)
VALID_CATEGORY1 = {"냉동", "냉장", "상온"}  # 보관유형

# ── 공장 코드 ↔ 한글명 매핑 ──
# DB(production_daily)·parse_sheet·daily 시트는 F-코드를 쓴다. 이 매핑은 '계획' 시트의
# 공장 컬럼(남양주1/2 등 한글명)과 plan_lookup 키 구성에 사용한다.
# (DB wide 데이터 시트는 더 이상 한글 공장명이 아니라 품목군 F-코드 시트로 저장된다 — ProductGroup 참고)
FACTORY_CODE_TO_KR: dict[str, str] = dict(DOMAIN_FACTORY_CODE_TO_KR)
FACTORY_KR_TO_CODE: dict[str, str] = dict(DOMAIN_FACTORY_KR_TO_CODE)
# 공장 표시 순서 참고용(한글명). DB wide 시트는 품목군(F-코드)로 저장되며 시트 순서는
# load_product_groups(RawDB) 정의 순서를 따른다.
FACTORY_SHEET_ORDER: list[str] = list(FACTORY_PHYSICAL_DISPLAY_ORDER)

# ── DB 파일(DB_생산실적.xlsx) 시트/컬럼 규약 ──
MASTER_SHEET = "제품마스터"          # 품목코드 → 제품명 + 보관유형 + 제품유형
PLAN_SHEET = "계획"                   # 연월 × 공장 × 품목코드 → 계획량(월 누계)
DAILY_SHEET = "daily"                 # 내부 호환용 tidy long (production_dw_sync_service 가 읽음)
WIDE_DATE_COL = "날짜"                # 공장별 wide 시트의 날짜 컬럼명 (DB_재공품.xlsx 와 동일)
MASTER_COLUMNS = ["품목코드", "제품명", "보관유형", "제품유형"]
PLAN_COLUMNS = ["연월", "공장", "품목코드", "계획량"]

# DB 파일에서 공장 시트가 아닌(스캔 제외 대상) 시트 이름들
NON_FACTORY_SHEETS = {MASTER_SHEET, PLAN_SHEET, DAILY_SHEET}

# ── Raw 파일(RawDB_생산실적.xlsx) 시트 1행 기간 마커 ──
# DW wide 그리드(01일~31일)에는 年月이 없어 시트명(F10_냉동)만으로는 月을 알 수 없다.
# RPA 가 paste 시 1행에 ["__PERIOD__", start(YYYY-MM-DD), end(YYYY-MM-DD)] 를 기입하고,
# 그리드 헤더는 2행부터 시작한다. 빌드는 이 마커로 (year, month, 업데이트 윈도우)를 얻는다.
PERIOD_MARKER = "__PERIOD__"

# ── 남양주(F10) F10A/F10B 자동 분리 룰 ──────────────────────────
# 사내 DW 추출 시점에는 남양주 1·2공장이 F10 통합으로 추출되어 (냉동/냉장)만 구분됨.
# 실제 운영은 두 공장이 별도 라인이며 다음 분담:
#   - 남양주1 (F10A): "냉장 + MY(유음료)" 만 전량 생산
#   - 남양주2 (F10B): 그 외 전부 (냉장 non-MY = FM/기타, 냉동, 상온)
# 본 모듈은 parse_sheet 단계에서 행마다 (factory, category1, category2)를 보고
# F10 → F10A/F10B 로 자동 재분류합니다. → consolidate / DB sync 모두 자동 적용.
def resolve_namyangju_factory(factory: str, category1: str, category2: str | None) -> str:
    """남양주(F10) 통합 코드를 (category1, category2) 룰로 F10A/F10B 로 분리.

    그 외 공장 코드는 변경 없이 반환.
    """
    if str(factory).upper() == NAMYANGJU_PARENT_CODE:
        if category1 == "냉장" and category2 == "MY":
            return NAMYANGJU_F10A_CODE
        return NAMYANGJU_F10B_CODE
    return factory

# ── category2 (제품유형) 코드 체계 ──
#   IC = Ice Cream      (보통 냉동)
#   MY = Milk & Yogurt  (냉장 또는 상온 - 멸균 유음료)
#   FM = Fermented Milk (냉장)
#   SN = Snack          (상온)
# * category1 과 category2 는 독립 차원. 예: (상온, MY) = 멸균 유음료
CATEGORY2_CODES = {
    "아이스크림": "IC",
    "유음료":     "MY",
    "발효유":     "FM",
    "스낵":       "SN",
}
CATEGORY2_LABELS = {v: k for k, v in CATEGORY2_CODES.items()}  # 역매핑 (IC→아이스크림)
VALID_CATEGORY2 = set(CATEGORY2_CODES.values())  # {"IC","MY","FM","SN"}

# 하위 호환 alias (이전 명칭으로 import 한 외부 코드 보호)
SUB_CATEGORY_CODES = CATEGORY2_CODES
SUB_CATEGORY_LABELS = CATEGORY2_LABELS
VALID_SUB_CATEGORIES = VALID_CATEGORY2

# 냉장 카테고리1 의 category2 추론 키워드 (Item 명 기반).
# 냉동/상온은 (factory, category1) 기본 규칙 또는 파일명 override 로 결정 (키워드 불필요).
_CATEGORY2_KEYWORDS: dict[str, list[str]] = {
    "FM": [  # 발효유
        "요플레", "요거트", "닥터캡슐", "발효유", "요거",
        "마시는요", "마시는 요", "큐티", "바이오플레",
    ],
    "MY": [  # 유음료
        "바나나", "바나나우유", "딸기우유", "초코우유", "우유", "쥬스", "주스", "음료", "두유", "라떼",
        "카페", "라이트", "저지방", "쥬시쿨", "타임", "춘식이", "990", "아이스티", "가공", "빙그레생크림",
        "프렌즈", "더단백", "아연", "에이드", "굿모닝",
    ],
}
# 호환 alias
_SUB_CATEGORY_KEYWORDS = _CATEGORY2_KEYWORDS

OUTPUT_COLUMNS = [
    "date",
    "item_code",
    "item_name",
    "factory",
    "category1",
    "category2",
    "planned_qty",
    "actual_qty",
]


# ── 데이터 클래스 ──────────────────────────────────────────────
@dataclass(frozen=True)
class FileMeta:
    """파일명에서 추출한 메타정보."""
    path: Path
    factory: str
    category1: str          # 보관유형 (냉동/냉장/상온)
    category2: str | None   # 제품유형 (IC/MY/FM/SN). None = Item 명으로 자동 추론


# ── 파일명 파서 ────────────────────────────────────────────────
def parse_filename(path: Path, strict: bool = False) -> FileMeta | None:
    """파일명에서 (factory, category1, category2)를 추출.

    예)
      광주_냉장.xlsx           → (광주, 냉장, None)
      광주_냉장_유음료.xlsx     → (광주, 냉장, FM)  ※ 한글 → 코드 변환
      김해_상온_MY.xlsx        → (김해, 상온, MY)  ※ 멸균 유음료 (독립 차원)
      남양주_상온.xlsx           → (남양주, 상온, None)

    strict=False (기본): factory가 'F\\d+' 형식이면 모두 허용,
                       카테고리명도 화이트리스트 아닌 값(예: '빙장') 허용.
    strict=True : VALID_FACTORIES / VALID_CATEGORY1 만 허용.
    """
    stem = path.stem.strip()
    if stem.startswith("~$") or stem.startswith("_") or stem.startswith("."):
        return None  # Excel 임시 잠금 파일 / 숨김 파일

    parts = stem.split("_")
    if len(parts) < 2:
        return None

    factory = parts[0].upper()
    cat1 = parts[1]
    cat2 = parts[2] if len(parts) >= 3 else None

    # factory 검증 (느슨: F + 숫자)
    if not re.match(r"^F\d{2,3}[A-Z]?$", factory):
        logger.debug(f"파일명 factory 형식 불일치 skip: {path.name}")
        return None
    if strict and factory not in VALID_FACTORIES:
        logger.warning(f"파일명 factory 화이트리스트 외: {path.name} → {factory}")
        return None
    if not strict and factory not in VALID_FACTORIES:
        logger.info(f"비표준 factory 허용: {factory} ({path.name})")

    if strict and cat1 not in VALID_CATEGORY1:
        logger.warning(f"파일명 category1 화이트리스트 외: {path.name} → {cat1}")
        return None
    if not strict and cat1 not in VALID_CATEGORY1:
        logger.info(f"비표준 category1 허용: {cat1} ({path.name})")

    # 한글 category2(예: '유음료', '발효유', '아이스크림', '스낵')는 코드로 변환
    if cat2 is not None and cat2 in CATEGORY2_CODES:
        cat2 = CATEGORY2_CODES[cat2]

    if cat2 is not None and cat2 not in VALID_CATEGORY2:
        if strict:
            logger.warning(f"파일명 category2 미인식 → 자동추론으로 대체: {path.name}")
            cat2 = None
        else:
            logger.info(f"비표준 category2 허용: {cat2} ({path.name})")

    return FileMeta(path=path, factory=factory, category1=cat1, category2=cat2)


# ── 시트명에서 (year, month) 추출 ──────────────────────────────
def parse_sheet_name(sheet_name: str) -> tuple[int, int] | None:
    """시트명에서 (year, month) 추출.

    허용 형식: 2024-01, 2024_01, 202401, 2024.01, 2024-1, 24-01
    """
    s = str(sheet_name).strip()
    patterns = [
        r"^(20\d{2})[-_.]?(\d{1,2})$",
        r"^(\d{2})[-_.](\d{1,2})$",
    ]
    for p in patterns:
        m = re.match(p, s)
        if m:
            y = int(m.group(1))
            mo = int(m.group(2))
            if y < 100:
                y += 2000
            if 1 <= mo <= 12 and 2000 <= y <= 2100:
                return y, mo
    return None


# ── 컬럼명 정규화 ──────────────────────────────────────────────
def _normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    """DW의 다양한 컬럼명 변형을 표준 키로 통일."""
    out = df.copy()
    rename_map: dict[str, str] = {}
    for c in out.columns:
        s = str(c).strip()
        if re.match(r"item\s*code", s, re.IGNORECASE) or "ItemCode" in s.replace(" ", ""):
            rename_map[c] = "item_code"
        elif re.match(r"item\s*명", s, re.IGNORECASE) or "품명" in s or "품목명" in s:
            rename_map[c] = "item_name"
        elif "물품대" in s or "단가" in s:
            rename_map[c] = "unit_price"
        elif "누계" in s and "계획" in s:
            rename_map[c] = "planned_cum"
        elif "누계" in s and ("실적" in s or "달성" in s):
            rename_map[c] = "actual_cum"
        elif "누계" in s and ("진척" in s or "달성률" in s):
            rename_map[c] = "progress_pct"
    return out.rename(columns=rename_map)


# ── 일자 컬럼 식별 ─────────────────────────────────────────────
def _identify_daily_columns(df: pd.DataFrame) -> list[tuple[int, str]]:
    """'01일', '1일', '1' 형태 컬럼을 (day, original_col_name) 리스트로 반환."""
    daily_cols: list[tuple[int, str]] = []
    for c in df.columns:
        s = str(c).strip()
        m = re.match(r"^(\d{1,2})\s*일?$", s)
        if m:
            day = int(m.group(1))
            if 1 <= day <= 31:
                daily_cols.append((day, c))
    daily_cols.sort()
    return daily_cols


# ── Item Code 정규화 ──────────────────────────────────────────
def _clean_item_code(raw: object) -> str:
    """Excel float→str 등 다양한 표현을 안정적으로 정규화."""
    if raw is None:
        return ""
    if isinstance(raw, float):
        if pd.isna(raw):
            return ""
        if raw.is_integer():
            return str(int(raw))
        return str(raw)
    s = str(raw).strip()
    if s.endswith(".0"):
        s = s[:-2]
    return s


# ── category2 자동 추론 ──────────────────────────────────────
def infer_category2(item_name: str, category1: str) -> str | None:
    """category1 + Item 명을 기반으로 category2 코드를 결정합니다.

    규칙 (파일명에 category2 명시되지 않은 경우의 기본값):
      - 냉동 → 'IC' (모든 품목, 키워드 매칭 불필요)
      - 상온 → 'SN' (기본값. 상온이지만 MY/FM 등 다른 제품유형이면 파일명에 명시 필요)
      - 냉장 → Item 명 키워드 매칭으로 'FM'(발효유) 또는 'MY'(유음료) 결정
               매칭 실패 시 None (커버리지 검증으로 모니터링)
      - 그 외(미인식 category1) → None
    """
    # 냉동/상온은 키워드 무관 일괄 부여 (파일명 override가 우선)
    if category1 == "냉동":
        return "IC"
    if category1 == "상온":
        return "SN"
    if category1 != "냉장":
        return None
    if not isinstance(item_name, str) or not item_name.strip():
        return None
    name = item_name.upper()
    # 발효유(FM) 키워드 우선 — '바나나요거트' 같은 하이브리드는 FM 분류
    for code in ("FM", "MY"):
        for kw in _CATEGORY2_KEYWORDS[code]:
            if kw.upper() in name:
                return code
    return None


# 하위 호환 alias (외부에서 옛 함수명으로 import 한 코드 보호)
infer_sub_category = infer_category2


# ── 단일 시트 → tidy DataFrame ────────────────────────────────
def parse_sheet(
    df_raw: pd.DataFrame,
    year: int,
    month: int,
    factory: str,
    category1: str,
    category2_override: str | None = None,
) -> pd.DataFrame:
    """단일 월 시트(wide) → tidy long DataFrame.

    출력 컬럼: OUTPUT_COLUMNS
    """
    df = _normalize_columns(df_raw)
    if "item_code" not in df.columns:
        return pd.DataFrame(columns=OUTPUT_COLUMNS)

    df = df.dropna(subset=["item_code"]).copy()
    df = df[df["item_code"].astype(str).str.strip() != ""]
    # 소계/합계/Total 행 제외
    df = df[
        ~df["item_code"]
        .astype(str)
        .str.contains(r"계|소계|합계|Total", case=False, regex=True, na=False)
    ]

    daily_cols = _identify_daily_columns(df)
    has_planned = "planned_cum" in df.columns

    rows: list[dict] = []
    for _, r in df.iterrows():
        item_code = _clean_item_code(r.get("item_code"))
        if not item_code:
            continue
        item_name = str(r.get("item_name", "")).strip() if "item_name" in df.columns else ""
        cat2 = category2_override or infer_category2(item_name, category1)

        planned_val = 0.0
        if has_planned:
            v = pd.to_numeric(r.get("planned_cum"), errors="coerce")
            planned_val = float(v) if pd.notna(v) else 0.0

        # 남양주(F10) 통합 추출 → F10A/F10B 자동 재분류 (행 단위 — category2 가 행마다 다를 수 있음)
        row_factory = resolve_namyangju_factory(factory, category1, cat2)

        for day, col in daily_cols:
            try:
                d = date(year, month, day)
            except ValueError:
                continue  # 해당 월에 없는 일자 (예: 2월 30일)
            v = pd.to_numeric(r.get(col), errors="coerce")
            actual = float(v) if pd.notna(v) else 0.0
            # 0인 일자도 보존 — 후속 상관관계 분석/가공 시 "미생산일"이 명시적으로 필요.
            rows.append({
                "date": d,
                "item_code": item_code,
                "item_name": item_name,
                "factory": row_factory,
                "category1": category1,
                "category2": cat2,
                "planned_qty": planned_val,
                "actual_qty": actual,
            })

    if not rows:
        return pd.DataFrame(columns=OUTPUT_COLUMNS)
    return pd.DataFrame(rows, columns=OUTPUT_COLUMNS)


# ── 단일 워크북 → tidy DataFrame ──────────────────────────────
def parse_workbook(meta: FileMeta) -> pd.DataFrame:
    """파일 1개 (factory × category)를 모든 월 시트 합쳐서 반환."""
    try:
        all_sheets = pd.read_excel(meta.path, sheet_name=None, engine="openpyxl")
    except Exception as exc:
        logger.error(f"엑셀 로드 실패: {meta.path.name} — {exc}")
        return pd.DataFrame(columns=OUTPUT_COLUMNS)

    parts: list[pd.DataFrame] = []
    for sn, df_raw in all_sheets.items():
        ym = parse_sheet_name(sn)
        if ym is None:
            logger.debug(f"시트명 미인식 skip: {meta.path.name} / {sn}")
            continue
        if df_raw is None or df_raw.empty:
            continue
        y, mo = ym
        df_one = parse_sheet(
            df_raw,
            year=y,
            month=mo,
            factory=meta.factory,
            category1=meta.category1,
            category2_override=meta.category2,
        )
        if not df_one.empty:
            parts.append(df_one)

    if not parts:
        return pd.DataFrame(columns=OUTPUT_COLUMNS)
    return pd.concat(parts, ignore_index=True)


# ── 폴더 전체 통합 ────────────────────────────────────────────
def consolidate_folder(
    src_folder: Path | str | None = None,
) -> pd.DataFrame:
    """폴더 내 모든 워크북을 파싱하여 단일 tidy DataFrame으로 합칩니다.

    Returns
    -------
    DataFrame with columns = OUTPUT_COLUMNS, sorted by (date, factory, item_code)
    """
    src = Path(src_folder) if src_folder else DEFAULT_SRC_FOLDER
    if not src.exists():
        raise FileNotFoundError(f"입력 폴더가 없습니다: {src}")

    files = sorted(src.glob("*.xlsx"))
    if not files:
        raise ValueError(f"폴더에 .xlsx 파일이 없습니다: {src}")

    parts: list[pd.DataFrame] = []
    for f in files:
        meta = parse_filename(f)
        if meta is None:
            logger.info(f"파일명 패턴 불일치 skip: {f.name}")
            continue
        df_one = parse_workbook(meta)
        if df_one.empty:
            logger.warning(f"빈 결과: {f.name}")
            continue
        parts.append(df_one)
        logger.info(f"파싱 완료: {f.name} (rows={len(df_one):,})")

    if not parts:
        return pd.DataFrame(columns=OUTPUT_COLUMNS)

    out = pd.concat(parts, ignore_index=True)
    out = out.sort_values(["date", "factory", "category1", "item_code"]).reset_index(drop=True)
    return out


# ── 월단위 집계 (AI 보고서용 핵심 함수) ───────────────────────
def aggregate_to_monthly(
    df_daily: pd.DataFrame,
    by: Iterable[str] = ("factory", "category1", "category2"),
) -> pd.DataFrame:
    """일별 데이터를 (year, month) + 그룹키로 집계.

    Parameters
    ----------
    by : 추가 그룹 키 (기본: factory + category1 + category2)
         item_code 단위로 집계하려면 by=("factory","category1","category2","item_code","item_name")

    Returns
    -------
    columns: year, month, *by, monthly_actual, monthly_plan, achievement_pct
        - monthly_plan : 그룹 내 first() — 같은 월 같은 그룹은 동일 계획값을 공유한다고 가정
                         (item_code 단위가 아니라면 sum이 더 정확하지만, plan은 item별로 발급되므로
                          item_code를 by에 포함하지 않으면 plan 합산은 부정확. 보수적으로 first 사용.)
    """
    if df_daily.empty:
        return pd.DataFrame()
    df = df_daily.copy()
    df["date"] = pd.to_datetime(df["date"])
    df["year"] = df["date"].dt.year
    df["month"] = df["date"].dt.month

    by_list = list(by)
    group_keys = ["year", "month"] + by_list

    # plan은 item_code 별로 월 1회만 잡혀 있으므로
    # item_code 가 그룹키에 포함될 때만 의미가 있음 → 사용자에게 위임
    if "item_code" in by_list:
        plan_agg = "first"
    else:
        # item_code별 월간 계획을 합산 (같은 그룹 내 모든 item plan의 sum)
        # 단, 같은 (year,month,그룹키) 내에서 item_code 별로 distinct 한 plan 만 더해야 함
        plan_agg = None  # 별도 처리

    if plan_agg == "first":
        out = df.groupby(group_keys, dropna=False).agg(
            monthly_actual=("actual_qty", "sum"),
            monthly_plan=("planned_qty", "first"),
        ).reset_index()
    else:
        # item_code 단위 plan 을 distinct 하게 한 번만 더하기
        plan_unique = (
            df.drop_duplicates(["year", "month", "item_code"] + by_list)
              .groupby(group_keys, dropna=False)["planned_qty"].sum()
              .rename("monthly_plan")
        )
        actual_sum = df.groupby(group_keys, dropna=False)["actual_qty"].sum().rename("monthly_actual")
        out = pd.concat([actual_sum, plan_unique], axis=1).reset_index()

    out["achievement_pct"] = out.apply(
        lambda r: (r["monthly_actual"] / r["monthly_plan"] * 100.0) if r["monthly_plan"] > 0 else 0.0,
        axis=1,
    )
    return out


# ── category2 커버리지 검증 ────────────────────────────────────
def validate_subcategory_coverage(df: pd.DataFrame) -> dict:
    """전체 데이터에서 category2 가 NULL 인 품목이 있는지 검사.

    category2 추론 실패 = Item 명 키워드 매칭 실패. (현재는 냉장 카테고리에서만 추론을 거침)
    그 합계가 곧 "키워드 보강이 필요한 품목들"의 규모.

    Returns
    -------
    dict 키:
      total_actual          : 전체 actual_qty 합
      classified_actual     : category2 NOT NULL 행의 합
      unclassified_actual   : category2 IS NULL 행의 합
      coverage_pct          : classified / total * 100
      unclassified_pct      : unclassified / total * 100
      unclassified_items    : (item_code, item_name) DataFrame (unique, 합계 내림차순)
      is_complete           : True 면 누락 0
    """
    empty_result = {
        "total_actual": 0.0,
        "classified_actual": 0.0,
        "unclassified_actual": 0.0,
        "coverage_pct": 100.0,
        "unclassified_pct": 0.0,
        "unclassified_items": pd.DataFrame(columns=["item_code", "item_name", "actual_qty"]),
        "is_complete": True,
    }
    if df.empty or "category2" not in df.columns:
        return empty_result

    total = float(df["actual_qty"].sum())
    classified = float(df[df["category2"].notna()]["actual_qty"].sum())
    unclassified_df = df[df["category2"].isna()]
    unclassified = float(unclassified_df["actual_qty"].sum())

    items_summary = (
        unclassified_df.groupby(["item_code", "item_name"], dropna=False)["actual_qty"]
        .sum()
        .reset_index()
        .sort_values("actual_qty", ascending=False)
    )

    result = {
        "total_actual": total,
        "classified_actual": classified,
        "unclassified_actual": unclassified,
        "coverage_pct": (classified / total * 100.0) if total > 0 else 100.0,
        "unclassified_pct": (unclassified / total * 100.0) if total > 0 else 0.0,
        "unclassified_items": items_summary,
        "is_complete": len(items_summary) == 0,
    }

    if not result["is_complete"]:
        logger.warning(
            f"category2 미분류 품목 {len(items_summary)}개 / "
            f"미분류 실적 {unclassified:,.0f} ({result['unclassified_pct']:.1f}%) — "
            f"production_dw_service._CATEGORY2_KEYWORDS 보강 필요"
        )

    return result


# ── 출력 저장 ─────────────────────────────────────────────────
def save_consolidated(
    df_daily: pd.DataFrame,
    output_path: Path | str | None = None,
    write_monthly_summary: bool = True,
) -> Path:
    """tidy 일별 데이터 + (옵션) 월별 요약을 하나의 xlsx에 저장.

    시트 구성:
      - daily                 : 일별 tidy long (0인 일자 포함)
      - monthly_factory_cat   : 공장×보관유형×제품유형 월별 합계
      - monthly_factory       : 공장 월별 합계
      - monthly_total         : 전사 월별 합계
      - category2_unclassified: (있을 때만) category2 미분류 품목 목록
    """
    out_path = Path(output_path) if output_path else DEFAULT_OUTPUT_PATH

    def _write(target: Path) -> None:
        with pd.ExcelWriter(target, engine="openpyxl") as writer:
            df_daily.to_excel(writer, sheet_name="daily", index=False)
            if write_monthly_summary and not df_daily.empty:
                mfc = aggregate_to_monthly(
                    df_daily, by=("factory", "category1", "category2")
                )
                mfc.to_excel(writer, sheet_name="monthly_factory_cat", index=False)

                mf = aggregate_to_monthly(df_daily, by=("factory",))
                mf.to_excel(writer, sheet_name="monthly_factory", index=False)

                mt = aggregate_to_monthly(df_daily, by=())
                mt.to_excel(writer, sheet_name="monthly_total", index=False)

                cov = validate_subcategory_coverage(df_daily)
                if not cov["is_complete"]:
                    cov["unclassified_items"].to_excel(
                        writer, sheet_name="category2_unclassified", index=False
                    )

    return _write_xlsx_atomically(out_path, _write)


# ─────────────────────────────────────────────────────────────
# 신구조: 단일 Raw 파일 → DB 파일(공장별 wide + 제품마스터 + 계획 + daily)
# (재공품 RawDB_재공품.xlsx / DB_재공품.xlsx 2-파일 구조와 통일)
# ─────────────────────────────────────────────────────────────

def parse_meta_from_sheet(sheet_name: str) -> FileMeta | None:
    """Raw 파일의 시트명(F10_냉동 등)에서 (factory, category1, category2) 추출.

    parse_filename 의 stem 파싱 규칙을 그대로 재사용. 공장 시트가 아닌
    메타 시트(제품마스터/계획/daily 등)는 factory 정규식에 안 걸려 None.
    """
    return parse_filename(Path(f"{sheet_name}.xlsx"))


# ── 품목군(Product Group) — DB 파일 wide 시트를 RawDB 와 동일하게 구성 ─────────
# 회사는 "공장 전체"가 아니라 "공장 × 보관유형[× 제품유형]" 단위(= RawDB 카테고리 시트)
# 로 실적을 관리한다. 따라서 DB_생산실적.xlsx 의 데이터 시트도 RawDB_생산실적.xlsx
# 의 시트명과 동일하게 맞춘다 (F10_냉동, F10_냉장, F20_상온_MY, F30_냉장_FM ...).
# 남양주 F10A/F10B 분석 구분은 숨김 daily 시트에 그대로 보존된다 (DB sync 무변경).
@dataclass(frozen=True)
class ProductGroup:
    """RawDB 카테고리 시트와 동일한 품목군 정의."""
    sheet_name: str         # 예: "F10_냉장", "F20_상온_MY"
    factory: str            # F-코드(남양주 통합): F10/F20/F30/F40
    category1: str          # 보관유형 (냉동/냉장/상온)
    category2: str | None   # 제품유형 override (None = 해당 (공장,보관유형) 전체 수용)


def _factory_to_parent(factory: str) -> str:
    """daily factory(F10A/F10B/...) → 품목군 시트용 부모 F-코드.

    남양주1/2(F10A/F10B)는 RawDB 와 동일하게 F10 으로 통합한다.
    """
    if str(factory) in (NAMYANGJU_F10A_CODE, NAMYANGJU_F10B_CODE):
        return NAMYANGJU_PARENT_CODE
    return str(factory)


def _groups_from_names(names: Iterable[str]) -> list[ProductGroup]:
    """시트명/파일 stem 목록 → 품목군 정의. 메타시트/패턴불일치는 제외."""
    groups: list[ProductGroup] = []
    for sn in names:
        if sn in NON_FACTORY_SHEETS:
            continue
        meta = parse_meta_from_sheet(sn)
        if meta is None:
            continue
        groups.append(ProductGroup(sn, meta.factory, meta.category1, meta.category2))
    return groups


def load_product_groups(ref_path: Path | str | None = None) -> list[ProductGroup]:
    """참조 파일(RawDB 또는 기존 DB)의 데이터 시트명 → 품목군 정의 목록.

    RawDB_생산실적.xlsx 의 카테고리 시트명이 품목군 구성의 권위 소스다.
    파일이 없거나 읽기 실패 시 빈 목록.
    """
    ref = Path(ref_path) if ref_path else DEFAULT_RAW_PATH
    if not ref.exists():
        return []
    try:
        names = pd.ExcelFile(ref, engine="openpyxl").sheet_names
    except Exception as exc:
        logger.warning(f"품목군 정의 로드 실패({ref}): {exc}")
        return []
    return _groups_from_names(names)


def merge_group_defs(
    primary: list[ProductGroup], secondary: list[ProductGroup]
) -> list[ProductGroup]:
    """품목군 정의 합집합 (sheet_name 기준, primary 우선).

    이번 회차 RawDB(primary) 에 없지만 기존 DB 이력에만 있는 품목군(secondary)도
    보존해, 부분 RawDB 로 빌드해도 과거 시트가 폴백명으로 쪼개지지 않게 한다.
    """
    seen = {g.sheet_name for g in primary}
    out = list(primary)
    for g in secondary:
        if g.sheet_name not in seen:
            out.append(g)
            seen.add(g.sheet_name)
    return out


def assign_product_group(
    factory: str,
    category1: str | None,
    category2: str | None,
    groups: list[ProductGroup],
) -> str | None:
    """daily 행(factory=F10A/.../F40) → 소속 품목군 시트명. 매칭 실패 시 None.

    (공장, 보관유형) 이 같은 후보 중 제품유형 override 가 맞는 시트를 우선,
    없으면 override 없는(= catch-all) 시트로 보낸다.
    """
    fcode = _factory_to_parent(factory)
    cands = [g for g in groups if g.factory == fcode and g.category1 == category1]
    if not cands:
        return None
    for g in cands:
        if g.category2 is not None and g.category2 == category2:
            return g.sheet_name
    for g in cands:
        if g.category2 is None:
            return g.sheet_name
    return None


def _fallback_group_name(
    factory: str, category1: str | None, category2: str | None
) -> str:
    """품목군 정의 밖 조합용 폴백 시트명 (RawDB 부재/이력 잔여 조합 방어)."""
    parts = [_factory_to_parent(factory), str(category1) if category1 else "기타"]
    if category2:
        parts.append(str(category2))
    return "_".join(parts)


def _parse_period_marker(first_row: pd.Series) -> tuple[date, date] | None:
    """Raw 시트 1행 마커 ["__PERIOD__", start, end] → (start_date, end_date)."""
    if first_row.empty:
        return None
    if str(first_row.iloc[0]).strip() != PERIOD_MARKER:
        return None
    try:
        start = pd.to_datetime(first_row.iloc[1]).date()
        end = pd.to_datetime(first_row.iloc[2]).date()
    except Exception:
        return None
    if start > end:
        start, end = end, start
    return start, end


def _read_raw_sheet(raw_path: Path, sheet_name: str) -> tuple[date, date, pd.DataFrame] | None:
    """Raw 파일의 단일 카테고리 시트 → (window_start, window_end, 그리드 DataFrame).

    1행은 기간 마커, 2행이 DW 그리드 헤더, 3행부터 데이터.
    마커가 없으면 ValueError (수동 붙여넣기 시 마커 누락 방지).
    """
    raw = pd.read_excel(raw_path, sheet_name=sheet_name, header=None, engine="openpyxl")
    if raw.empty or raw.shape[0] < 2:
        logger.warning(f"빈 시트 skip: {sheet_name}")
        return None

    period = _parse_period_marker(raw.iloc[0])
    if period is None:
        raise ValueError(
            f"[{sheet_name}] 1행에 기간 마커({PERIOD_MARKER})가 없습니다. "
            f"RPA(production_daily_rpa)가 자동 기입합니다. 수동 붙여넣기 시 "
            f"1행에 ['{PERIOD_MARKER}', 시작일, 종료일]을 입력하세요."
        )
    start, end = period

    grid = raw.iloc[1:].reset_index(drop=True)
    grid.columns = grid.iloc[0]
    grid = grid.iloc[1:].reset_index(drop=True)
    if grid.empty:
        logger.warning(f"그리드 데이터 없음 skip: {sheet_name}")
        return None
    return start, end, grid


def consolidate_raw_file(raw_path: Path | str | None = None) -> pd.DataFrame:
    """Raw 파일의 모든 카테고리 시트를 파싱하여 tidy long(F-코드) DataFrame.

    각 시트는 최신 조회 月의 그리드만 담는다. parse_sheet 로 (F10 분리·category2
    추론·planned 포함) tidy long 생성 후, 마커 기간 [start, end] 로 클립한다.
    """
    raw = Path(raw_path) if raw_path else DEFAULT_RAW_PATH
    if not raw.exists():
        raise FileNotFoundError(f"Raw 파일이 없습니다: {raw}")

    xl_sheets = pd.ExcelFile(raw, engine="openpyxl").sheet_names
    parts: list[pd.DataFrame] = []

    for sn in xl_sheets:
        if sn in NON_FACTORY_SHEETS:
            continue
        meta = parse_meta_from_sheet(sn)
        if meta is None:
            logger.info(f"시트명 패턴 불일치 skip: {sn}")
            continue

        read = _read_raw_sheet(raw, sn)
        if read is None:
            continue
        start, end, grid = read

        df_one = parse_sheet(
            grid,
            year=start.year,
            month=start.month,
            factory=meta.factory,
            category1=meta.category1,
            category2_override=meta.category2,
        )
        if df_one.empty:
            logger.warning(f"빈 파싱 결과: {sn}")
            continue

        # 마커 기간 [start, end] 로 클립 (D-2 이후 미래 일자의 0 은 이력에 쓰지 않음)
        df_one = df_one[(df_one["date"] >= start) & (df_one["date"] <= end)]
        if df_one.empty:
            continue
        parts.append(df_one)
        logger.info(f"파싱 완료: {sn} (rows={len(df_one):,}, {start}~{end})")

    if not parts:
        return pd.DataFrame(columns=OUTPUT_COLUMNS)
    out = pd.concat(parts, ignore_index=True)
    return out.sort_values(["date", "factory", "item_code"]).reset_index(drop=True)


# ── 기존 DB 파일 로드 (공장별 wide / 제품마스터 / 계획) ────────
def _load_existing_factory_long(output_path: Path) -> dict[str, pd.DataFrame]:
    """기존 DB 파일의 품목군 wide 시트 → {F-코드: long(date, item_code, actual_qty)}.

    품목군 시트(F10_냉동 등)를 melt 한 뒤 시트명 메타와 제품마스터로 daily factory
    코드(F10A/F10B/F20/...)를 재구성한다. 남양주 냉장 시트(F10A·F10B 혼재)만 제품유형
    조회로 행별 재분리하고, 그 외 시트는 시트명의 F-코드를 그대로 쓴다.
    NaN 셀(미생산/미등장)은 제외 — 0 인 일자는 명시 0 으로 보존된다.
    """
    result: dict[str, pd.DataFrame] = {}
    if not output_path.exists():
        return result
    try:
        existing_sheets = pd.ExcelFile(output_path, engine="openpyxl").sheet_names
    except Exception:
        return result

    # 남양주 냉장 시트(F10A/F10B 혼재)의 행별 재분리를 위해 제품유형(category2) 조회
    master = _load_existing_master(output_path)

    parts_by_factory: dict[str, list[pd.DataFrame]] = {}
    for sn in existing_sheets:
        if sn in NON_FACTORY_SHEETS:
            continue
        meta = parse_meta_from_sheet(sn)
        if meta is None:
            continue
        wide = pd.read_excel(output_path, sheet_name=sn, engine="openpyxl")
        if wide.empty or WIDE_DATE_COL not in wide.columns:
            continue
        wide[WIDE_DATE_COL] = pd.to_datetime(wide[WIDE_DATE_COL], errors="coerce").dt.date
        wide = wide.dropna(subset=[WIDE_DATE_COL])
        long = wide.melt(
            id_vars=[WIDE_DATE_COL], var_name="item_code", value_name="actual_qty"
        ).rename(columns={WIDE_DATE_COL: "date"})
        long["item_code"] = long["item_code"].map(_clean_item_code)
        long["actual_qty"] = pd.to_numeric(long["actual_qty"], errors="coerce")
        long = long.dropna(subset=["actual_qty"])  # NaN(미등장) 제외, 0 은 보존
        if long.empty:
            continue

        # 시트 → daily factory 코드 재구성
        if meta.factory != NAMYANGJU_PARENT_CODE:
            long["__factory"] = meta.factory
        else:
            def _resolve(code: str) -> str:
                c2 = meta.category2
                if c2 is None:  # F10_냉장 등 catch-all → 품목 제품유형으로 분기
                    info = master.get(code)
                    c2 = info["category2"] if info else None
                return resolve_namyangju_factory(
                    NAMYANGJU_PARENT_CODE, meta.category1, c2
                )
            long["__factory"] = long["item_code"].map(_resolve)

        for fcode, sub in long.groupby("__factory"):
            parts_by_factory.setdefault(fcode, []).append(
                sub[["date", "item_code", "actual_qty"]]
            )

    for fcode, parts in parts_by_factory.items():
        combined = pd.concat(parts, ignore_index=True)
        # 같은 (date,item) 중복(여러 시트 기여) 합산 방어 — 정상 데이터는 영향 없음
        combined = combined.groupby(
            ["date", "item_code"], as_index=False
        )["actual_qty"].sum()
        result[fcode] = combined.sort_values(["date", "item_code"]).reset_index(drop=True)
    if not result and DAILY_SHEET in existing_sheets:
        daily_result = _load_existing_daily_long(output_path)
        if daily_result:
            logger.info(f"기존 daily 시트에서 생산 이력 로드: {output_path}")
            return daily_result
    return result


def _load_existing_daily_long(output_path: Path) -> dict[str, pd.DataFrame]:
    """기존 daily 시트 → {factory: long(date, item_code, actual_qty)}.

    구구조 백업은 품목군 wide 시트가 없을 수 있으므로 daily 시트를 이력 소스로 사용한다.
    """
    result: dict[str, pd.DataFrame] = {}
    if not output_path.exists():
        return result
    try:
        sheets = pd.ExcelFile(output_path, engine="openpyxl").sheet_names
    except Exception as exc:
        logger.warning(f"기존 daily 시트 목록 로드 실패({output_path}): {exc}")
        return result
    if DAILY_SHEET not in sheets:
        return result
    try:
        df = pd.read_excel(output_path, sheet_name=DAILY_SHEET, engine="openpyxl")
    except Exception as exc:
        logger.warning(f"기존 daily 시트 로드 실패({output_path}): {exc}")
        return result

    required = {"date", "factory", "item_code", "actual_qty"}
    if not required.issubset(df.columns):
        logger.warning(f"기존 daily 시트 필수 컬럼 누락({output_path}): {required - set(df.columns)}")
        return result

    daily = df[["date", "factory", "item_code", "actual_qty"]].copy()
    daily["date"] = pd.to_datetime(daily["date"], errors="coerce").dt.date
    daily["factory"] = daily["factory"].astype(str).str.strip()
    daily["item_code"] = daily["item_code"].map(_clean_item_code)
    daily["actual_qty"] = pd.to_numeric(daily["actual_qty"], errors="coerce")
    daily = daily.dropna(subset=["date", "factory", "item_code", "actual_qty"])
    daily = daily[daily["factory"] != ""]
    daily = daily[daily["item_code"] != ""]

    for fcode, sub in daily.groupby("factory"):
        combined = sub.groupby(
            ["date", "item_code"], as_index=False
        )["actual_qty"].sum()
        result[str(fcode)] = combined.sort_values(["date", "item_code"]).reset_index(drop=True)
    return result


def _load_existing_master(output_path: Path) -> dict[str, dict]:
    """기존 제품마스터 시트 → {item_code: {name, category1, category2}}."""
    out: dict[str, dict] = {}
    if not output_path.exists():
        return out
    try:
        sheets = pd.ExcelFile(output_path, engine="openpyxl").sheet_names
    except Exception as exc:
        logger.warning(f"기존 제품마스터 로드 실패({output_path}): {exc}")
        return out
    if MASTER_SHEET not in sheets:
        return out
    df = pd.read_excel(output_path, sheet_name=MASTER_SHEET, engine="openpyxl")
    for _, r in df.iterrows():
        code = _clean_item_code(r.get("품목코드"))
        if not code:
            continue
        c2 = r.get("제품유형")
        out[code] = {
            "name": str(r.get("제품명", "") or ""),
            "category1": (str(r.get("보관유형")) if pd.notna(r.get("보관유형")) else None),
            "category2": (str(c2) if pd.notna(c2) and str(c2).strip() else None),
        }
    return out


def _load_existing_plan(output_path: Path) -> pd.DataFrame:
    """기존 계획 시트 → DataFrame[연월, 공장, 품목코드, 계획량]."""
    if not output_path.exists():
        return pd.DataFrame(columns=PLAN_COLUMNS)
    try:
        sheets = pd.ExcelFile(output_path, engine="openpyxl").sheet_names
    except Exception as exc:
        logger.warning(f"기존 계획 로드 실패({output_path}): {exc}")
        return pd.DataFrame(columns=PLAN_COLUMNS)
    if PLAN_SHEET not in sheets:
        return pd.DataFrame(columns=PLAN_COLUMNS)
    df = pd.read_excel(output_path, sheet_name=PLAN_SHEET, engine="openpyxl")
    if df.empty:
        return pd.DataFrame(columns=PLAN_COLUMNS)
    df["연월"] = df["연월"].astype(str).str.strip()
    df["공장"] = df["공장"].astype(str).str.strip()
    df["품목코드"] = df["품목코드"].map(_clean_item_code)
    df["계획량"] = pd.to_numeric(df["계획량"], errors="coerce").fillna(0.0)
    return df[PLAN_COLUMNS]


# ── 병합: 기존 이력 + 신규 月 ────────────────────────────────
def _merge_factory_long(
    existing: dict[str, pd.DataFrame],
    new_df: pd.DataFrame,
) -> dict[str, pd.DataFrame]:
    """공장별 long 병합. 신규 데이터에 등장한 날짜는 교체, 나머지 이력은 보존."""
    merged: dict[str, pd.DataFrame] = {}
    factories = set(existing) | set(new_df["factory"].unique())

    for code in factories:
        old = existing.get(code, pd.DataFrame(columns=["date", "item_code", "actual_qty"]))
        new_fac = new_df[new_df["factory"] == code][["date", "item_code", "actual_qty"]].copy()
        # 같은 (date,item) 중복 합산 (여러 시트가 같은 공장에 기여하는 경우 방어)
        if not new_fac.empty:
            new_fac = (
                new_fac.groupby(["date", "item_code"], as_index=False)["actual_qty"].sum()
            )
            replace_dates = set(new_fac["date"])
            old = old[~old["date"].isin(replace_dates)]
        combined = pd.concat([old, new_fac], ignore_index=True)
        if combined.empty:
            continue
        combined["item_code"] = combined["item_code"].map(_clean_item_code)
        merged[code] = combined.sort_values(["date", "item_code"]).reset_index(drop=True)
    return merged


def _merge_master(existing: dict[str, dict], new_df: pd.DataFrame) -> pd.DataFrame:
    """제품마스터 병합 (신규가 기존을 덮어씀) → DataFrame[MASTER_COLUMNS]."""
    master = dict(existing)
    for _, r in new_df.drop_duplicates("item_code").iterrows():
        code = _clean_item_code(r.get("item_code"))
        if not code:
            continue
        master[code] = {
            "name": str(r.get("item_name", "") or ""),
            "category1": (str(r.get("category1")) if pd.notna(r.get("category1")) else None),
            "category2": (str(r.get("category2")) if pd.notna(r.get("category2")) and str(r.get("category2")).strip() else None),
        }
    rows = [
        {"품목코드": c, "제품명": m["name"], "보관유형": m["category1"], "제품유형": m["category2"]}
        for c, m in master.items()
    ]
    df = pd.DataFrame(rows, columns=MASTER_COLUMNS)
    return df.sort_values("품목코드").reset_index(drop=True)


def _merge_plan(existing_plan: pd.DataFrame, new_df: pd.DataFrame) -> pd.DataFrame:
    """계획 병합. 신규 데이터의 (연월) 은 교체, 나머지 이력 보존."""
    if new_df.empty:
        return existing_plan
    nd = new_df.copy()
    nd["date"] = pd.to_datetime(nd["date"])
    nd["연월"] = nd["date"].dt.strftime("%Y-%m")
    nd["공장"] = nd["factory"].map(FACTORY_CODE_TO_KR).fillna(nd["factory"])
    # (연월, 공장, 품목코드) 별 계획량(월 누계) — 동일 그룹 내 동일 값이므로 first
    new_plan = (
        nd.groupby(["연월", "공장", "품목코드" if "품목코드" in nd.columns else "item_code"], as_index=False)
        .agg(계획량=("planned_qty", "first"))
    )
    new_plan = new_plan.rename(columns={"item_code": "품목코드"})
    new_plan = new_plan[PLAN_COLUMNS]
    new_plan["품목코드"] = new_plan["품목코드"].map(_clean_item_code)

    replace_ym = set(new_plan["연월"])
    kept = existing_plan[~existing_plan["연월"].isin(replace_ym)] if not existing_plan.empty else existing_plan
    out = pd.concat([kept, new_plan], ignore_index=True)
    return out.sort_values(["연월", "공장", "품목코드"]).reset_index(drop=True)


# ── daily(호환용 tidy long) 재생성 ───────────────────────────
def regenerate_daily(
    merged_long: dict[str, pd.DataFrame],
    master_df: pd.DataFrame,
    plan_df: pd.DataFrame,
) -> pd.DataFrame:
    """공장별 long + 제품마스터 + 계획 → tidy long(OUTPUT_COLUMNS).

    production_dw_sync_service 가 읽는 'daily' 시트와 동일 스키마로 복원한다.
    """
    if not merged_long:
        return pd.DataFrame(columns=OUTPUT_COLUMNS)

    master_map = {
        _clean_item_code(r["품목코드"]): r for _, r in master_df.iterrows()
    } if not master_df.empty else {}

    plan_lookup: dict[tuple, float] = {}
    if not plan_df.empty:
        for _, r in plan_df.iterrows():
            plan_lookup[(str(r["연월"]), str(r["공장"]), _clean_item_code(r["품목코드"]))] = float(r["계획량"])

    frames: list[pd.DataFrame] = []
    for code, long in merged_long.items():
        if long.empty:
            continue
        df = long.copy()
        df["item_code"] = df["item_code"].map(_clean_item_code)
        df["factory"] = code
        df["date"] = pd.to_datetime(df["date"])
        kr = FACTORY_CODE_TO_KR.get(code, code)
        ym = df["date"].dt.strftime("%Y-%m")

        df["item_name"] = df["item_code"].map(lambda c: str(master_map[c]["제품명"]) if c in master_map else "")
        df["category1"] = df["item_code"].map(lambda c: master_map[c]["보관유형"] if c in master_map else None)
        df["category2"] = df["item_code"].map(lambda c: master_map[c]["제품유형"] if c in master_map else None)
        df["planned_qty"] = [
            plan_lookup.get((m, kr, c), 0.0) for m, c in zip(ym, df["item_code"])
        ]
        df["date"] = df["date"].dt.date
        frames.append(df[OUTPUT_COLUMNS])

    if not frames:
        return pd.DataFrame(columns=OUTPUT_COLUMNS)
    out = pd.concat(frames, ignore_index=True)
    return out.sort_values(["date", "factory", "category1", "item_code"]).reset_index(drop=True)


# ── DB 파일 저장 (공장별 wide + 제품마스터 + 계획 + daily) ────
def _format_wide_worksheet(ws) -> None:
    from openpyxl.styles import Alignment, Font

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.column_dimensions["A"].width = 14
    for col_idx in range(2, ws.max_column + 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 12
    ws.freeze_panes = "B2"
    for row in range(2, ws.max_row + 1):
        ws.cell(row=row, column=1).number_format = "yyyy-mm-dd"


def _build_group_wides(
    daily_df: pd.DataFrame, groups: list[ProductGroup]
) -> tuple[dict[str, pd.DataFrame], list[str]]:
    """daily tidy long → {품목군 시트명: wide(날짜 × 품목코드)} + 출력 순서.

    각 행을 (factory→F-코드, 보관유형, 제품유형) 으로 품목군 시트에 배정한 뒤
    날짜×품목코드 wide 로 피벗한다. NaN(미등장) 보존, 0 보존.
    """
    if daily_df is None or daily_df.empty:
        return {}, []
    df = daily_df.copy()
    df["item_code"] = df["item_code"].map(_clean_item_code)
    df["__group"] = [
        assign_product_group(f, c1, c2, groups) or _fallback_group_name(f, c1, c2)
        for f, c1, c2 in zip(df["factory"], df["category1"], df["category2"])
    ]

    sheets: dict[str, pd.DataFrame] = {}
    for sn, sub in df.groupby("__group"):
        wide = sub.pivot_table(
            index="date", columns="item_code", values="actual_qty", aggfunc="sum"
        )
        # 전체 기간 연속 날짜 보정 (재공품 ensure_continuous_dates 와 동일 취지)
        wide.index = pd.to_datetime(wide.index)
        full = pd.date_range(wide.index.min(), wide.index.max(), freq="D")
        wide = wide.reindex(full)
        wide.index = wide.index.date
        wide = wide.reset_index().rename(columns={"index": WIDE_DATE_COL})
        sheets[sn] = wide

    # 출력 순서: 품목군 정의(RawDB) 순서 우선, 그 외(폴백) 알파벳
    defined = [g.sheet_name for g in groups if g.sheet_name in sheets]
    extra = sorted(s for s in sheets if s not in defined)
    return sheets, defined + extra


def save_db_file(
    master_df: pd.DataFrame,
    plan_df: pd.DataFrame,
    daily_df: pd.DataFrame,
    output_path: Path | str | None = None,
    groups: list[ProductGroup] | None = None,
) -> Path:
    """품목군별 wide + 제품마스터 + 계획 + daily(호환) 시트를 하나의 xlsx 에 저장.

    품목군 시트는 RawDB_생산실적.xlsx 와 동일한 (공장 × 보관유형[× 제품유형]) 구성이다.
    groups 미지정 시 RawDB 에서 품목군 정의를 로드한다.
    """
    from openpyxl import load_workbook

    out_path = Path(output_path) if output_path else DEFAULT_OUTPUT_PATH

    if groups is None:
        groups = load_product_groups(DEFAULT_RAW_PATH)

    group_sheets, group_order = _build_group_wides(daily_df, groups)

    def _write(target: Path) -> None:
        with pd.ExcelWriter(target, engine="openpyxl") as writer:
            # 품목군 시트 (정의 순서)
            for sn in group_order:
                group_sheets[sn].to_excel(writer, sheet_name=sn, index=False)
            master_df.to_excel(writer, sheet_name=MASTER_SHEET, index=False)
            plan_df.to_excel(writer, sheet_name=PLAN_SHEET, index=False)
            daily_df.to_excel(writer, sheet_name=DAILY_SHEET, index=False)

        # 서식 + daily 시트 숨김(내부 호환용)
        wb = load_workbook(target)
        for sn in group_order:
            if sn in wb.sheetnames:
                _format_wide_worksheet(wb[sn])
        if DAILY_SHEET in wb.sheetnames:
            wb[DAILY_SHEET].sheet_state = "hidden"
        wb.save(target)

    _write_xlsx_atomically(out_path, _write)

    logger.info(f"DB 파일 저장: {out_path} (품목군 시트 {len(group_order)}개)")
    return out_path


# ── 한 번에 통합 + 저장 ──────────────────────────────────────
def build_dataset(
    raw_path: Path | str | None = None,
    output_path: Path | str | None = None,
    src_folder: Path | str | None = None,
) -> tuple[pd.DataFrame, Path]:
    """Raw 파일 → DB 파일 빌드 → (daily DataFrame, 출력 경로) 반환.

    - 신구조(기본): build_dataset(raw_path=..., output_path=...)
    - 레거시 폴더: build_dataset(src_folder=...) → 구버전 폴더 통합(consolidate_folder)
      (migrate_f10_legacy.py 등 기존 호출 호환)
    """
    if src_folder is not None:
        df = consolidate_folder(src_folder)
        out_path = save_consolidated(df, output_path)
        return df, out_path

    raw = Path(raw_path) if raw_path else DEFAULT_RAW_PATH
    out = Path(output_path) if output_path else DEFAULT_OUTPUT_PATH
    existing_source = _existing_db_source(out)

    new_df = consolidate_raw_file(raw)

    existing_long = _load_existing_factory_long(existing_source)
    existing_master = _load_existing_master(existing_source)
    existing_plan = _load_existing_plan(existing_source)

    merged_long = _merge_factory_long(existing_long, new_df)
    master_df = _merge_master(existing_master, new_df)
    plan_df = _merge_plan(existing_plan, new_df)
    daily_df = regenerate_daily(merged_long, master_df, plan_df)

    # 품목군 정의: 이번 회차 RawDB ∪ 기존 DB 시트(이력 보존)
    groups = merge_group_defs(load_product_groups(raw), load_product_groups(existing_source))
    save_db_file(master_df, plan_df, daily_df, out, groups=groups)
    return daily_df, out

