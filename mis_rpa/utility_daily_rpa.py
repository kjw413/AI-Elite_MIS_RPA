# MIS 유틸리티 일자별 사용량 추이 데이터 자동 샘플링 RPA (pywinauto + openpyxl 기반)
"""
사내 MIS '(신)종합정보' 시스템에서 유틸리티 일자별 사용량 추이 데이터를
사업장별로 자동 조회 → 클립보드 복사 → Excel 시트별 붙여넣기하는 RPA 프로그램.

업무 절차:
  1. MIS 앱 연결 (pywinauto UIA backend)
  2. 사업장별 순회 (F1A→F1B→F20→F30→F40)
     a. ORG 드롭다운에서 공장 선택
     b. 기준년월 설정 (D-2 기준)
     c. 조회 버튼 클릭 + 로딩 대기
     d. 그리드 Ctrl+A → Ctrl+C → 확인 다이얼로그
     e. 클립보드 TSV 파싱 (필드명 제거, 데이터만)
     f. Excel 해당 시트 마지막 열 다음에 날짜+데이터 추가
  3. Excel 저장

Usage:
  python utility_daily_rpa.py              # 기본: D-2 기준월
  python utility_daily_rpa.py --ym 2026-05 # 특정 월 지정
  python utility_daily_rpa.py --dry-run    # MIS 조회만, Excel 미기록
"""

import sys
import time
import os
import json
import shutil
import logging
import argparse
import re
from collections import OrderedDict
from datetime import datetime, date, timedelta
from pathlib import Path

import openpyxl
import pandas as pd
from pywinauto import Application
from pywinauto.keyboard import send_keys
from pywinauto.timings import Timings

# pywinauto 내부 click/keys 대기 시간 단축 (MIS는 즉시 반응한다는 가정)
Timings.after_clickinput_wait = 0.01
Timings.after_setfocus_wait = 0.01
Timings.after_sendkeys_key_wait = 0.001

# 프로젝트 루트(app/tools/...) import 보장 — _common.fast_click 사용을 위함
_PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(_PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(_PROJECT_ROOT))

from mis_rpa._common import (  # noqa: E402
    atomic_save_workbook,
    fast_click,
    find_mis_window,
    get_clipboard_sequence,
    sampled_db_path,
    wait_for_clipboard_change,
)

# ---------------------------------------------------------------------------
# 로깅 설정
# ---------------------------------------------------------------------------
LOG_DIR = Path(__file__).parent / "logs"
LOG_DIR.mkdir(exist_ok=True)
log = logging.getLogger(__name__)


def _setup_logging() -> None:
    if logging.getLogger().handlers:
        return
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(message)s",
        handlers=[
            logging.StreamHandler(sys.stdout),
            logging.FileHandler(
                LOG_DIR / f"rpa_{datetime.now():%Y%m%d_%H%M%S}.log",
                encoding="utf-8",
            ),
        ],
    )

# ---------------------------------------------------------------------------
# 상수 정의
# ---------------------------------------------------------------------------
# 사업장 코드 → Excel 시트명 매핑 (순회 순서 보장)
FACTORY_SHEET_MAP = OrderedDict([
    ("F1A", "남양주1"),
    ("F1B", "남양주2"),
    ("F20", "김해"),
    ("F30", "광주"),
    ("F40", "논산"),
])

# Excel 파일 경로
EXCEL_PATH = sampled_db_path("RawDB_에너지.xlsx", "ENERGY_SOURCE_XLSX")

# MIS 그리드 행 순서 (Excel row 2~11은 MIS 그리드 원본 그대로)
FIELD_ORDER = [
    "냉동전력량",
    "공압기",
    "전력량",
    "연료량",
    "용수량",
    "폐수량",
    "믹스",         # 믹스생산량 / 믹스량, 단위: kg
    "전력원단위",
    "연료원단위",
    "용수원단위",
]

# 대기 시간 기본값 (utility_coords.json의 "wait" 값으로 덮어씌워짐)
WAIT_SHORT = 0.05        # 클릭/타이핑 후 미세 대기 (MIS는 즉시 반응)
WAIT_MEDIUM = 0.2        # 클립보드 fill / 필드 클릭 후 약간 더 긴 대기
WAIT_DROPDOWN = 0.01     # 드롭다운 펼침 후 항목 클릭 전 대기
WAIT_SCREEN_LOAD = 1.0   # 사이드바 메뉴 클릭 → MIS 화면 전환 로딩
WAIT_QUERY_LOAD = 1.0    # 조회 버튼 → 그리드 데이터 로딩
WAIT_COPY_CONFIRM = 0.4  # 복사 버튼 클릭 후 확인 팝업이 포커스를 받을 때까지 대기


# ---------------------------------------------------------------------------
# 클립보드 헬퍼
# ---------------------------------------------------------------------------
def get_clipboard_text() -> str:
    """Windows 클립보드에서 MIS 데이터를 읽어온다 (win32clipboard)."""
    import win32clipboard

    # MIS가 사용하는 'Csv' 커스텀 클립보드 포맷 ID
    csv_fmt = win32clipboard.RegisterClipboardFormat("Csv")

    for attempt in range(3):
        try:
            win32clipboard.OpenClipboard()
            try:
                # 1차: Csv 포맷 시도
                try:
                    data = win32clipboard.GetClipboardData(csv_fmt)
                    if data:
                        # bytes인 경우 디코딩
                        if isinstance(data, bytes):
                            for enc in ("utf-8", "euc-kr", "cp949"):
                                try:
                                    text = data.decode(enc).rstrip("\x00")
                                    if text.strip():
                                        log.info(f"  클립보드 읽기 성공 (Csv/{enc}, {len(text)}자)")
                                        return text
                                except UnicodeDecodeError:
                                    continue
                        elif isinstance(data, str) and data.strip():
                            log.info(f"  클립보드 읽기 성공 (Csv/str, {len(data)}자)")
                            return data
                except Exception:
                    pass

                # 2차: 표준 텍스트 포맷 시도
                try:
                    text = win32clipboard.GetClipboardData(win32clipboard.CF_UNICODETEXT)
                    if text and text.strip():
                        log.info(f"  클립보드 읽기 성공 (UNICODETEXT, {len(text)}자)")
                        return text
                except Exception:
                    pass

                try:
                    text = win32clipboard.GetClipboardData(win32clipboard.CF_TEXT)
                    if text:
                        decoded = text.decode("utf-8", errors="replace") if isinstance(text, bytes) else text
                        if decoded.strip():
                            log.info(f"  클립보드 읽기 성공 (TEXT, {len(decoded)}자)")
                            return decoded
                except Exception:
                    pass

            finally:
                win32clipboard.CloseClipboard()

        except Exception as e:
            log.warning(f"  클립보드 시도 {attempt + 1}/3 실패: {e}")

        time.sleep(0.5)

    log.warning("  클립보드 읽기 실패 (3회 시도)")
    return ""


# ---------------------------------------------------------------------------
# 클립보드 TSV 파싱
# ---------------------------------------------------------------------------
def parse_grid_clipboard(raw_text: str, year_month: str):
    """
    MIS 그리드에서 복사된 CSV/TSV 텍스트를 파싱한다.

    Returns:
        dates: list[str]   - 날짜 목록 ['26-05-01', '26-05-02', ...]
        data:  list[list]  - 10행 × N열 숫자 데이터 (0~9행: MIS 그리드 원본 = FIELD_ORDER)
    """
    import csv
    import io

    raw_text = raw_text.strip()
    if not raw_text:
        raise ValueError("클립보드 데이터가 비어 있습니다.")

    # 디버깅: 첫 300자 출력
    log.info(f"  클립보드 데이터 미리보기: {repr(raw_text[:300])}")

    # 구분자 감지 (탭 vs 쉼표)
    first_line = raw_text.split("\n")[0]
    if "\t" in first_line and first_line.count("\t") >= 2:
        sep = "\t"
    else:
        sep = ","
    log.info(f"  구분자 감지: {'TAB' if sep == chr(9) else 'COMMA'}")

    # csv.reader 사용 (인용부호 내 쉼표 = 천단위 구분기호 정상 처리)
    reader = csv.reader(io.StringIO(raw_text), delimiter=sep)
    rows = list(reader)

    if not rows:
        raise ValueError("파싱된 행이 없습니다.")

    # 첫 행: 헤더 (일자)
    header_cols = rows[0]
    day_numbers = []
    for col in header_cols:
        m = re.search(r"(\d+)", col.strip())
        if m:
            day_numbers.append(int(m.group(1)))

    # 날짜 객체 생성 (datetime.date)
    # 텍스트(`'26-05-14`)로 저장하면 sync 단계에서 pd.to_datetime이 NaT로 처리해
    # 신규 행이 silently 누락되므로(실제 발생 사례), datetime.date 객체로 저장한다.
    ym_parts = year_month.split("-")
    year_int = int(ym_parts[0])
    month_int = int(ym_parts[1])
    dates = [date(year_int, month_int, d) for d in day_numbers]

    # 데이터 행 파싱 (첫 열 = 필드명 제거, 나머지 = 숫자)
    data_rows = []
    for row in rows[1:]:
        if len(row) < 2:
            continue
        values = []
        for val_str in row[1:]:
            # 천단위 구분기호, 인용부호, 공백 제거
            val_str = val_str.strip().replace(",", "").replace('"', '')
            try:
                if "." in val_str:
                    values.append(float(val_str))
                else:
                    values.append(int(val_str) if val_str else 0)
            except ValueError:
                values.append(0)
        data_rows.append(values)

    # 폐수 원단위(폐수량/(믹스/1000))는 폐기됨 — 별도 계산 행을 추가하지 않고
    # MIS 그리드 원본(FIELD_ORDER) 그대로 적재한다. 폐수/용수 비는 화면/메일에서
    # 폐수량·용수량 raw 값으로 즉석 계산한다.

    log.info(f"파싱 완료: {len(dates)}일 × {len(data_rows)}행")
    if data_rows:
        log.info(f"  첫 행 미리보기: {data_rows[0][:5]}")
    return dates, data_rows


# ---------------------------------------------------------------------------
# Excel 기록
# ---------------------------------------------------------------------------
def write_to_excel(excel_path: str, sheet_name: str,
                   dates: list, data_rows: list):
    """
    Excel 시트에 월 전체 일자 데이터를 기록한다.
    - 기존 날짜 열이 있으면 → 최신 데이터로 덮어쓰기
    - 새 날짜 열이면     → 마지막 열 다음에 추가
    """
    wb = openpyxl.load_workbook(excel_path)
    if sheet_name not in wb.sheetnames:
        log.warning(f"시트 '{sheet_name}'이(가) 없습니다. 스킵합니다.")
        wb.close()
        return 0

    ws = wb[sheet_name]

    # 기존 날짜 헤더 → 열 번호 매핑 (Row 1 스캔)
    # 셀 값이 datetime/date/문자열 어느 형식이든 "YY-MM-DD" 키로 통일해 비교한다.
    existing_date_col = {}   # { "YY-MM-DD": 열번호 }
    last_col = 1
    for col_idx in range(1, ws.max_column + 2):
        cell_val = ws.cell(row=1, column=col_idx).value
        if cell_val is not None:
            if isinstance(cell_val, (datetime, date)):
                normalized = cell_val.strftime("%y-%m-%d")
            else:
                normalized = str(cell_val).strip().lstrip("'")
            existing_date_col[normalized] = col_idx
            last_col = col_idx
        else:
            break

    next_col = last_col + 1   # 새 날짜 추가 시작 위치
    overwritten = 0
    appended = 0

    for day_idx, date_obj in enumerate(dates):
        # 비교 키 정규화 — date 객체와 기존 datetime/문자열 헤더 모두 매칭
        if isinstance(date_obj, (datetime, date)):
            date_key = date_obj.strftime("%y-%m-%d")
        else:
            date_key = str(date_obj).lstrip("'")

        if date_key in existing_date_col:
            # ── 기존 날짜: 덮어쓰기 (데이터 변동 반영) ──
            target_col = existing_date_col[date_key]
            for row_idx, row_data in enumerate(data_rows):
                if day_idx < len(row_data):
                    ws.cell(row=row_idx + 2, column=target_col,
                            value=row_data[day_idx])
            overwritten += 1
        else:
            # ── 새 날짜: 마지막 열 다음에 추가 (date 객체로 저장 + YY-MM-DD 표시 서식) ──
            new_cell = ws.cell(row=1, column=next_col, value=date_obj)
            new_cell.number_format = "YY-MM-DD"
            for row_idx, row_data in enumerate(data_rows):
                if day_idx < len(row_data):
                    ws.cell(row=row_idx + 2, column=next_col,
                            value=row_data[day_idx])
            appended += 1
            next_col += 1

    atomic_save_workbook(wb, excel_path)
    wb.close()
    log.info(f"  [{sheet_name}] 덮어쓰기 {overwritten}열 / 신규 추가 {appended}열 "
             f"(총 {len(dates)}일)")
    return appended


# ---------------------------------------------------------------------------
# MIS RPA 클래스
# ---------------------------------------------------------------------------
class MISUtilityRPA:
    """MIS 유틸리티 일자별 사용량 추이 데이터 자동 샘플링 RPA"""

    def __init__(self, year_month: str = None, dry_run: bool = False):
        if year_month is None:
            ref_date = datetime.now() - timedelta(days=2)
            self.year_month = ref_date.strftime("%Y-%m")
        else:
            self.year_month = year_month
        self.dry_run = dry_run
        
        # 좌표 설정 로드
        self.coords = self._load_coords()

        self.app = None
        self.main_window = None
        log.info(f"=== MIS 유틸리티 RPA 초기화 ===")
        log.info(f"  기준년월: {self.year_month}  (D-2 자동 계산)")
        log.info(f"  Dry-run: {self.dry_run}")

    # -----------------------------------------------------------------------
    # MIS 연결
    # -----------------------------------------------------------------------
    def attach_existing_window(self, app, main_window) -> None:
        """오케스트레이터가 이미 연결한 MIS 윈도우를 주입 — 재연결 생략."""
        self.app = app
        self.main_window = main_window

    def connect_mis(self):
        """실행 중인 MIS (신)종합정보 창에 연결한다."""
        if self.main_window is not None:
            log.info("MIS 연결 재사용 (이전 단계에서 연결됨)")
            return
        log.info("MIS 앱 연결 중...")
        # Win32 로 HWND 를 먼저 찾고 handle 로 connect 한다. UIA title_re 스캔은
        # 바탕화면 전체 트리를 순회해 수 초~20초+ 까지 들쭉날쭉 걸리지만, 핸들
        # 연결은 트리 스캔이 없어 ~10ms 로 끝난다.
        hwnd, title = find_mis_window("(신)종합정보")
        if not hwnd:
            log.error("MIS (신)종합정보 창을 찾을 수 없습니다.")
            log.error("MIS (신)종합정보를 먼저 실행해주세요.")
            raise SystemExit(1)
        try:
            self.app = Application(backend="uia").connect(handle=hwnd)
            self.main_window = self.app.window(handle=hwnd)
            log.info(f"MIS 연결 성공: {title}")
        except Exception as e:
            log.error(f"MIS 앱 연결 실패: {e}")
            log.error("MIS (신)종합정보를 먼저 실행해주세요.")
            raise SystemExit(1)

    # -----------------------------------------------------------------------
    # 메뉴 진입
    # -----------------------------------------------------------------------
    def navigate_to_utility_screen(self):
        """좌측 트리에서 '유틸리티 일자별 사용량 추이' 메뉴를 더블클릭한다."""
        log.info("유틸리티 일자별 사용량 추이 화면으로 이동 중...")
        # 좌표 기반 더블클릭 — UIA child_window(TreeItem) 검색은 트리 전체를
        # 순회해 수 초~20초+ 까지 들쭉날쭉 걸리므로 좌표 클릭으로 대체.
        x, y = self.coords.get("tree_menu", [171, 163])
        fast_click(self.main_window, x, y, double=True)
        log.info(f"  트리메뉴 더블클릭 ({x}, {y})")
        time.sleep(WAIT_SCREEN_LOAD)
        log.info("화면 이동 완료")

    def _load_coords(self):
        """JSON 파일에서 좌표 및 대기 시간 설정을 로드한다."""
        global WAIT_SHORT, WAIT_MEDIUM, WAIT_DROPDOWN, WAIT_SCREEN_LOAD, WAIT_QUERY_LOAD
        global WAIT_COPY_CONFIRM

        coord_path = os.path.join(os.path.dirname(__file__), "utility_coords.json")
        try:
            with open(coord_path, "r", encoding="utf-8") as f:
                config = json.load(f)
                log.info(f"설정 파일 로드 완료: {coord_path}")

                # 대기 시간 로드
                wait = config.get("wait", {})
                WAIT_SHORT = wait.get("short", WAIT_SHORT)
                WAIT_MEDIUM = wait.get("medium", WAIT_MEDIUM)
                WAIT_DROPDOWN = wait.get("dropdown", WAIT_DROPDOWN)
                WAIT_SCREEN_LOAD = wait.get("screen_load", WAIT_SCREEN_LOAD)
                WAIT_QUERY_LOAD = wait.get("query_load", WAIT_QUERY_LOAD)
                WAIT_COPY_CONFIRM = wait.get("copy_confirm", WAIT_COPY_CONFIRM)
                log.info(f"  대기 시간: short={WAIT_SHORT}s, medium={WAIT_MEDIUM}s, "
                         f"dropdown={WAIT_DROPDOWN}s, screen={WAIT_SCREEN_LOAD}s, "
                         f"query={WAIT_QUERY_LOAD}s, copy_confirm={WAIT_COPY_CONFIRM}s")

                return config.get("coords", {})
        except Exception as e:
            log.warning(f"설정 파일을 읽을 수 없습니다 ({e}). 기본값을 사용합니다.")
            return {}

    # -----------------------------------------------------------------------
    # ORG 공장 선택 (순수 좌표 기반 — MIS는 커스텀 렌더링 UI로 UIA 미지원)
    # -----------------------------------------------------------------------
    def select_factory(self, org_code: str):
        """ORG 드롭다운에서 공장을 선택한다 (좌표 기반)."""
        org_name = FACTORY_SHEET_MAP.get(org_code, org_code)
        log.info(f"공장 선택: {org_code} ({org_name})")

        # 1. 드롭다운 클릭하여 열기
        x, y = self.coords.get("factory_dropdown", [417, 106])
        fast_click(self.main_window, x, y)
        log.info(f"  드롭다운 클릭 ({x}, {y})")
        time.sleep(WAIT_DROPDOWN)

        # 2. 드롭다운이 열린 상태에서 해당 공장 항목 좌표 클릭
        factory_list = self.coords.get("factory_list", {})
        item_y = factory_list.get(org_code)

        if item_y is None:
            # JSON에 없는 공장 → 간격 기반 계산
            start_y = factory_list.get("F1A", 167)
            order = ["F1A", "F1B", "F30", "F20", "F10", "F40", "F50"]
            idx = order.index(org_code) if org_code in order else 0
            item_y = start_y + (idx * 18)

        fast_click(self.main_window, x, item_y)
        log.info(f"  공장 항목 클릭 ({x}, {item_y})")
        time.sleep(WAIT_SHORT)

    # -----------------------------------------------------------------------
    # 기준년월 설정 (순수 좌표 기반)
    # -----------------------------------------------------------------------
    def set_year_month(self):
        """기준년월 필드에 값을 설정한다 (좌표 기반)."""
        log.info(f"기준년월 설정: {self.year_month}")

        # 기준년월 입력 영역 클릭
        x, y = self.coords.get("month_filter", [601, 107])
        log.info(f"  기준년월 클릭 ({x}, {y})")
        fast_click(self.main_window, x, y)
        time.sleep(WAIT_SHORT)
        send_keys("^a")
        time.sleep(WAIT_SHORT)
        send_keys(self.year_month, with_spaces=True)
        time.sleep(WAIT_SHORT)

    # -----------------------------------------------------------------------
    # 조회
    # -----------------------------------------------------------------------
    def click_query(self):
        """조회 버튼을 클릭하고 데이터 로딩을 기다린다 (좌표 기반)."""
        log.info("조회 버튼 클릭...")

        x, y = self.coords.get("query_button", [328, 66])
        fast_click(self.main_window, x, y)
        log.info(f"  조회 클릭 ({x}, {y})")

        # 로딩 대기
        log.info("  데이터 로딩 대기 중...")
        time.sleep(WAIT_QUERY_LOAD)
        log.info("  조회 완료")

    # -----------------------------------------------------------------------
    # 그리드 복사
    # -----------------------------------------------------------------------
    def copy_grid_data(self) -> str:
        """
        그리드 상단의 복사 버튼을 클릭하여 클립보드에 데이터를 복사하고,
        확인 팝업을 닫은 후 클립보드 텍스트를 반환한다.

        로딩 도중 클릭이 흡수되면 클립보드가 갱신되지 않을 수 있으므로
        ClipboardSequenceNumber 폴링으로 갱신 여부를 검증하고, 안 바뀌면 재클릭한다.
        """
        log.info("그리드 데이터 클립보드 복사 중...")

        x, y = self.coords.get("copy_button", [332, 725])
        log.info(f"  복사 버튼 클릭 ({x}, {y})")

        clipboard_text = ""
        for attempt in range(1, 4):  # 최대 3회 시도
            seq_before = get_clipboard_sequence()
            fast_click(self.main_window, x, y)
            changed = wait_for_clipboard_change(seq_before, timeout=2.0)
            if changed:
                if attempt > 1:
                    log.info(f"  클립보드 갱신 확인 (재시도 {attempt}/3)")
                # 클립보드 읽기 (팝업 닫기 전에! — MIS가 닫을 때 클립보드를 비울 수 있음)
                clipboard_text = get_clipboard_text()
                self._handle_copy_confirm_dialog()
                break
            self._handle_copy_confirm_dialog(use_ok_click=False)
            log.warning(f"  복사 시도 {attempt}/3 — 2s 내 클립보드 변화 없음")
            time.sleep(WAIT_MEDIUM)
        else:
            log.warning("  클립보드 변경 감지 실패 → 기존 클립보드 그대로 읽기")
            time.sleep(WAIT_MEDIUM)
            clipboard_text = get_clipboard_text()

        # 4. 결과 확인
        if not clipboard_text.strip():
            # pandas로 재시도 (OLE/DataObject 포맷 대응)
            log.info("  ctypes 실패 → pandas.read_clipboard 시도")
            try:
                df = pd.read_clipboard(sep="\t", header=None)
                clipboard_text = df.to_csv(sep="\t", index=False, header=False)
                log.info(f"  pandas 클립보드 읽기 성공: {df.shape}")
            except Exception as e:
                log.warning(f"  pandas.read_clipboard 실패: {e}")

        if not clipboard_text.strip():
            log.warning("클립보드가 비어 있습니다!")
        else:
            lines = clipboard_text.strip().split("\n")
            log.info(f"  클립보드 데이터: {len(lines)}행")

        return clipboard_text

    def _handle_copy_confirm_dialog(self, use_ok_click: bool = True):
        """'데이터가 클립보드로 복사되었습니다' 확인 팝업을 닫는다."""
        time.sleep(WAIT_COPY_CONFIRM)
        ok_xy = self.coords.get("confirm_popup_ok") if use_ok_click else None
        if ok_xy:
            fast_click(self.main_window, ok_xy[0], ok_xy[1])
            log.info(f"  확인 팝업 닫기 완료 (OK 클릭, {ok_xy[0]}, {ok_xy[1]})")
        else:
            send_keys("{ENTER}")
            action = "완료" if use_ok_click else "시도"
            log.info(f"  확인 팝업 닫기 {action} (Enter)")
        time.sleep(WAIT_SHORT)

    # -----------------------------------------------------------------------
    # 전체 실행
    # -----------------------------------------------------------------------
    def run(self):
        """모든 사업장을 순회하며 데이터를 추출 → Excel에 기록한다."""
        log.info("=" * 60)
        log.info("MIS 유틸리티 일자별 사용량 추이 RPA 시작")
        log.info("=" * 60)

        # 1. MIS 연결
        self.connect_mis()
        self.main_window.set_focus()
        time.sleep(WAIT_MEDIUM)

        # 2. 화면 진입
        self.navigate_to_utility_screen()

        # 3. 기준년월 설정 (최초 1회)
        self.set_year_month()

        # 3-1. Excel 백업 (최초 1회)
        if not self.dry_run:
            excel_path = EXCEL_PATH
            backup_dir = os.path.join(os.path.dirname(excel_path), "backup")
            os.makedirs(backup_dir, exist_ok=True)
            base_name = os.path.splitext(os.path.basename(excel_path))[0]
            ext = os.path.splitext(excel_path)[1]
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            backup_path = os.path.join(backup_dir, f"{base_name}_backup_{timestamp}{ext}")
            shutil.copy2(excel_path, backup_path)
            log.info(f"백업 생성 완료: {backup_path}")

        # 4. 사업장 순회
        total_added = 0
        for org_code, sheet_name in FACTORY_SHEET_MAP.items():
            log.info("-" * 40)
            log.info(f"▶ 사업장 처리: {org_code} → {sheet_name}")
            log.info("-" * 40)

            try:
                # 4-1. 공장 선택
                self.select_factory(org_code)
                time.sleep(WAIT_SHORT)

                # 4-2. 조회
                self.click_query()

                # 4-3. 그리드 복사
                clipboard_text = self.copy_grid_data()
                if not clipboard_text.strip():
                    log.warning(f"  {org_code}: 데이터 없음 → 스킵")
                    continue

                # 4-4. 클립보드 파싱
                dates, data_rows = parse_grid_clipboard(
                    clipboard_text, self.year_month
                )

                if not dates:
                    log.warning(f"  {org_code}: 파싱된 날짜 없음 → 스킵")
                    continue

                # 4-5. Excel 기록
                if self.dry_run:
                    log.info(f"  [DRY-RUN] {sheet_name}: "
                             f"{len(dates)}일 × {len(data_rows)}행 (기록 안함)")
                else:
                    added = write_to_excel(
                        EXCEL_PATH, sheet_name, dates, data_rows
                    )
                    total_added += added

                # MIS 창으로 포커스 복귀
                self.main_window.set_focus()
                time.sleep(WAIT_SHORT)

            except Exception as e:
                log.error(f"  {org_code} 처리 중 오류: {e}", exc_info=True)
                # 다음 공장 계속 처리
                try:
                    self.main_window.set_focus()
                except Exception:
                    pass
                continue

        log.info("=" * 60)
        if self.dry_run:
            log.info("DRY-RUN 완료 (Excel 미기록)")
        else:
            log.info(f"RPA 완료: 총 {total_added}개 날짜 열 추가")
        log.info("=" * 60)


# ---------------------------------------------------------------------------
# 메인 진입점
# ---------------------------------------------------------------------------
def main():
    _setup_logging()
    parser = argparse.ArgumentParser(
        description="MIS 유틸리티 일자별 사용량 추이 RPA"
    )
    parser.add_argument(
        "--ym", type=str, default=None,
        help="기준년월 (YYYY-MM). 미지정 시 D-2 자동 계산"
    )
    parser.add_argument(
        "--dry-run", action="store_true",
        help="MIS 조회만 실행, Excel 기록하지 않음"
    )
    args = parser.parse_args()

    rpa = MISUtilityRPA(year_month=args.ym, dry_run=args.dry_run)
    rpa.run()


if __name__ == "__main__":
    main()
