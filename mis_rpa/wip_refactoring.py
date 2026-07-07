# -*- coding: utf-8 -*-
"""
일일 재공품 실적 정리 자동화 스크립트 (멀티 공장 버전)

[최종 규칙]
1. RawData의 기간(또는 설정한 고정 기간) 안에서는 모든 날짜를 생성
2. 해당 날짜에 지정 ItemCode 실적이 없으면 0으로 기록
3. 기존 출력파일은 유지
4. 단, 업데이트는 이번 RawData 기간 안에서만 수행
5. 업데이트 기간 밖의 기존 값은 절대 변경하지 않음
6. RawData에 잠시 미등장한 기존 ItemCode 컬럼도 삭제하지 않음
7. 출력은 공장별 시트(남양주1, 남양주2, 김해, 광주, 논산)로 저장
   (구버전 F-코드 시트가 있으면 자동 이관)
"""

from __future__ import annotations

import os
import shutil
import sys
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font


# =========================================================
# 1) 사용자 설정
# =========================================================

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from mis_rpa.config import sampled_db_path_str


INPUT_FILE = sampled_db_path_str("RawDB_재공품.xlsx", "WIP_ITEM_MASTER_XLSX")
OUTPUT_FILE = sampled_db_path_str("DB_재공품.xlsx", "WIP_SUMMARY_XLSX")

PLANTS = ["F10A", "F10B", "F20", "F30", "F40"]

# 공장별 ItemCode 화이트리스트 (PLANT_ITEMCODES).
# 모든 공장이 비어 있음 = RawData 에 등장한 ItemCode 를 그대로 추적
# (resolve_target_itemcodes 의 fallback). 광주(F30)도 동일.
#
# 분모 보정용 ItemCode + 환산계수는 별도 모듈
# (app/services/production_correction_service.py) 에서 관리하므로,
# 여기서 광주 7품목을 다시 hardcode 할 필요 없다.
PLANT_ITEMCODES: Dict[str, List[str]] = {}

INPUT_SHEET_BY_PLANT: Dict[str, str] = {
    "F10A": "남양주1",
    "F10B": "남양주2",
    "F20": "김해",
    "F30": "광주",
    "F40": "논산",
}

# RawDB의 시트명이 아직 한글로 통일되지 않은 환경(또는 다시 F-코드로 되돌린
# 환경)을 위해 구버전 F-코드 시트도 폴백 후보로 둡니다.
LEGACY_INPUT_SHEET_BY_PLANT: Dict[str, List[str]] = {
    "F10A": ["F10A"],
    "F10B": ["F10B"],
    "F20": ["F20"],
    "F30": ["F30"],
    "F40": ["F40"],
}

OUTPUT_SHEET_BY_PLANT: Dict[str, str] = {
    "F10A": "남양주1",
    "F10B": "남양주2",
    "F20": "김해",
    "F30": "광주",
    "F40": "논산",
}

# 새 시트명이 아직 없으면 동일 공장의 구버전 시트(F-코드 또는 옛 'Sheet1')를
# 그대로 읽어와 새 시트로 이관합니다. 한 공장에 대해 여러 후보가 필요한 경우
# 첫 매칭 항목을 사용합니다.
LEGACY_OUTPUT_SHEET_BY_PLANT: Dict[str, List[str]] = {
    "F10A": ["F10A"],
    "F10B": ["F10B"],
    "F20": ["F20"],
    "F30": ["F30", "Sheet1"],
    "F40": ["F40"],
}

# True면 아래 기간 강제 사용
# False면 RawData에 존재하는 날짜 최소~최대 사용
USE_FIXED_PERIOD = False
START_DATE = "2025-01-01"
END_DATE = "2026-03-19"

CREATE_BACKUP = True

COL_DATE = "일"
COL_ITEM = "ItemCode"
COL_ACTUAL = "실적량"
OUT_DATE_COL = "날짜"


# =========================================================
# 2) 공통 함수
# =========================================================

def log(msg: str):
    print(f"[INFO] {msg}")


def validate_file_exists(file_path: str):
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"파일이 존재하지 않습니다: {file_path}")


def ensure_parent_folder(file_path: str):
    folder = os.path.dirname(file_path)
    if folder and not os.path.exists(folder):
        os.makedirs(folder, exist_ok=True)


def backup_output_file(file_path: str):
    if CREATE_BACKUP and os.path.exists(file_path):
        folder = os.path.dirname(file_path)
        name, ext = os.path.splitext(os.path.basename(file_path))
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = os.path.join(folder, f"{name}_backup_{ts}{ext}")
        shutil.copy2(file_path, backup_path)
        log(f"백업 생성 완료: {backup_path}")


def clean_datetime(series: pd.Series) -> pd.Series:
    s = series.astype(str).str.strip()
    s = s.replace({
        "": pd.NA,
        "nan": pd.NA,
        "NaT": pd.NA,
        "None": pd.NA,
    })
    dt = pd.to_datetime(s, errors="coerce")
    return dt.dt.normalize()


def clean_numeric(series: pd.Series) -> pd.Series:
    return pd.to_numeric(series, errors="coerce").fillna(0)


def normalize_itemcodes(itemcodes: List[str]) -> List[str]:
    normalized: List[str] = []
    seen = set()

    for code in itemcodes:
        txt = str(code).strip()
        if not txt:
            continue
        if txt.lower() in {"nan", "none", "nat"}:
            continue
        # float→str 변환 잔재 ".0" 제거 (ItemCode 는 항상 정수 코드)
        if txt.endswith(".0") and txt[:-2].isdigit():
            txt = txt[:-2]
        if txt in seen:
            continue
        seen.add(txt)
        normalized.append(txt)

    return normalized


def get_excel_sheet_names(file_path: str) -> List[str]:
    with pd.ExcelFile(file_path) as xl:
        return list(xl.sheet_names)


def resolve_input_sheet_name(plant: str, input_sheet_names: List[str]) -> str:
    requested = INPUT_SHEET_BY_PLANT[plant]

    if requested in input_sheet_names:
        return requested

    lower_map = {name.strip().lower(): name for name in input_sheet_names}
    if requested.strip().lower() in lower_map:
        return lower_map[requested.strip().lower()]

    # 한글 시트가 없으면 구버전 F-코드 시트로 폴백
    for legacy in LEGACY_INPUT_SHEET_BY_PLANT.get(plant, []):
        if legacy in input_sheet_names:
            log(f"[{plant}] 입력 시트 '{requested}' 미존재 -> 레거시 시트 '{legacy}' 사용")
            return legacy
        if legacy.strip().lower() in lower_map:
            mapped = lower_map[legacy.strip().lower()]
            log(f"[{plant}] 입력 시트 '{requested}' 미존재 -> 레거시 시트 '{mapped}' 사용")
            return mapped

    if len(input_sheet_names) == 1:
        fallback = input_sheet_names[0]
        log(
            f"[{plant}] 입력 시트 '{requested}' 미존재 -> 단일 시트 '{fallback}'로 대체"
        )
        return fallback

    raise KeyError(
        f"[{plant}] 입력 시트 '{requested}'를 찾을 수 없습니다. "
        f"현재 시트: {input_sheet_names}"
    )


def resolve_existing_output_sheet_name(
    plant: str,
    output_sheet_names: List[str],
) -> Optional[str]:
    target = OUTPUT_SHEET_BY_PLANT[plant]
    if target in output_sheet_names:
        return target

    for legacy in LEGACY_OUTPUT_SHEET_BY_PLANT.get(plant, []):
        if legacy in output_sheet_names:
            log(f"[{plant}] 레거시 시트 '{legacy}' 데이터를 '{target}'로 이관 반영")
            return legacy

    return None


def validate_configuration():
    # PLANT_ITEMCODES 는 명시 화이트리스트가 필요한 공장만 등록 (현재 비어 있음 = 모든 공장 raw 추적).
    # 누락 = "RawData ItemCode 자동 추적" 의미이므로 검증하지 않는다.

    missing_in_input_sheet = [p for p in PLANTS if p not in INPUT_SHEET_BY_PLANT]
    if missing_in_input_sheet:
        raise KeyError(f"INPUT_SHEET_BY_PLANT 설정 누락: {missing_in_input_sheet}")

    missing_in_output_sheet = [p for p in PLANTS if p not in OUTPUT_SHEET_BY_PLANT]
    if missing_in_output_sheet:
        raise KeyError(f"OUTPUT_SHEET_BY_PLANT 설정 누락: {missing_in_output_sheet}")

    duplicates = pd.Series(list(OUTPUT_SHEET_BY_PLANT.values()))
    dup_values = sorted(duplicates[duplicates.duplicated()].unique().tolist())
    if dup_values:
        raise ValueError(f"출력 시트명이 중복됩니다: {dup_values}")

    for plant in PLANTS:
        codes = normalize_itemcodes(PLANT_ITEMCODES.get(plant, []))
        if not codes:
            log(f"[정보] [{plant}] PLANT_ITEMCODES 미지정 -> RawData 자동 추적 모드")


def get_update_period(raw_df: pd.DataFrame) -> Tuple[pd.Timestamp, pd.Timestamp]:
    if USE_FIXED_PERIOD:
        start = pd.to_datetime(START_DATE).normalize()
        end = pd.to_datetime(END_DATE).normalize()
        if start > end:
            raise ValueError(f"기간 설정 오류: {start.date()} > {end.date()}")
        return start, end

    valid_dates = raw_df[COL_DATE].dropna()
    if valid_dates.empty:
        raise ValueError("입력 RawData에 유효한 날짜가 없습니다.")

    raw_min = valid_dates.min().normalize()
    raw_max = valid_dates.max().normalize()

    # raw 의 시작 날짜만 제외.
    # MIS 특성상 조회 시작일은 전일 실적이 끼어드는 케이스가 있어 update window 에서 제외한다.
    # 끝 날짜는 보존 — RPA 가 D-2까지만 조회한다는 전제 하에 raw_max 는 완료된 영업일이므로
    # 잘라내면 분석/예측에 D-2 데이터가 1일 늦게 반영되는 부작용이 더 큼.
    span_days = (raw_max - raw_min).days
    if span_days >= 1:
        start = raw_min + pd.Timedelta(days=1)
        end = raw_max
        log(
            f"[기간 보정] raw {raw_min.date()}~{raw_max.date()} 중 시작일 제외 "
            f"-> 업데이트 윈도우 {start.date()}~{end.date()}"
        )
    else:
        # raw 기간이 1일이면 그 하루만 사용.
        start = raw_min
        end = raw_max
        log(
            f"[기간 보정] raw 기간 1일 -> raw 전체 {start.date()}~{end.date()} 사용"
        )

    if start > end:
        raise ValueError(f"기간 설정 오류: {start.date()} > {end.date()}")

    return start, end


# =========================================================
# 3) RawData 로드
# =========================================================

def load_raw_data(input_file: str, sheet_name: str, plant: str) -> pd.DataFrame:
    log(f"[{plant}] RawData 읽는 중... (시트: {sheet_name})")
    df = pd.read_excel(input_file, sheet_name=sheet_name)

    required = {COL_DATE, COL_ITEM, COL_ACTUAL}
    missing = required - set(df.columns)
    if missing:
        raise KeyError(
            f"[{plant}] 입력파일 시트 '{sheet_name}'에 필수 컬럼이 없습니다: {missing}\n"
            f"현재 컬럼: {list(df.columns)}"
        )

    df = df.copy()
    df[COL_DATE] = clean_datetime(df[COL_DATE])
    df[COL_ITEM] = df[COL_ITEM].astype(str).str.strip()
    # raw 시트에 빈 셀이 섞이면 pandas 가 ItemCode 컬럼을 float64 로 추론하여
    # str 변환 시 "260015.0" 처럼 ".0" 이 잔재로 붙는다. 이러면 기존 DB 컬럼명("260015")과
    # 매칭이 안 되어 load_existing_output 이 기존 컬럼을 잘라내고 0 으로 덮어쓰는 버그 발생.
    df[COL_ITEM] = df[COL_ITEM].str.replace(r"\.0$", "", regex=True)
    df[COL_ACTUAL] = clean_numeric(df[COL_ACTUAL])

    df = df[~df[COL_ITEM].str.lower().isin({"", "nan", "none", "nat"})]
    df = df.dropna(subset=[COL_DATE])
    return df


def resolve_target_itemcodes(
    raw_df: pd.DataFrame,
    configured_itemcodes: List[str],
    existing_itemcodes: List[str],
    plant: str,
) -> List[str]:
    """공장별 추적 ItemCode 결정.

    - configured_itemcodes 가 있으면 그대로 사용 (PLANT_ITEMCODES 화이트리스트)
    - 비어있으면 기존 DB 컬럼을 먼저 보존하고 RawData 신규 ItemCode 를 뒤에 추가
    """
    selected = normalize_itemcodes(configured_itemcodes)
    if selected:
        return selected

    existing = normalize_itemcodes(existing_itemcodes)
    raw_itemcodes = normalize_itemcodes(raw_df[COL_ITEM].tolist())

    merged_itemcodes = normalize_itemcodes(existing + raw_itemcodes)
    if not merged_itemcodes:
        raise ValueError(
            f"[{plant}] ItemCode 리스트가 비어 있고 기존 DB/RawData 에도 유효 ItemCode 가 없습니다."
        )

    preserved_only = [code for code in existing if code not in raw_itemcodes]
    new_only = [code for code in raw_itemcodes if code not in existing]
    log(
        f"[{plant}] PLANT_ITEMCODES 미지정 -> 기존 DB {len(existing)}개 보존 + "
        f"RawData 신규 {len(new_only)}개 추가 = {len(merged_itemcodes)}개 추적"
    )
    if preserved_only:
        log(f"[{plant}] RawData 미등장 기존 ItemCode 보존: {len(preserved_only)}개")
    return merged_itemcodes


def load_existing_itemcodes(
    output_file: str,
    sheet_name: Optional[str],
    plant: str,
) -> List[str]:
    """기존 출력 시트의 ItemCode 컬럼명을 헤더에서 읽는다."""
    if not os.path.exists(output_file) or not sheet_name:
        return []

    log(f"[{plant}] 기존 출력 ItemCode 헤더 확인 중... (시트: {sheet_name})")
    header_df = pd.read_excel(output_file, sheet_name=sheet_name, nrows=0)
    if OUT_DATE_COL not in header_df.columns:
        raise KeyError(f"[{plant}] 기존 출력 시트 '{sheet_name}'에 '{OUT_DATE_COL}' 컬럼이 없습니다.")

    raw_columns = [str(col).strip() for col in header_df.columns]
    itemcode_columns = [
        col
        for col in raw_columns
        if col != OUT_DATE_COL and col and not col.lower().startswith("unnamed")
    ]
    return normalize_itemcodes(itemcode_columns)


# =========================================================
# 4) 업데이트용 데이터 생성
# =========================================================

def build_update_df(
    raw_df: pd.DataFrame,
    itemcodes: List[str],
    plant: str,
) -> Tuple[pd.DataFrame, pd.Timestamp, pd.Timestamp]:
    itemcodes = normalize_itemcodes(itemcodes)

    start, end = get_update_period(raw_df)
    log(f"[{plant}] 업데이트 기간: {start.date()} ~ {end.date()}")

    df_period = raw_df[(raw_df[COL_DATE] >= start) & (raw_df[COL_DATE] <= end)].copy()
    df_target = df_period[df_period[COL_ITEM].isin(itemcodes)].copy()

    grouped = df_target.groupby([COL_DATE, COL_ITEM], as_index=False)[COL_ACTUAL].sum()

    if grouped.empty:
        # raw 에 itemcodes 매칭 데이터 0건 → 빈 update_df.
        # merge 단계가 base(기존 DB) 를 그대로 유지하도록 한다.
        update_df = pd.DataFrame(columns=[OUT_DATE_COL] + list(itemcodes))
        log(f"[{plant}] 업데이트용 데이터 없음 (raw 매칭 itemcode 0건) — 기존 DB 유지")
        return update_df, start, end

    # raw 에 등장한 (날짜, ItemCode) 만 wide 변환. 미등장 셀은 NaN 으로 둔다
    # (= merge 단계에서 base 의 기존 값을 그대로 유지). full_dates 로 강제 확장하지 않음.
    pivot = grouped.pivot(
        index=COL_DATE,
        columns=COL_ITEM,
        values=COL_ACTUAL,
    )
    pivot = pivot.reindex(columns=itemcodes)  # 미등장 itemcodes 는 NaN 컬럼

    update_df = pivot.reset_index().rename(columns={COL_DATE: OUT_DATE_COL, "index": OUT_DATE_COL})
    update_df[OUT_DATE_COL] = pd.to_datetime(update_df[OUT_DATE_COL]).dt.normalize()

    for code in itemcodes:
        # NaN 보존 (fillna(0) 제거) — raw 미등장 셀은 그대로 NaN
        update_df[code] = (
            pd.to_numeric(update_df[code], errors="coerce")
            .round(5)
            .astype("float64")
        )

    log(f"[{plant}] 업데이트용 데이터 행 수: {len(update_df)}")
    return update_df, start, end


# =========================================================
# 5) 기존 출력파일 로드
# =========================================================

def load_existing_output(
    output_file: str,
    sheet_name: Optional[str],
    itemcodes: List[str],
    plant: str,
) -> pd.DataFrame:
    itemcodes = normalize_itemcodes(itemcodes)

    if not os.path.exists(output_file):
        log(f"[{plant}] 기존 출력파일 없음 -> 신규 시트 생성")
        return pd.DataFrame(columns=[OUT_DATE_COL] + itemcodes)

    if not sheet_name:
        log(f"[{plant}] 기존 출력 시트 없음 -> 신규 시트 생성")
        return pd.DataFrame(columns=[OUT_DATE_COL] + itemcodes)

    log(f"[{plant}] 기존 출력 시트 읽는 중... (시트: {sheet_name})")
    old_df = pd.read_excel(output_file, sheet_name=sheet_name)

    if OUT_DATE_COL not in old_df.columns:
        raise KeyError(f"[{plant}] 기존 출력 시트 '{sheet_name}'에 '{OUT_DATE_COL}' 컬럼이 없습니다.")

    old_df = old_df.copy()
    old_df[OUT_DATE_COL] = pd.to_datetime(old_df[OUT_DATE_COL], errors="coerce").dt.normalize()
    old_df = old_df.dropna(subset=[OUT_DATE_COL])

    old_df.columns = [OUT_DATE_COL] + [str(c) for c in old_df.columns[1:]]

    for code in itemcodes:
        if code not in old_df.columns:
            old_df[code] = pd.NA  # 신규 컬럼은 NaN — raw 가 채워주기 전엔 "데이터 없음"

    old_df = old_df[[OUT_DATE_COL] + itemcodes]

    for code in itemcodes:
        # NaN 보존 (fillna(0) 제거)
        old_df[code] = (
            pd.to_numeric(old_df[code], errors="coerce")
            .round(5)
            .astype("float64")
        )

    log(f"[{plant}] 기존 출력 데이터 행 수: {len(old_df)}")
    return old_df


# =========================================================
# 6) 병합
# =========================================================

def merge_by_update_window(
    old_df: pd.DataFrame,
    update_df: pd.DataFrame,
    itemcodes: List[str],
    update_start: pd.Timestamp,
    update_end: pd.Timestamp,
) -> pd.DataFrame:
    itemcodes = normalize_itemcodes(itemcodes)

    if old_df.empty:
        result = update_df.copy()
        result = result.sort_values(OUT_DATE_COL).reset_index(drop=True)
        result[OUT_DATE_COL] = pd.to_datetime(result[OUT_DATE_COL]).dt.date
        return result

    base = old_df.copy().set_index(OUT_DATE_COL)

    if update_df.empty:
        # raw 매칭 데이터 0건 — base 그대로 반환
        result = base.reset_index().rename(columns={"index": OUT_DATE_COL})
        result = result.sort_values(OUT_DATE_COL).reset_index(drop=True)
        result[OUT_DATE_COL] = pd.to_datetime(result[OUT_DATE_COL]).dt.date
        return result

    upd = update_df.copy().set_index(OUT_DATE_COL)

    all_dates = base.index.union(upd.index)
    base = base.reindex(all_dates)

    update_mask = (base.index >= update_start) & (base.index <= update_end)
    target_dates = base.index[update_mask]

    upd = upd.reindex(target_dates)  # NaN 유지

    # raw 에 (date, item) 데이터가 있는 셀만 갱신, 없으면 base 기존 값 보존.
    # combine_first(base) 가 upd 의 NaN 자리에 base 값을 채워줌.
    for code in itemcodes:
        if code not in base.columns:
            base[code] = pd.NA
        if code not in upd.columns:
            continue
        base.loc[target_dates, code] = upd[code].combine_first(base.loc[target_dates, code])

    for code in itemcodes:
        # NaN 보존 (fillna(0) 제거)
        base[code] = (
            pd.to_numeric(base[code], errors="coerce")
            .round(5)
            .astype("float64")
        )

    result = base.reset_index().rename(columns={"index": OUT_DATE_COL})
    result = result.sort_values(OUT_DATE_COL).reset_index(drop=True)

    return result


# =========================================================
# 7) 전체 기간 연속 날짜 보정
# =========================================================

def ensure_continuous_dates(df: pd.DataFrame, itemcodes: List[str]) -> pd.DataFrame:
    if df.empty:
        return df

    itemcodes = normalize_itemcodes(itemcodes)

    df = df.copy()
    df[OUT_DATE_COL] = pd.to_datetime(df[OUT_DATE_COL], errors="coerce").dt.normalize()
    df = df.dropna(subset=[OUT_DATE_COL])

    df = df.set_index(OUT_DATE_COL).sort_index()

    full_dates = pd.date_range(df.index.min(), df.index.max(), freq="D")
    df = df.reindex(full_dates)

    for code in itemcodes:
        if code not in df.columns:
            df[code] = pd.NA
        # NaN 보존 (fillna(0) 제거) — raw 에 없었던 (date, item) 셀은 빈 셀로 둔다
        df[code] = pd.to_numeric(df[code], errors="coerce").round(5).astype("float64")

    df = df.reset_index().rename(columns={"index": OUT_DATE_COL})
    df[OUT_DATE_COL] = pd.to_datetime(df[OUT_DATE_COL]).dt.date

    return df


# =========================================================
# 8) 저장
# =========================================================

def format_worksheet(ws):
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.column_dimensions["A"].width = 14
    for col_idx in range(2, ws.max_column + 1):
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[col_letter].width = 12

    ws.freeze_panes = "B2"

    for row in range(2, ws.max_row + 1):
        ws.cell(row=row, column=1).number_format = "yyyy-mm-dd"


def save_to_excel(sheet_df_map: Dict[str, pd.DataFrame], output_file: str):
    ensure_parent_folder(output_file)

    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        for sheet_name, df in sheet_df_map.items():
            df.to_excel(writer, index=False, sheet_name=sheet_name)

    wb = load_workbook(output_file)
    for sheet_name in sheet_df_map:
        ws = wb[sheet_name]
        format_worksheet(ws)

    wb.save(output_file)
    log(f"저장 완료: {output_file}")


# =========================================================
# 9) 메인
# =========================================================

def main():
    log("===== 일일 재공품 실적 정리 시작 (멀티 공장) =====")

    validate_file_exists(INPUT_FILE)
    validate_configuration()

    input_sheet_names = get_excel_sheet_names(INPUT_FILE)
    output_sheet_names = get_excel_sheet_names(OUTPUT_FILE) if os.path.exists(OUTPUT_FILE) else []

    final_by_sheet: Dict[str, pd.DataFrame] = {}
    summary_rows: List[Tuple[str, str, int, pd.Timestamp, pd.Timestamp, int, object, object]] = []

    for plant in PLANTS:
        input_sheet = resolve_input_sheet_name(plant, input_sheet_names)
        output_sheet = OUTPUT_SHEET_BY_PLANT[plant]
        existing_sheet = resolve_existing_output_sheet_name(plant, output_sheet_names)

        raw_df = load_raw_data(INPUT_FILE, input_sheet, plant)
        existing_itemcodes = load_existing_itemcodes(OUTPUT_FILE, existing_sheet, plant)
        itemcodes = resolve_target_itemcodes(
            raw_df=raw_df,
            configured_itemcodes=PLANT_ITEMCODES.get(plant, []),
            existing_itemcodes=existing_itemcodes,
            plant=plant,
        )
        update_df, update_start, update_end = build_update_df(raw_df, itemcodes, plant)
        old_df = load_existing_output(OUTPUT_FILE, existing_sheet, itemcodes, plant)

        merged_df = merge_by_update_window(
            old_df=old_df,
            update_df=update_df,
            itemcodes=itemcodes,
            update_start=update_start,
            update_end=update_end,
        )

        final_df = ensure_continuous_dates(merged_df, itemcodes)
        final_by_sheet[output_sheet] = final_df

        if final_df.empty:
            final_start = None
            final_end = None
        else:
            final_start = final_df[OUT_DATE_COL].min()
            final_end = final_df[OUT_DATE_COL].max()

        summary_rows.append((
            plant,
            output_sheet,
            len(itemcodes),
            update_start,
            update_end,
            len(final_df),
            final_start,
            final_end,
        ))

    backup_output_file(OUTPUT_FILE)
    save_to_excel(final_by_sheet, OUTPUT_FILE)

    print()
    print("===== 실행 결과 =====")
    print(f"- 입력 파일: {INPUT_FILE}")
    print(f"- 출력 파일: {OUTPUT_FILE}")
    print(f"- 처리 공장 수: {len(PLANTS)}")

    for plant, output_sheet, item_count, update_start, update_end, row_count, final_start, final_end in summary_rows:
        print(
            f"- [{plant}] 시트={output_sheet}, "
            f"ItemCode 수={item_count}, "
            f"업데이트 기간={update_start.date()} ~ {update_end.date()}, "
            f"최종 행 수={row_count}, "
            f"최종 기간={final_start} ~ {final_end}"
        )


if __name__ == "__main__":
    main()
