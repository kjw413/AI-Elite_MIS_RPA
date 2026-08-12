# MIS 기간별 제품 생산실적 자동 샘플링 RPA (pywinauto + openpyxl 기반)
"""
사내 MIS '(신)종합정보' 시스템에서
기간별 제품 생산실적(일/월) 데이터를
단일 Raw 파일(E:/DB_MIS/RawDB_생산실적.xlsx)의 카테고리 시트별로 자동 조회하여
각 시트 1행에 기간 마커(__PERIOD__)를 적고 2행부터 그리드를 통째로 붙여넣고,
이어서 그 Raw 파일을 합쳐 DB_생산실적.xlsx 까지 자동 생성한다.
(재공품 RawDB_재공품.xlsx / DB_재공품.xlsx 2-파일 구조와 통일)

시트명 규칙: F{공장}_{카테고리}[_{접미사}]
  예) F10_냉동, F20_상온_MY, F30_냉장_FM

업무 절차:
  1. MIS 앱 연결 (pywinauto UIA backend)
  2. RawDB_생산실적.xlsx 시트명 스캔 → (sheet_name, factory, category) 목록 생성
  3. 공장별로 묶어 순회 (공장 드롭다운 변경 최소화)
     a. ORG 드롭다운 공장 선택 (공장이 바뀔 때만)
     b. 기본 조회 = D-1가 속한 월의 1일 ~ D-1
        직전 수집일 이후 누락이 있으면 누락 구간을 월별로 먼저 조회·통합
     c. category1 드롭다운 선택
     d. 항목구분 드롭다운 "중량" 선택 (최초 1회)
     e. 조회 + 로딩 대기
     f. 그리드 좌상단 헤더 셀(복사 버튼) 클릭 → 클립보드 복사
     g. 확인 팝업 Enter
     h. RawDB_생산실적.xlsx 의 해당 시트에 1행 기간 마커 + 2행부터 그리드 통째 쓰기
  4. (자동 연결) RawDB_생산실적.xlsx → DB_생산실적.xlsx 통합
     - app.services.production_dw_service.build_dataset(raw_path=...) 호출
     - 출력: 공장별 wide 시트 + 제품마스터 + 계획 + daily(내부 호환)
     - DB(production_daily) UPSERT는 다음 앱 기동 시
       production_dw_sync_service 가 파일 mtime 변화를 보고 자동 처리

Usage:
  python production_daily_rpa.py              # D-1 기준 + 이전 누락 자동 복구 + DW 통합
  python production_daily_rpa.py --date 2026-05-14
  python production_daily_rpa.py --dry-run    # MIS 조회만, Excel/DW 미기록
  python production_daily_rpa.py --skip-dw-build   # Raw 샘플링만, DW 통합 생략
  python production_daily_rpa.py --dw-output "E:/tmp/test.xlsx"
"""

import sys
import time
import os
import json
import hashlib
import shutil
import logging
import argparse
import re
import tempfile
from datetime import date, datetime, timedelta
from pathlib import Path

import pandas as pd
from pywinauto import Application
from pywinauto.keyboard import send_keys
from pywinauto.timings import Timings

# pywinauto 내부 click/keys 대기 시간 단축 (MIS는 즉시 반응한다는 가정)
#   after_clickinput_wait : click_input 직후 자동 sleep (default 0.09s → 0.01s)
#   after_setfocus_wait   : set_focus 직후 자동 sleep (default 0.06s → 0.01s)
#   after_sendkeys_key_wait : 각 키 입력 사이 (default 0.01s → 0.001s)
Timings.after_clickinput_wait = 0.01
Timings.after_setfocus_wait = 0.01
Timings.after_sendkeys_key_wait = 0.001

# 프로젝트 루트(app/...) import 보장. tools/AI-Elite-MIS_RPA/file.py → 2단계 위가 루트.
_PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(_PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(_PROJECT_ROOT))

from _common import (
    fast_click,
    find_mis_window,
    get_clipboard_sequence,
    get_clipboard_text,
    parse_clipboard_rows,
    paste_to_sheet,
    sampled_db_path,
    wait_for_clipboard_change,
)

# ---------------------------------------------------------------------------
# 로깅 — module-import 시점에는 핸들러를 부착하지 않는다.
# 단독 실행(__main__) 시 _setup_logging() 이 핸들러를 부착하고,
# 오케스트레이터(run_all_rpa) 가 import 해서 쓸 때는 root 로거의 기존 핸들러로
# propagate 된다 (오케스트레이터가 자체 핸들러를 미리 설정해 둠).
# ---------------------------------------------------------------------------
LOG_DIR = Path(__file__).parent / "logs"
LOG_DIR.mkdir(exist_ok=True)
log = logging.getLogger(__name__)


def _setup_logging() -> None:
    if logging.getLogger().handlers:
        return
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(message)s",
        handlers=[
            logging.StreamHandler(sys.stdout),
            logging.FileHandler(
                LOG_DIR / f"production_rpa_{datetime.now():%Y%m%d_%H%M%S}.log",
                encoding="utf-8",
            ),
        ],
    )

# ---------------------------------------------------------------------------
# 상수 정의
# ---------------------------------------------------------------------------
RAW_FILE = sampled_db_path("RawDB_생산실적.xlsx", "PRODUCTION_RAW_XLSX")
ITEM_TYPE = "중량"

# 시트명 규칙: F{공장}_{카테고리}[_{접미사}]  (확장자 없는 시트명에 매칭)
SHEETNAME_RE = re.compile(r"^(F\d+)_(냉동|냉장|상온)(?:_(.+))?$")

# Raw 파일에서 데이터 시트가 아닌(스캔 제외) 메타 시트
NON_TARGET_SHEETS = {"제품마스터", "계획", "daily"}

# 기간 마커 — 빌드(production_dw_service)가 시트 1행에서 (year, month, 윈도우) 산출.
# DW wide 그리드(01일~31일)에는 年月이 없으므로 RPA가 직접 기입한다.
PERIOD_MARKER = "__PERIOD__"

# 대기 시간 기본값 (production_coords.json의 "wait" 값으로 덮어씌워짐)
WAIT_SHORT = 0.05        # 클릭/타이핑 후 미세 대기 (MIS는 즉시 반응)
WAIT_MEDIUM = 0.2        # 클립보드 fill / 필드 클릭 후 약간 더 긴 대기
WAIT_DROPDOWN = 0.01     # 드롭다운 펼침 후 항목 클릭 전 대기
WAIT_SCREEN_LOAD = 1.0   # 사이드바 메뉴 클릭 → MIS 화면 전환 로딩
WAIT_QUERY_LOAD = 1.5    # 조회 버튼 → 그리드 데이터 로딩
WAIT_COPY_CONFIRM = 0.4  # 복사 버튼 클릭 후 확인 팝업이 포커스를 받을 때까지 대기
GRID_VERIFY_ATTEMPTS = 3  # 직전 그리드 재복사 감지 시 공장·카테고리 재선택 횟수


# ---------------------------------------------------------------------------
# 폴더 스캔 → 작업 대상 목록 생성
# ---------------------------------------------------------------------------
def discover_targets(raw_file: str):
    """
    단일 Raw 파일의 카테고리 시트명을 스캔해서 (sheet_name, factory, category, suffix) 추출.
    같은 공장끼리 인접하도록 정렬한다 (공장 드롭다운 변경 최소화).
    """
    import openpyxl

    if not os.path.exists(raw_file):
        log.error(f"Raw 파일이 없습니다: {raw_file}")
        log.error("먼저 tools/scripts/migrate_production_to_new_structure.py 로 "
                  "RawDB_생산실적.xlsx 를 생성하세요.")
        return []

    wb = openpyxl.load_workbook(raw_file, read_only=True)
    sheet_names = list(wb.sheetnames)
    wb.close()

    targets = []
    for sn in sheet_names:
        if sn in NON_TARGET_SHEETS:
            continue
        m = SHEETNAME_RE.match(sn)
        if not m:
            log.warning(f"시트명 규칙 미일치 (스킵): {sn}")
            continue
        factory, category, suffix = m.group(1), m.group(2), m.group(3) or ""
        targets.append({
            "sheet_name": sn,
            "factory": factory,
            "category": category,
            "suffix": suffix,
        })

    # 공장별 → 카테고리별 정렬 (공장 변경 최소화)
    targets.sort(key=lambda t: (t["factory"], t["category"], t["suffix"]))
    log.info(f"대상 시트 {len(targets)}건 발견:")
    for t in targets:
        log.info(f"  - {t['sheet_name']:20s} (factory={t['factory']}, "
                 f"category={t['category']}, suffix={t['suffix'] or '-'})")
    return targets


def _as_date(value) -> date | None:
    """Excel 날짜 셀 값을 date 로 정규화한다."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%y-%m-%d", "%y/%m/%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def load_collected_dates_by_sheet(
    output_path: str | Path,
    sheet_names: list[str],
) -> dict[str, set[date]]:
    """DB_생산실적의 품목군 시트별 수집 완료 날짜를 읽는다.

    각 품목군 wide 시트에 날짜 행이 있고 품목 값이 하나 이상 있을 때만 수집
    완료로 본다. 값이 모두 빈 날짜 행은 출력 피벗이 만든 자리표시자일 수 있다.
    파일을 아직 만들지 않았거나 읽을 수 없으면 빈 이력을 반환해 기존의 현재 월
    조회만 수행하게 한다.
    """
    path = Path(output_path)
    if not path.exists():
        return {}

    import openpyxl

    try:
        wb = openpyxl.load_workbook(
            path, read_only=True, data_only=True, keep_links=False
        )
    except Exception as exc:
        log.warning(f"생산실적 수집 이력 확인 실패({path}): {exc}")
        return {}

    result: dict[str, set[date]] = {}
    try:
        for sheet_name in sheet_names:
            dates: set[date] = set()
            result[sheet_name] = dates
            if sheet_name not in wb.sheetnames:
                continue

            rows = wb[sheet_name].iter_rows(values_only=True)
            try:
                header = next(rows)
            except StopIteration:
                continue

            date_index = next(
                (
                    idx
                    for idx, value in enumerate(header)
                    if str(value).strip() == "날짜"
                ),
                None,
            )
            if date_index is None:
                continue

            for row in rows:
                if date_index >= len(row):
                    continue
                parsed = _as_date(row[date_index])
                has_actual_cell = any(
                    value is not None
                    for idx, value in enumerate(row)
                    if idx != date_index
                )
                if parsed is not None and has_actual_cell:
                    dates.add(parsed)
    finally:
        wb.close()

    return result


def plan_collection_periods(
    end_date: date,
    collected_dates_by_sheet: dict[str, set[date]],
) -> list[tuple[date, date]]:
    """마지막 수집일 이후 누락 구간과 현재 월 조회 구간을 월별로 계획한다.

    생산실적 그리드의 일자 열은 01일처럼 연월 정보가 없으므로 한 조회가
    두 달에 걸치면 안 된다. 현재 월 시작일의 직전 날짜가 어느 대상 시트에서든
    빠졌다면, 해당 시트들의 마지막 수집일 다음 날 중 가장 이른 날짜부터 직전
    월 말일까지 별도 조회한다.
    """
    current_start = end_date.replace(day=1)
    current_period = (current_start, end_date)
    if not collected_dates_by_sheet:
        return [current_period]

    previous_day = current_start - timedelta(days=1)
    if all(
        previous_day in dates
        for dates in collected_dates_by_sheet.values()
    ):
        return [current_period]

    if not any(collected_dates_by_sheet.values()):
        return [current_period]

    previous_month_start = previous_day.replace(day=1)
    missing_starts: list[date] = []
    for dates in collected_dates_by_sheet.values():
        if previous_day in dates:
            continue
        collected_in_previous_month = [
            day
            for day in dates
            if previous_month_start <= day <= previous_day
        ]
        if collected_in_previous_month:
            missing_starts.append(max(collected_in_previous_month) + timedelta(days=1))
        else:
            missing_starts.append(previous_month_start)

    if not missing_starts:
        return [current_period]

    recovery_start = min(missing_starts)
    return [(recovery_start, previous_day), current_period]


# ---------------------------------------------------------------------------
# MIS RPA 클래스
# ---------------------------------------------------------------------------
class MISProductionRPA:
    """MIS 기간별 제품 생산실적 자동 샘플링 RPA"""

    def __init__(
        self,
        ref_date: str = None,
        dry_run: bool = False,
        build_dw: bool = True,
        dw_output: str | None = None,
    ):
        self.auto_recover_missing_dates = ref_date is None
        if ref_date is None:
            d = datetime.now() - timedelta(days=1)
        else:
            d = datetime.strptime(ref_date, "%Y-%m-%d")

        self.requested_end_date = d.date()
        self._set_collection_period(
            self.requested_end_date.replace(day=1),
            self.requested_end_date,
        )

        self.dry_run = dry_run
        self.build_dw = build_dw
        self.dw_output = dw_output  # None → production_dw_service.DEFAULT_OUTPUT_PATH
        self.coords = self._load_coords()
        self._pending_raw_snapshots: list[tuple[str, str, bytes]] = []
        self.collection_success_count = 0
        # 좌표 선택/화면 로딩 실패로 이전 대상의 그리드를 다시 복사한 사례.
        # (시작일, 종료일, 대상 시트, 실제로 복사된 이전 시트)
        self.duplicate_grids: list[tuple[str, str, str, str]] = []

        self.app = None
        self.main_window = None
        log.info("=== MIS 생산실적 RPA 초기화 ===")
        log.info(f"  기준일자: {self.start_date} ~ {self.end_date}")
        log.info(f"  시트명  : {self.sheet_name}")
        log.info(f"  Dry-run : {self.dry_run}")
        log.info(f"  DW 통합 : {'실행' if self.build_dw else '생략'}")

    def _set_collection_period(self, start: date, end: date) -> None:
        """현재 MIS 조회 기간을 같은 월 안의 구간으로 설정한다."""
        if start > end:
            raise ValueError(f"조회 기간 오류: {start} > {end}")
        if (start.year, start.month) != (end.year, end.month):
            raise ValueError(f"생산실적 조회는 월을 넘길 수 없습니다: {start} ~ {end}")

        self.start_date_obj = datetime(start.year, start.month, start.day)
        self.end_date_obj = datetime(end.year, end.month, end.day)
        self.start_date = start.strftime("%Y-%m-%d")
        self.end_date = end.strftime("%Y-%m-%d")
        self.sheet_name = start.strftime("%Y-%m")

    def _plan_collection_periods(self, targets: list[dict]) -> list[tuple[date, date]]:
        """자동 실행이면 시트별 마지막 수집일을 확인해 누락 구간을 추가한다."""
        current_period = (
            self.requested_end_date.replace(day=1),
            self.requested_end_date,
        )
        # --date 는 명시된 월을 재수집하는 용도이므로 자동 누락 복구를 섞지 않는다.
        if not self.auto_recover_missing_dates:
            return [current_period]

        output_path = self.dw_output or sampled_db_path(
            "DB_생산실적.xlsx", "PRODUCTION_DW_XLSX"
        )
        target_sheets = [target["sheet_name"] for target in targets]
        collected = load_collected_dates_by_sheet(output_path, target_sheets)
        periods = plan_collection_periods(self.requested_end_date, collected)

        if len(periods) > 1:
            recovery_start, recovery_end = periods[0]
            missing_sheets = [
                name
                for name, dates in collected.items()
                if recovery_end not in dates
            ]
            log.warning(
                f"이전 수집 누락 감지: {recovery_start} ~ {recovery_end} "
                f"(대상 시트 {len(missing_sheets)}개) → 현재 월보다 먼저 복구"
            )
        elif collected and any(collected.values()):
            previous_day = current_period[0] - timedelta(days=1)
            log.info(f"이전 수집일 확인 완료: {previous_day} 누락 없음")
        else:
            log.info("기존 수집 이력이 없어 현재 월 기본 조회만 수행합니다.")
        return periods

    # -----------------------------------------------------------------------
    # 좌표 설정 로드
    # -----------------------------------------------------------------------
    def _load_coords(self):
        global WAIT_SHORT, WAIT_MEDIUM, WAIT_DROPDOWN, WAIT_SCREEN_LOAD, WAIT_QUERY_LOAD
        global WAIT_COPY_CONFIRM

        coord_path = os.path.join(os.path.dirname(__file__),
                                   "production_coords.json")
        try:
            with open(coord_path, "r", encoding="utf-8") as f:
                config = json.load(f)
                log.info(f"설정 파일 로드: {coord_path}")
                wait = config.get("wait", {})
                WAIT_SHORT = wait.get("short", WAIT_SHORT)
                WAIT_MEDIUM = wait.get("medium", WAIT_MEDIUM)
                WAIT_DROPDOWN = wait.get("dropdown", WAIT_DROPDOWN)
                WAIT_SCREEN_LOAD = wait.get("screen_load", WAIT_SCREEN_LOAD)
                WAIT_QUERY_LOAD = wait.get("query_load", WAIT_QUERY_LOAD)
                WAIT_COPY_CONFIRM = wait.get("copy_confirm", WAIT_COPY_CONFIRM)
                log.info(f"  대기시간: short={WAIT_SHORT}s, medium={WAIT_MEDIUM}s, "
                         f"dropdown={WAIT_DROPDOWN}s, screen={WAIT_SCREEN_LOAD}s, "
                         f"query={WAIT_QUERY_LOAD}s, copy_confirm={WAIT_COPY_CONFIRM}s")
                return config.get("coords", {})
        except Exception as e:
            log.warning(f"설정 파일을 읽을 수 없습니다 ({e}). 기본값 사용.")
            return {}

    # -----------------------------------------------------------------------
    # MIS 연결
    # -----------------------------------------------------------------------
    def attach_existing_window(self, app, main_window) -> None:
        """오케스트레이터가 이미 연결한 MIS 윈도우를 주입.
        connect_mis() 가 호출되어도 재연결을 생략한다."""
        self.app = app
        self.main_window = main_window

    def connect_mis(self):
        if self.main_window is not None:
            log.info("MIS 연결 재사용 (이전 단계에서 연결됨)")
            return
        log.info("MIS 앱 연결 중...")
        # Win32 로 HWND 를 먼저 찾고 handle 로 connect 한다. UIA title_re 스캔은
        # 바탕화면 전체 트리를 순회해 수 초~20초+ 까지 들쭉날쭉 걸리지만, 핸들
        # 연결은 트리 스캔이 없어 ~10ms 로 끝난다.
        hwnd, title = find_mis_window("(신)종합정보")
        if not hwnd:
            log.error("MIS (신)종합정보 창을 찾을 수 없습니다.")
            log.error("MIS (신)종합정보를 먼저 실행해주세요.")
            raise SystemExit(1)
        try:
            self.app = Application(backend="uia").connect(handle=hwnd)
            self.main_window = self.app.window(handle=hwnd)
            log.info(f"MIS 연결 성공: {title}")
        except Exception as e:
            log.error(f"MIS 앱 연결 실패: {e}")
            log.error("MIS (신)종합정보를 먼저 실행해주세요.")
            raise SystemExit(1)

    # -----------------------------------------------------------------------
    # 메뉴 진입
    # -----------------------------------------------------------------------
    def navigate_to_production_screen(self):
        log.info("'기간별 제품 생산실적(일/월)' 화면으로 이동 중...")
        # 좌표 기반 더블클릭 — UIA child_window(TreeItem) 검색은 트리 전체를
        # 순회해 수 초~20초+ 까지 들쭉날쭉 걸리므로 좌표 클릭으로 대체.
        x, y = self.coords.get("tree_menu", [171, 211])
        fast_click(self.main_window, x, y, double=True)
        log.info(f"  트리메뉴 더블클릭 ({x}, {y})")
        time.sleep(WAIT_SCREEN_LOAD)
        log.info("화면 이동 완료")

    # -----------------------------------------------------------------------
    # 공장 선택
    # -----------------------------------------------------------------------
    def select_factory(self, org_code: str):
        log.info(f"공장 선택: {org_code}")
        x, y = self.coords.get("factory_dropdown", [417, 106])
        fast_click(self.main_window, x, y)
        log.info(f"  드롭다운 클릭 ({x}, {y})")
        time.sleep(WAIT_DROPDOWN)

        factory_list = self.coords.get("factory_list", {})
        item_y = factory_list.get(org_code)
        if item_y is None:
            log.error(f"  {org_code} 좌표 미정의 → JSON에 추가 필요")
            raise RuntimeError(f"factory_list[{org_code}] 좌표 없음")

        fast_click(self.main_window, x, item_y)
        log.info(f"  공장 항목 클릭 ({x}, {item_y})")
        time.sleep(WAIT_SHORT)

    # -----------------------------------------------------------------------
    # 기준일자 설정
    # -----------------------------------------------------------------------
    def set_date_range(self):
        """
        시작일 필드 클릭 → 입력 → TAB(자동 종료일 포커스) → 종료일 입력 → ENTER(확정).
        MIS 날짜 필드는 YYYY-MM-DD 형식을 요구한다.
        """
        log.info(f"기준일자 설정: {self.start_date} ~ {self.end_date}")

        # 시작일 클릭 + 입력
        x, y = self.coords.get("start_date_field", [609, 108])
        fast_click(self.main_window, x, y)
        log.info(f"  시작일 클릭 ({x}, {y})")
        time.sleep(WAIT_MEDIUM)
        send_keys("^a")
        time.sleep(WAIT_SHORT)
        send_keys(self.start_date, with_spaces=True)
        log.info(f"  시작일 입력: {self.start_date}")
        time.sleep(WAIT_SHORT)

        # TAB → 종료일로 자동 포커스 이동
        send_keys("{TAB}")
        time.sleep(WAIT_SHORT)

        # 종료일 입력 (클릭 불필요)
        send_keys("^a")
        time.sleep(WAIT_SHORT)
        send_keys(self.end_date, with_spaces=True)
        log.info(f"  종료일 입력: {self.end_date}")
        time.sleep(WAIT_SHORT)

        # ENTER → 확정
        send_keys("{ENTER}")
        time.sleep(WAIT_SHORT)

    # -----------------------------------------------------------------------
    # category1 선택
    # -----------------------------------------------------------------------
    def select_category(self, category: str):
        log.info(f"카테고리1 선택: {category}")
        x, y = self.coords.get("category1_dropdown", [830, 107])
        fast_click(self.main_window, x, y)
        time.sleep(WAIT_DROPDOWN)

        cat_list = self.coords.get("category1_list", {})
        item_y = cat_list.get(category)
        if item_y is None:
            log.error(f"  {category} 좌표 미정의")
            raise RuntimeError(f"category1_list[{category}] 좌표 없음")

        fast_click(self.main_window, x, item_y)
        log.info(f"  카테고리 클릭 ({x}, {item_y})")
        time.sleep(WAIT_SHORT)

    # -----------------------------------------------------------------------
    # 항목구분 선택 (중량 고정)
    # -----------------------------------------------------------------------
    def select_item_type(self):
        log.info(f"항목구분 선택: {ITEM_TYPE}")
        x, y = self.coords.get("item_type_dropdown", [970, 107])
        fast_click(self.main_window, x, y)
        time.sleep(WAIT_DROPDOWN)

        item_list = self.coords.get("item_type_list", {})
        item_y = item_list.get(ITEM_TYPE)
        if item_y is None:
            log.error(f"  {ITEM_TYPE} 좌표 미정의")
            raise RuntimeError(f"item_type_list[{ITEM_TYPE}] 좌표 없음")

        fast_click(self.main_window, x, item_y)
        log.info(f"  항목구분 클릭 ({x}, {item_y})")
        time.sleep(WAIT_SHORT)

    # -----------------------------------------------------------------------
    # 조회
    # -----------------------------------------------------------------------
    def click_query(self):
        log.info("조회 버튼 클릭...")
        x, y = self.coords.get("query_button", [328, 66])
        fast_click(self.main_window, x, y)
        log.info(f"  조회 클릭 ({x}, {y})")
        log.info("  데이터 로딩 대기 중...")
        time.sleep(WAIT_QUERY_LOAD)
        log.info("  조회 완료")

    # -----------------------------------------------------------------------
    # 그리드 복사
    # -----------------------------------------------------------------------
    def copy_grid_data(self) -> str:
        # 로딩 도중 클릭이 흡수되어 클립보드가 갱신되지 않는 경우 방어 —
        # ClipboardSequenceNumber 폴링으로 실제 갱신을 검증하고, 안 바뀌면 재클릭.
        log.info("그리드 좌상단 복사 버튼 클릭...")
        x, y = self.coords.get("copy_button", [332, 200])
        log.info(f"  복사 버튼 클릭 ({x}, {y})")

        clipboard_text = ""
        for attempt in range(1, 4):  # 최대 3회 시도
            seq_before = get_clipboard_sequence()
            fast_click(self.main_window, x, y)
            changed = wait_for_clipboard_change(seq_before, timeout=2.0)
            if changed:
                if attempt > 1:
                    log.info(f"  클립보드 갱신 확인 (재시도 {attempt}/3)")
                clipboard_text = get_clipboard_text()
                self._handle_copy_confirm_dialog()
                break
            self._handle_copy_confirm_dialog(use_ok_click=False)
            log.warning(f"  복사 시도 {attempt}/3 — 2s 내 클립보드 변화 없음")
            time.sleep(WAIT_MEDIUM)
        else:
            log.warning("  클립보드 변경 감지 실패 → 기존 클립보드 그대로 읽기")
            time.sleep(WAIT_MEDIUM)
            clipboard_text = get_clipboard_text()

        if not clipboard_text.strip():
            log.info("  win32clipboard 실패 → pandas.read_clipboard 시도")
            try:
                df = pd.read_clipboard(sep="\t", header=None)
                clipboard_text = df.to_csv(sep="\t", index=False, header=False)
                log.info(f"  pandas 클립보드 읽기 성공: {df.shape}")
            except Exception as e:
                log.warning(f"  pandas.read_clipboard 실패: {e}")

        if not clipboard_text.strip():
            log.warning("클립보드가 비어 있습니다!")
        else:
            lines = clipboard_text.strip().split("\n")
            log.info(f"  클립보드 데이터: {len(lines)}행")
        return clipboard_text

    def _handle_copy_confirm_dialog(self, use_ok_click: bool = True):
        time.sleep(WAIT_COPY_CONFIRM)
        ok_xy = self.coords.get("confirm_popup_ok") if use_ok_click else None
        if ok_xy:
            fast_click(self.main_window, ok_xy[0], ok_xy[1])
            log.info(f"  확인 팝업 닫기 완료 (OK 클릭, {ok_xy[0]}, {ok_xy[1]})")
        else:
            send_keys("{ENTER}")
            action = "완료" if use_ok_click else "시도"
            log.info(f"  확인 팝업 닫기 {action} (Enter)")
        time.sleep(WAIT_SHORT)

    @staticmethod
    def _grid_fingerprint(clipboard_text: str) -> str:
        """줄바꿈 차이를 제거한 MIS 그리드 내용 지문을 만든다."""
        normalized = (
            clipboard_text.strip()
            .replace("\r\n", "\n")
            .replace("\r", "\n")
        )
        return hashlib.sha1(normalized.encode("utf-8", "replace")).hexdigest()

    def _copy_verified_grid(
        self,
        target: dict,
        fingerprints: dict[str, str],
    ) -> tuple[str, str | None, bool]:
        """앞서 수집한 그리드와 같으면 재조회하고, 계속 같으면 적재를 거부한다.

        ClipboardSequenceNumber는 복사 동작 자체만 검증할 수 있다. MIS 조회가 아직
        끝나지 않았거나 공장 선택이 적용되지 않으면 새 클립보드에도 직전 그리드가
        정상 복사될 수 있으므로 내용 지문까지 비교한다.
        """
        for attempt in range(1, GRID_VERIFY_ATTEMPTS + 1):
            clipboard_text = self.copy_grid_data()
            if not clipboard_text.strip():
                return "", None, False

            fingerprint = self._grid_fingerprint(clipboard_text)
            twin = fingerprints.get(fingerprint)
            if twin is None:
                return clipboard_text, fingerprint, False

            if attempt >= GRID_VERIFY_ATTEMPTS:
                log.error(
                    f"  {target['sheet_name']}: 그리드가 '{twin}' 과(와) 완전히 동일 "
                    f"({attempt}/{GRID_VERIFY_ATTEMPTS}) → 적재하지 않습니다."
                )
                self.duplicate_grids.append(
                    (self.start_date, self.end_date, target["sheet_name"], twin)
                )
                return "", None, True

            log.warning(
                f"  {target['sheet_name']}: '{twin}' 그리드가 다시 복사됨 "
                f"({attempt}/{GRID_VERIFY_ATTEMPTS}) → 공장·카테고리 재선택 후 재조회"
            )
            self.select_factory(target["factory"])
            time.sleep(WAIT_SHORT)
            self.select_category(target["category"])
            self.click_query()

        return "", None, False

    # -----------------------------------------------------------------------
    # Excel 백업 (단일 Raw 파일)
    # -----------------------------------------------------------------------
    def backup_raw_file(self):
        if self.dry_run or not os.path.exists(RAW_FILE):
            return
        backup_root = os.path.join(os.path.dirname(RAW_FILE), "backup")
        os.makedirs(backup_root, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = os.path.join(backup_root, f"RawDB_생산실적_{timestamp}.xlsx")
        try:
            shutil.copy2(RAW_FILE, backup_path)
            log.info(f"백업 생성: {backup_path}")
        except Exception as e:
            log.warning(f"  백업 실패: {e}")

    # -----------------------------------------------------------------------
    # DW 통합 (Raw_생산실적/*.xlsx → DB_생산실적.xlsx)
    # -----------------------------------------------------------------------
    def consolidate_to_dw(self, raw_path: str | Path | None = None) -> bool:
        """
        RawDB_생산실적.xlsx 의 카테고리 시트들을 합쳐
        DB_생산실적.xlsx (공장별 wide + 제품마스터 + 계획 + daily)로 저장한다.

        production_dw_service.build_dataset 을 그대로 호출 — 빌드 로직은 한 곳에서 관리.
        DB UPSERT(production_daily)는 app 기동 시 production_dw_sync_service 가
        파일 mtime 변화를 감지해 자동 수행하므로 여기선 파일까지만.
        """
        raw_source = str(raw_path or RAW_FILE)
        log.info("=" * 60)
        log.info("DW 통합 단계 시작 (RawDB_생산실적 → DB_생산실적.xlsx)")
        log.info(f"  입력 Raw: {raw_source}")
        log.info("=" * 60)
        try:
            from production_builder import (
                DEFAULT_OUTPUT_PATH,
                build_dataset,
                validate_subcategory_coverage,
            )
        except Exception as exc:
            log.error(f"production_dw_service import 실패: {exc}", exc_info=True)
            return False

        out_path = self.dw_output or str(DEFAULT_OUTPUT_PATH)
        try:
            t0 = datetime.now()
            df, saved_path = build_dataset(raw_path=raw_source, output_path=out_path)
            dt = (datetime.now() - t0).total_seconds()
        except Exception as exc:
            log.error(f"DW 통합 실패: {exc}", exc_info=True)
            return False

        if df.empty:
            log.warning(f"DW 통합 결과가 비어 있음 (저장 경로: {saved_path})")
            return False

        log.info(f"DW 통합 완료 — {dt:.1f}s, {len(df):,}행")
        log.info(f"  날짜 범위 : {df['date'].min()} ~ {df['date'].max()}")
        log.info(f"  공장      : {sorted(df['factory'].unique())}")
        log.info(f"  category1 : {sorted(df['category1'].unique())}")
        c2_vals = sorted([s for s in df['category2'].dropna().unique()])
        log.info(f"  category2 : {c2_vals if c2_vals else '(없음)'}")
        log.info(f"  unique items: {df['item_code'].nunique():,}")
        log.info(f"  출력 파일 : {saved_path}")

        # category2 커버리지 — 미분류가 크면 키워드 보강 필요
        try:
            cov = validate_subcategory_coverage(df)
            if cov["total_actual"] > 0 and not cov["is_complete"]:
                log.warning(
                    f"  category2 미분류 {len(cov['unclassified_items'])}품목, "
                    f"실적 {cov['unclassified_actual']:,.0f} "
                    f"({cov['unclassified_pct']:.1f}%) — 'category2_unclassified' 시트 확인"
                )
        except Exception as exc:
            log.warning(f"  category2 커버리지 검증 스킵: {exc}")

        log.info("DB UPSERT는 다음 앱 기동 시 production_dw_sync_service 가 자동 처리합니다.")
        return True

    def _collect_period(self, targets: list[dict]) -> tuple[int, int, int]:
        """현재 설정된 한 달 내 기간을 모든 대상 시트에서 수집한다."""
        period_marker = [PERIOD_MARKER, self.start_date, self.end_date]
        current_factory = None
        # 같은 기간에 앞서 정상 수집한 그리드 지문. 직전 화면 재복사를 검출한다.
        fingerprints: dict[str, str] = {}

        total_rows = 0
        success = 0
        failed = 0

        for target in targets:
            log.info("=" * 50)
            log.info(
                f"▶ {target['sheet_name']} (factory={target['factory']}, "
                f"category={target['category']}, suffix={target['suffix'] or '-'})"
            )
            log.info("=" * 50)

            try:
                if target["factory"] != current_factory:
                    self.select_factory(target["factory"])
                    time.sleep(WAIT_SHORT)
                    current_factory = target["factory"]

                self.select_category(target["category"])
                self.click_query()
                (
                    clipboard_text,
                    fingerprint,
                    duplicate_rejected,
                ) = self._copy_verified_grid(
                    target, fingerprints
                )

                if not clipboard_text.strip():
                    if duplicate_rejected:
                        log.warning("  중복 그리드 거부 → 기존 시트 유지")
                    else:
                        log.warning("  데이터 없음 → 스킵")
                    failed += 1
                    continue

                rows = parse_clipboard_rows(clipboard_text)
                if not rows:
                    log.warning("  파싱 결과 없음 → 스킵")
                    failed += 1
                    continue

                if fingerprint is not None:
                    fingerprints[fingerprint] = target["sheet_name"]

                rows_with_marker = [period_marker] + rows
                if self.dry_run:
                    log.info(
                        f"  [DRY-RUN] {target['sheet_name']} "
                        f"({self.start_date}~{self.end_date}): "
                        f"{len(rows)}행 × {len(rows[0])}열 (+마커, 기록 안함)"
                    )
                    success += 1
                else:
                    written = paste_to_sheet(
                        RAW_FILE, target["sheet_name"], rows_with_marker
                    )
                    total_rows += written
                    if written > 0:
                        success += 1
                    else:
                        failed += 1

                self.main_window.set_focus()
                time.sleep(WAIT_SHORT)

            except Exception as exc:
                log.error(f"  처리 중 오류: {exc}", exc_info=True)
                failed += 1
                try:
                    self.main_window.set_focus()
                except Exception:
                    pass

        return success, failed, total_rows

    # -----------------------------------------------------------------------
    # 수집 단계
    # -----------------------------------------------------------------------
    def collect(self) -> bool:
        log.info("=" * 60)
        log.info("MIS 생산실적 수집 단계 시작")
        log.info("=" * 60)

        self._pending_raw_snapshots = []
        self.collection_success_count = 0
        self.duplicate_grids = []

        targets = discover_targets(RAW_FILE)
        if not targets:
            log.error(f"작업 대상 시트가 없습니다: {RAW_FILE}")
            return False

        periods = self._plan_collection_periods(targets)

        self.connect_mis()
        self.main_window.set_focus()
        time.sleep(WAIT_MEDIUM)

        self.navigate_to_production_screen()
        self.backup_raw_file()

        total_rows = 0
        success = 0
        failed = 0

        for period_index, (period_start, period_end) in enumerate(periods, start=1):
            self._set_collection_period(period_start, period_end)
            log.info("=" * 60)
            log.info(
                f"수집 기간 [{period_index}/{len(periods)}]: "
                f"{self.start_date} ~ {self.end_date}"
            )
            log.info("=" * 60)
            self.set_date_range()
            if period_index == 1:
                self.select_item_type()

            period_success, period_failed, period_rows = self._collect_period(targets)
            success += period_success
            failed += period_failed
            total_rows += period_rows

            if (
                not self.dry_run
                and period_success > 0
            ):
                self._capture_raw_snapshot()

        log.info("=" * 60)
        if self.dry_run:
            log.info(f"DRY-RUN 완료 (성공 {success} / 실패 {failed})")
        else:
            log.info(f"RPA 완료: 성공 {success} / 실패 {failed} / 총 {total_rows}행 기록")
        log.info("=" * 60)

        self.collection_success_count = success
        self.report_duplicate_grids()
        return success > 0

    def report_duplicate_grids(self) -> bool:
        """이전 대상 그리드가 반복 복사되어 적재를 거부한 시트를 요약한다."""
        if not self.duplicate_grids:
            return False
        log.error(
            f"⚠ 이전 생산 그리드 재복사 {len(self.duplicate_grids)}건 "
            "— 해당 시트는 기존 RawDB 내용을 보존했습니다:"
        )
        for start, end, sheet_name, twin in self.duplicate_grids:
            log.error(f"    [{start}~{end}] {sheet_name} = {twin} 의 그리드")
        log.error(
            "  → production_coords.json의 factory/category 좌표와 "
            "query_load 대기를 확인하세요."
        )
        return True

    def _capture_raw_snapshot(self) -> None:
        """현재 월 Raw 파일을 메모리에 보존해 다음 월 수집의 덮어쓰기를 막는다."""
        raw_path = Path(RAW_FILE)
        if not raw_path.exists():
            raise FileNotFoundError(f"생산실적 Raw 파일이 없습니다: {raw_path}")
        payload = raw_path.read_bytes()
        self._pending_raw_snapshots.append(
            (self.start_date, self.end_date, payload)
        )
        log.info(
            f"Raw 스냅샷 보관: {self.start_date} ~ {self.end_date} "
            f"({len(payload):,} bytes)"
        )

    # -----------------------------------------------------------------------
    # 가공 단계
    # -----------------------------------------------------------------------
    def process_collected_data(self) -> bool:
        """수집 단계에서 보관한 월별 Raw를 순서대로 DW에 통합한다."""
        if self.dry_run:
            log.info("생산실적 가공 단계 생략 (dry-run 모드)")
            return True
        if not self._pending_raw_snapshots:
            log.error("가공할 생산실적 Raw 스냅샷이 없습니다.")
            return False

        log.info("=" * 60)
        log.info(
            f"생산실적 가공 단계 시작 — "
            f"{len(self._pending_raw_snapshots)}개 기간 순차 통합"
        )
        log.info("=" * 60)
        with tempfile.TemporaryDirectory(prefix="production_raw_stage_") as tmp:
            tmp_dir = Path(tmp)
            for index, (start, end, payload) in enumerate(
                self._pending_raw_snapshots, start=1
            ):
                snapshot_path = tmp_dir / f"production_raw_{index:02d}.xlsx"
                snapshot_path.write_bytes(payload)
                log.info(
                    f"가공 기간 [{index}/{len(self._pending_raw_snapshots)}]: "
                    f"{start} ~ {end}"
                )
                if not self.consolidate_to_dw(snapshot_path):
                    log.error(f"생산실적 가공 실패: {start} ~ {end}")
                    return False
        return True

    # -----------------------------------------------------------------------
    # 단독 실행 호환: 수집 후 가공
    # -----------------------------------------------------------------------
    def run(self) -> bool:
        if not self.collect():
            return False
        if not self.build_dw:
            log.info("생산실적 가공 단계 생략 (--skip-dw-build)")
            return True
        return self.process_collected_data()


# ---------------------------------------------------------------------------
# 메인 진입점
# ---------------------------------------------------------------------------
def main():
    _setup_logging()
    parser = argparse.ArgumentParser(
        description="MIS 기간별 제품 생산실적 RPA "
                    "(Raw_생산실적 샘플링 → DB_생산실적.xlsx 통합까지 일괄 수행)"
    )
    parser.add_argument(
        "--date", type=str, default=None,
        help="기준 종료일 (YYYY-MM-DD). 미지정 시 D-1 및 이전 누락 자동 복구. "
             "명시 시 해당 월 1일부터 지정일까지 재수집."
    )
    parser.add_argument(
        "--dry-run", action="store_true",
        help="MIS 조회만 실행, Excel 기록하지 않음 (DW 통합도 생략)"
    )
    parser.add_argument(
        # 3종 RPA 가 같은 이름을 쓰도록 --skip-build 로 통일. 구 이름은 별칭 유지.
        "--skip-build", "--skip-dw-build", dest="skip_build", action="store_true",
        help="MIS 수집만 하고 DB_생산실적.xlsx 통합(가공)은 생략 "
             "(가공은 build_production_dataset.py 로 별도 수행)"
    )
    parser.add_argument(
        "--dw-output", type=str, default=None,
        help="DW 통합 결과 출력 경로. 미지정 시 "
             "production_dw_service.DEFAULT_OUTPUT_PATH 사용 "
             "(E:\\DB_MIS\\DB_생산실적.xlsx)"
    )
    args = parser.parse_args()

    rpa = MISProductionRPA(
        ref_date=args.date,
        dry_run=args.dry_run,
        build_dw=not args.skip_build,
        dw_output=args.dw_output,
    )
    rpa.run()


if __name__ == "__main__":
    main()
