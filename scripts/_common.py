# MIS RPA 공통 헬퍼
"""
MIS RPA 스크립트들이 공유하는 헬퍼:
  - get_clipboard_text   : Windows 클립보드에서 텍스트 추출 (Csv/UNICODETEXT/TEXT 폴백)
  - parse_clipboard_rows : 클립보드 텍스트 → 2D 리스트 (TAB/COMMA 자동 감지)
  - coerce_value         : 문자열 → 숫자 자동 변환 (천단위 쉼표 허용)
  - paste_to_sheet       : 워크북의 특정 시트에 통째 paste (기존 데이터 클리어 후 교체)
"""
from __future__ import annotations

import csv
import io
import logging
import os
import sys
import time
from pathlib import Path

import openpyxl

logger = logging.getLogger(__name__)

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from mis_rpa.config import sampled_db_path_str


def sampled_db_path(filename: str, env_name: str) -> str:
    """Return a source workbook path, allowing .env to override the default sampled DB dir."""
    return sampled_db_path_str(filename, env_name)


# ---------------------------------------------------------------------------
# 빠른 raw mouse click — pywinauto UIA 의 click_input 오버헤드 우회
# ---------------------------------------------------------------------------
# pywinauto.BaseWrapper.click_input() 은 호출마다 내부적으로
#   window_text(), is_dialog(), is_visible(), is_enabled() 등 UIA 속성 접근을
#   여러 번 수행 → UIA 백엔드에서 클릭 한 번에 300~800ms 소요.
# 이 함수는 pywinauto.mouse 의 raw SendInput 이벤트를 절대 좌표로 직접 보냄.
#   - main_window 의 절대 좌상단 좌표는 최초 한 번만 측정해 _WINDOW_ORIGIN_CACHE
#     에 캐싱 → 이후 호출은 UIA 호출 0건.
#   - 윈도우가 이동했거나 새 main_window 객체로 교체된 경우 reset_window_origin_cache().
_WINDOW_ORIGIN_CACHE: dict = {}


def fast_click(main_window, x: int, y: int, double: bool = False) -> None:
    """절대 좌표 기반 마우스 클릭 — click_input 대비 ~50x 빠름.

    데스크탑이 비활성(화면 잠금/RDP 끊김/모니터 절전) 상태면 SetCursorPos 가
    'There is no active desktop' RuntimeError 를 던진다. 일시적인 절전·잠금
    가능성을 위해 최대 30초 동안 1초 간격으로 폴링 재시도하고, 그래도 안되면
    사용자가 바로 원인을 알 수 있는 메시지로 raise.
    """
    from pywinauto.mouse import click as _raw_click, double_click as _raw_dclick

    key = id(main_window)
    origin = _WINDOW_ORIGIN_CACHE.get(key)
    if origin is None:
        rect = main_window.rectangle()
        origin = (rect.left, rect.top)
        _WINDOW_ORIGIN_CACHE[key] = origin

    abs_xy = (origin[0] + x, origin[1] + y)
    do_click = _raw_dclick if double else _raw_click

    try:
        do_click(coords=abs_xy)
        return
    except RuntimeError as e:
        if "no active desktop" not in str(e).lower():
            raise
        logger.warning(
            "  데스크탑 비활성 감지 (화면잠금/RDP끊김/모니터절전?) — 30초 대기 후 재시도"
        )

    deadline = time.monotonic() + 30.0
    while time.monotonic() < deadline:
        time.sleep(1.0)
        try:
            do_click(coords=abs_xy)
            logger.info("  데스크탑 복귀 확인 — 클릭 재시도 성공")
            return
        except RuntimeError as e:
            if "no active desktop" not in str(e).lower():
                raise

    raise RuntimeError(
        "데스크탑이 비활성 상태입니다 (30초 대기 후에도 복구 안됨). "
        "RPA 실행 중에는 화면 잠금 / 화면 보호기 / 모니터 절전 / RDP 세션 끊김이 "
        "발생하지 않도록 해주세요. "
        "(전원 옵션 → 디스플레이 끄기 시간을 '안 함'으로, Windows+L 잠금 금지)"
    )


def reset_window_origin_cache() -> None:
    """MIS 윈도우가 이동·재생성된 경우 호출."""
    _WINDOW_ORIGIN_CACHE.clear()


# ---------------------------------------------------------------------------
# MIS 윈도우 빠른 탐색 — UIA title_re 스캔 회피
# ---------------------------------------------------------------------------
# Application(backend="uia").connect(title_re=...) 는 바탕화면 전체 UIA 트리를
# 순회하며 제목 정규식을 매칭한다. UIA 트리가 크거나 다른 앱이 바쁜 순간엔 이
# 스캔이 수 초~20초+ 까지 들쭉날쭉 걸린다 (connect 자체의 timeout 과는 무관).
#   해법: 순수 Win32 EnumWindows(~수십 ms)로 HWND 를 먼저 찾고
#         connect(handle=hwnd) 로 붙는다. 핸들 연결은 트리 스캔이 없어 ~10ms.
def find_mis_window(title_substring: str = "(신)종합정보"):
    """제목에 title_substring 을 포함하는 visible top-level 윈도우의 (hwnd, title).

    없으면 (None, None). 여러 개면 첫 visible 매칭을 반환 → title_re connect 의
    ElementAmbiguousError 를 구조적으로 회피한다.
    """
    import win32gui

    matches: list = []

    def _cb(hwnd, _):
        if win32gui.IsWindowVisible(hwnd):
            title = win32gui.GetWindowText(hwnd)
            if title_substring in title:
                matches.append((hwnd, title))

    win32gui.EnumWindows(_cb, None)
    return matches[0] if matches else (None, None)


# ---------------------------------------------------------------------------
# Atomic xlsx 저장 — 중간 크래시로 인한 파일 손상 방지
# ---------------------------------------------------------------------------
# openpyxl 의 wb.save(path) 는 path 를 직접 덮어쓴다. 저장 도중 프로세스가
# 크래시·강제종료되면 path 가 잘린 zip 으로 남아 다음 load_workbook() 에서
# "There is no item named '[Content_Types].xml'" 같은 오류가 난다.
# 이 헬퍼는 temp 파일에 먼저 save 한 후 os.replace 로 atomic rename — 실패해도
# 원본은 그대로 유지된다.
def atomic_save_workbook(wb, target_path: str) -> None:
    """openpyxl Workbook 을 손상 위험 없이 저장한다."""
    target = os.fspath(target_path)
    tmp = f"{target}.tmp_{os.getpid()}"
    try:
        wb.save(tmp)
        # os.replace 는 Windows·POSIX 모두 atomic rename
        os.replace(tmp, target)
    except Exception:
        # save 가 실패했어도 tmp 파일이 남아있을 수 있으니 정리
        if os.path.exists(tmp):
            try:
                os.remove(tmp)
            except OSError:
                pass
        raise


# ---------------------------------------------------------------------------
# 클립보드 변경 감지 (sequence number 기반)
# ---------------------------------------------------------------------------
# 문제: MIS 그리드 좌상단 복사 버튼은 로딩 중에 클릭하면 입력이 흡수되어
#       클립보드가 갱신되지 않을 수 있다. 기존 코드는 클릭 후 고정 sleep
#       (WAIT_MEDIUM=0.2s) 으로 대기하고 무조건 한 번 읽었기 때문에
#       이전 사이클 데이터가 그대로 다음 시트에 paste 되는 사고 가능성이 있다.
#
# 해법: GetClipboardSequenceNumber() — Windows 가 유지하는 DWORD 카운터로,
#       클립보드 내용이 갱신될 때마다 자동 증가한다. OpenClipboard 락이 필요
#       없는 단일 Win32 콜이라 폴링 비용이 무시할 수준(μs 단위).
#       → 클릭 전 seq 를 스냅, 클릭 후 seq 가 바뀔 때까지 짧은 간격으로 폴링.
#         happy path 는 ~30~50ms 안에 통과(기존 200ms sleep 보다 빠름),
#         로딩이 느린 경우만 길게 기다린다.
def get_clipboard_sequence() -> int:
    """현재 시스템 클립보드 시퀀스 번호.

    클립보드가 갱신될 때마다 OS 가 자동 증가시키는 DWORD. 본 프로세스가 아닌
    다른 프로세스(MIS)가 SetClipboardData 를 호출해도 증가한다. 0 이면 호출
    실패(보통 윈도우 스테이션 접근 불가) — 그 경우 변경 감지 폴백 불가.
    """
    try:
        import win32clipboard
        return win32clipboard.GetClipboardSequenceNumber()
    except Exception:
        return 0


def wait_for_clipboard_change(
    seq_before: int,
    timeout: float = 2.0,
    poll_interval: float = 0.03,
) -> bool:
    """클립보드 시퀀스 번호가 seq_before 에서 바뀔 때까지 대기.

    Returns True if changed within timeout, False otherwise.
    seq_before == 0 (헬퍼 사용 불가) 면 즉시 False 반환 → 호출부가 폴백 처리.
    """
    if seq_before == 0:
        return False
    deadline = time.monotonic() + timeout
    while time.monotonic() < deadline:
        if get_clipboard_sequence() != seq_before:
            return True
        time.sleep(poll_interval)
    return False


# 클립보드 헬퍼
def get_clipboard_text() -> str:
    """Windows 클립보드에서 MIS 데이터를 읽어온다 (win32clipboard).

    MIS 그리드 복사 결과는 Csv / UNICODETEXT / TEXT 어느 포맷에 들어있을지 모름.
    세 가지 포맷을 순서대로 시도, 비어있지 않은 첫 결과를 반환.
    """
    import win32clipboard

    csv_fmt = win32clipboard.RegisterClipboardFormat("Csv")

    for attempt in range(3):
        try:
            win32clipboard.OpenClipboard()
            try:
                try:
                    data = win32clipboard.GetClipboardData(csv_fmt)
                    if data:
                        if isinstance(data, bytes):
                            for enc in ("utf-8", "euc-kr", "cp949"):
                                try:
                                    text = data.decode(enc).rstrip("\x00")
                                    if text.strip():
                                        logger.info(f"  클립보드 읽기 성공 (Csv/{enc}, {len(text)}자)")
                                        return text
                                except UnicodeDecodeError:
                                    continue
                        elif isinstance(data, str) and data.strip():
                            logger.info(f"  클립보드 읽기 성공 (Csv/str, {len(data)}자)")
                            return data
                except Exception:
                    pass

                try:
                    text = win32clipboard.GetClipboardData(win32clipboard.CF_UNICODETEXT)
                    if text and text.strip():
                        logger.info(f"  클립보드 읽기 성공 (UNICODETEXT, {len(text)}자)")
                        return text
                except Exception:
                    pass

                try:
                    text = win32clipboard.GetClipboardData(win32clipboard.CF_TEXT)
                    if text:
                        decoded = text.decode("utf-8", errors="replace") if isinstance(text, bytes) else text
                        if decoded.strip():
                            logger.info(f"  클립보드 읽기 성공 (TEXT, {len(decoded)}자)")
                            return decoded
                except Exception:
                    pass
            finally:
                win32clipboard.CloseClipboard()
        except Exception as e:
            logger.warning(f"  클립보드 시도 {attempt + 1}/3 실패: {e}")

        time.sleep(0.5)

    logger.warning("  클립보드 읽기 실패 (3회 시도)")
    return ""


# 클립보드 텍스트 → 행렬
def parse_clipboard_rows(raw_text: str) -> list[list[str]]:
    """클립보드 텍스트를 (행, 열) 2차원 리스트로 파싱한다.

    첫 줄에 TAB이 2개 이상이면 TAB 구분, 아니면 COMMA. 빈 행은 제거.
    """
    raw_text = raw_text.strip()
    if not raw_text:
        raise ValueError("클립보드 데이터가 비어 있습니다.")

    logger.info(f"  클립보드 미리보기: {repr(raw_text[:200])}")

    first_line = raw_text.split("\n")[0]
    if "\t" in first_line and first_line.count("\t") >= 2:
        sep = "\t"
    else:
        sep = ","
    logger.info(f"  구분자 감지: {'TAB' if sep == chr(9) else 'COMMA'}")

    reader = csv.reader(io.StringIO(raw_text), delimiter=sep)
    rows = [r for r in reader if any(cell.strip() for cell in r)]
    logger.info(f"  파싱 완료: {len(rows)}행")
    return rows


# 셀 값 변환 (문자열 → 숫자 자동 변환)
def coerce_value(v):
    """문자열을 숫자로 변환 시도 (천단위 쉼표 제거, 정수/실수 자동 판단)."""
    if v is None:
        return None
    v = v.strip().strip('"')
    if v == "":
        return None
    v_num = v.replace(",", "")
    try:
        if "." in v_num:
            return float(v_num)
        return int(v_num)
    except ValueError:
        return v


# 시트 통째 쓰기 (A1셀부터 붙여넣기)
def paste_to_sheet(filepath: str, sheet_name: str, rows: list) -> int:
    """
    파일의 sheet_name 시트 A1셀부터 rows를 통째로 쓴다.
      - 시트 존재 시: 기존 데이터 클리어 후 교체
      - 시트 없으면 : 새로 생성
      - 파일 없으면 : 새로 만들지 않고 경고 후 0 반환 (관리 외 파일 방지)
    """
    if not os.path.exists(filepath):
        logger.warning(f"  대상 파일이 없습니다 (스킵): {filepath}")
        return 0

    lock_path = os.path.join(
        os.path.dirname(filepath), f"~${os.path.basename(filepath)}"
    )
    if os.path.exists(lock_path):
        logger.error(f"  파일이 다른 곳에서 열려있습니다 (스킵): {filepath}")
        return 0

    try:
        wb = openpyxl.load_workbook(filepath)
    except Exception as e:
        logger.error(f"  파일 열기 실패: {e}")
        return 0

    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if ws.max_row >= 1:
            ws.delete_rows(1, ws.max_row)
        logger.info(f"  기존 시트 '{sheet_name}' 데이터 클리어")
    else:
        ws = wb.create_sheet(sheet_name)
        logger.info(f"  신규 시트 '{sheet_name}' 생성")

    for r_idx, row in enumerate(rows, start=1):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=coerce_value(val))

    try:
        atomic_save_workbook(wb, filepath)
    except PermissionError as e:
        logger.error(f"  저장 실패 (파일이 열려있을 수 있음): {e}")
        wb.close()
        return 0
    wb.close()
    logger.info(f"  [{os.path.basename(filepath)} / {sheet_name}] {len(rows)}행 쓰기 완료")
    return len(rows)
