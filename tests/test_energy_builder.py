# 에너지 수집·재가공 파이프라인 검증 (MIS 화면 변경 2026-07 대응).
#
# MIS/E: 드라이브 없이 임시 fixture 로
#   utility_daily_rpa.parse_unit_input_clipboard  ('원단위 실적입력' 클립보드 파싱)
#   utility_daily_rpa.resolve_org_codes           (--factories 해석)
#   energy_builder.write_raw / read_raw            (RawDB_에너지 적재·읽기)
# 를 검증한다.
#
# 실행: python tests/test_energy_builder.py
from __future__ import annotations

import sys
import tempfile
from datetime import date
from pathlib import Path

from openpyxl import Workbook, load_workbook

try:
    sys.stdout.reconfigure(encoding="utf-8")  # type: ignore[attr-defined]
    sys.stderr.reconfigure(encoding="utf-8")  # type: ignore[attr-defined]
except Exception:
    pass

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

import energy_builder as eb  # noqa: E402
import utility_daily_rpa as rpa  # noqa: E402

# 2026-07 남양주1 실측 발췌 (E:\DB_MIS\RawDB_에너지_new.xlsx Sheet1).
# 컬럼: 일자, 냉동전력, 공압기, 전력사용량, 전력비, 전력단가,
#       연료사용량, 연료비, 연료단가, 용수, 폐수, 원수COD, 배출수COD
_SAMPLE_DAYS = [
    (1, 9393, 2435, 35136, 6823950.9, 194.22, 2614, 2496056.32, 954.88, 652, 505, 520, 10),
    (2, 8984, 2458, 34128, 6607428.1, 193.61, 2435, 2325132.8, 954.88, 610, 484, 480, 11),
    (3, 9252, 2815, 34740, 6786505.7, 195.35, 2340, 2234419.2, 954.88, 674, 534, 540, 12),
]
_TOTAL = ("TOTAL", 254784, 62321, 907669, 170117726.6, 185.65,
          58516, 55875758.08, 954.88, 15467, 12116, 17960, 287)
_TRAILING_BLANKS = 5   # MIS 그리드가 뒤에 붙여 내보내는 빈 열 5개


def _cells(row) -> list[str]:
    return [str(v) for v in row] + [" "] * _TRAILING_BLANKS


def _unit_input_clipboard(swap_cost_price: bool = False) -> str:
    """신규 화면 클립보드 텍스트를 재현한다 (머리글 빈 행 + 일자 행 + TOTAL)."""
    rows: list[list[str]] = [[""] * (13 + _TRAILING_BLANKS)]   # MIS 는 첫 행을 비워 내보낸다
    for day in _SAMPLE_DAYS:
        day = list(day)
        if swap_cost_price:
            day[4], day[5] = day[5], day[4]     # 전력비 ↔ 전력단가
            day[7], day[8] = day[8], day[7]     # 연료비 ↔ 연료단가
        rows.append(_cells(day))
    # 아직 실적이 없는 미래 일자 (일 번호만 있고 값은 공백)
    for day_no in (29, 30, 31):
        rows.append([str(day_no)] + [" "] * (12 + _TRAILING_BLANKS))
    rows.append(_cells(_TOTAL))
    return "\n".join(",".join(c for c in row) for row in rows)


def test_parse_unit_input() -> None:
    parsed = rpa.parse_unit_input_clipboard(_unit_input_clipboard(), "2026-07")

    assert set(parsed) == {date(2026, 7, 1), date(2026, 7, 2), date(2026, 7, 3)}, \
        f"일자 파싱 실패: {sorted(parsed)}"

    day1 = parsed[date(2026, 7, 1)]
    assert day1["freezing_power_kwh"] == 9393
    assert day1["air_compressor_kwh"] == 2435
    assert day1["total_power_kwh"] == 35136
    assert day1["power_cost_krw"] == 6823950.9, day1
    assert day1["power_price_krw_kwh"] == 194.22, day1
    assert day1["fuel_nm3"] == 2614
    assert day1["fuel_cost_krw"] == 2496056.32, day1
    assert day1["fuel_price_krw_nm3"] == 954.88, day1
    assert day1["water_ton"] == 652
    assert day1["wastewater_ton"] == 505
    assert day1["influent_cod_ppm"] == 520
    assert day1["effluent_cod_ppm"] == 10
    print("  ✓ 신규 화면 파싱 — TOTAL/빈 일자 제외, 13개 항목 매핑")


def test_cost_price_autodetect() -> None:
    """화면 머리글 순서('사용량→단가→비용')로 나와도 올바르게 배치돼야 한다."""
    parsed = rpa.parse_unit_input_clipboard(
        _unit_input_clipboard(swap_cost_price=True), "2026-07"
    )
    day1 = parsed[date(2026, 7, 1)]
    assert day1["power_cost_krw"] == 6823950.9, day1
    assert day1["power_price_krw_kwh"] == 194.22, day1
    assert day1["fuel_cost_krw"] == 2496056.32, day1
    assert day1["fuel_price_krw_nm3"] == 954.88, day1
    print("  ✓ 비용/단가 열 자동 판별 — 순서가 뒤바뀌어도 교정")


def test_write_and_read_raw() -> None:
    """적재/재읽기 + 과거 값 보존 + 신규 원단위 수식 자동 채움."""
    with tempfile.TemporaryDirectory() as tmp:
        raw_path = Path(tmp) / "RawDB_에너지.xlsx"
        production_path = Path(tmp) / "DB_생산실적.xlsx"
        production_wb = Workbook()
        production_ws = production_wb.active
        production_ws.title = "daily"
        production_ws.append([
            "date", "item_code", "item_name", "factory",
            "category1", "category2", "planned_qty", "actual_qty",
        ])
        for day_number, quantities in {
            1: (100_000, 20_000),
            2: (110_000, 30_000),
            3: (120_000, 40_000),
        }.items():
            for item_code, quantity in zip(("A", "B"), quantities):
                production_ws.append([
                    date(2026, 7, day_number), item_code, item_code, "F10A",
                    "냉장", "MY", 0, quantity,
                ])
        production_wb.save(production_path)
        production_wb.close()

        # DB 통합보다 먼저 완료되는 최신 생산 RawDB. 같은 월의 이전 날짜가
        # 수정되면 이 값이 DB보다 우선해 N열을 다시 덮어써야 한다.
        production_raw_path = Path(tmp) / "RawDB_생산실적.xlsx"
        production_raw_wb = Workbook()
        production_raw_ws = production_raw_wb.active
        production_raw_ws.title = "F10_냉장"
        production_raw_ws.append([
            "__PERIOD__", date(2026, 7, 1), date(2026, 7, 3),
        ])
        production_raw_ws.append([
            "Item Code", "Item 명", "물품대", "누계 계획", "누계 실적",
            "누계 진척률", "01일", "02일", "03일",
        ])
        production_raw_ws.append([
            "A", "바나나우유", 0, 0, 0, 0, 110_000, 130_000, 150_000,
        ])
        production_raw_ws.append([
            "B", "딸기우유", 0, 0, 0, 0, 20_000, 20_000, 20_000,
        ])
        production_raw_wb.save(production_raw_path)
        production_raw_wb.close()

        # 구 화면으로 수집해 둔 과거 값 — 6/30 행에 믹스/원단위가 들어있는 상태
        eb.write_raw({"남양주1": {date(2026, 6, 30): {
            "total_power_kwh": 30000, "mix_prod_kg": 111111,
            "power_per_ton_kwh": 270.0, "fuel_per_ton_nm3": 20.0,
            "water_per_ton_ton": 5.0,
        }}}, raw_path, sync_production=False, recalculate=False)

        # 신규 수집 — unit_input 항목만
        collected = {"남양주1": rpa.parse_unit_input_clipboard(
            _unit_input_clipboard(), "2026-07")}
        eb.write_raw(
            collected, raw_path,
            production_path=production_path, production_raw_path=production_raw_path,
            recalculate=False,
        )

        round_tripped = eb.read_raw(raw_path)
        assert set(round_tripped) == {"남양주1"}
        by_date = round_tripped["남양주1"]
        assert set(by_date) == {date(2026, 6, 30), date(2026, 7, 1),
                                date(2026, 7, 2), date(2026, 7, 3)}, sorted(by_date)

        day1 = by_date[date(2026, 7, 1)]
        assert day1["power_cost_krw"] == 6823950.9
        assert day1["power_price_krw_kwh"] == 194.22
        unit_input_keys = {f.key for f in eb.FIELDS if f.source == "unit_input"}
        assert set(day1) == unit_input_keys | {"mix_prod_kg"}, (
            f"수집 항목 불일치: {sorted(day1)}"
        )
        assert day1["mix_prod_kg"] == 130_000, "최신 Raw 생산량이 DB보다 우선하지 않음"

        # openpyxl은 수식을 계산하지 않으므로 data_only 재읽기에는 원단위가 없지만,
        # 실제 워크북의 신규 날짜 O~S 셀에는 수식이 빠짐없이 들어가야 한다.
        formula_wb = load_workbook(raw_path, data_only=False)
        formula_ws = formula_wb["남양주1"]
        row_by_date = {
            formula_ws.cell(row=row_idx, column=1).value.date(): row_idx
            for row_idx in range(2, formula_ws.max_row + 1)
        }
        for day in (date(2026, 7, 1), date(2026, 7, 2), date(2026, 7, 3)):
            row_idx = row_by_date[day]
            assert formula_ws.cell(row_idx, 15).value == (
                f"=B{row_idx}*1000/$N{row_idx}"
            )
            assert formula_ws.cell(row_idx, 16).value == (
                f"=C{row_idx}*1000/$N{row_idx}"
            )
            assert formula_ws.cell(row_idx, 17).value == (
                f"=D{row_idx}*1000/$N{row_idx}"
            )
            assert formula_ws.cell(row_idx, 18).value == (
                f"=G{row_idx}*1000/$N{row_idx}"
            )
            assert formula_ws.cell(row_idx, 19).value == (
                f"=J{row_idx}*1000/$N{row_idx}"
            )
        formula_wb.close()

        # 과거 legacy 값은 그대로 살아 있어야 한다
        old = by_date[date(2026, 6, 30)]
        assert old["mix_prod_kg"] == 111111, "과거 믹스생산량이 지워짐"
        assert old["power_per_ton_kwh"] == 270.0
        print(f"  ✓ 적재/재읽기 — {len(unit_input_keys)}개 수집 항목 왕복, "
              f"legacy 과거 값 보존")

        # 최신 생산 RawDB에서 이전 날짜가 수정되면 같은 월 재적재 때 N열도 갱신한다.
        production_raw_wb = load_workbook(production_raw_path)
        production_raw_wb["F10_냉장"]["G3"] = 120_000
        production_raw_wb.save(production_raw_path)
        production_raw_wb.close()
        eb._PRODUCTION_RAW_CACHE = None

        # 같은 데이터 재적재 시 행이 늘지 않아야 한다 (upsert)
        eb.write_raw(
            collected, raw_path,
            production_path=production_path, production_raw_path=production_raw_path,
            recalculate=False,
        )
        wb = load_workbook(raw_path)
        assert wb["남양주1"].max_row == 1 + 4, wb["남양주1"].max_row
        row_by_date = {
            wb["남양주1"].cell(row=row_idx, column=1).value.date(): row_idx
            for row_idx in range(2, wb["남양주1"].max_row + 1)
        }
        assert wb["남양주1"].cell(row_by_date[date(2026, 7, 1)], 14).value == 140_000
        wb.close()
        print("  ✓ 월 전체 재동기화 — 이전 날짜 생산량 수정 반영 + 행 중복 없음")


def _make_production_db(path: Path, quantities: dict) -> None:
    """{(공장코드, date): kg} 로 DB_생산실적.xlsx daily 시트 fixture 를 만든다."""
    wb = Workbook()
    ws = wb.active
    ws.title = "daily"
    ws.append(["date", "item_code", "item_name", "factory",
               "category1", "category2", "planned_qty", "actual_qty"])
    for (factory, day), qty in sorted(quantities.items(), key=lambda kv: kv[0][1]):
        ws.append([day, "A", "A", factory, "냉장", "MY", 0, qty])
    wb.save(path)
    wb.close()


def _mix_by_date(raw_path: Path, sheet_name: str) -> dict:
    wb = load_workbook(raw_path, data_only=False)
    ws = wb[sheet_name]
    mix_col = eb._raw_column("mix_prod_kg")
    out = {}
    for row_idx in range(2, ws.max_row + 1):
        day = ws.cell(row=row_idx, column=1).value
        if day is None:
            continue
        out[day.date() if hasattr(day, "date") else day] = ws.cell(
            row=row_idx, column=mix_col).value
    wb.close()
    return out


def test_resync_production() -> None:
    """MIS 없이 믹스생산량만 재집계 — 기간 필터·dry-run·미수집 보존."""
    with tempfile.TemporaryDirectory() as tmp:
        raw_path = Path(tmp) / "RawDB_에너지.xlsx"
        production_path = Path(tmp) / "DB_생산실적.xlsx"

        days = [date(2026, 7, d) for d in (1, 2, 3)]
        # 7/1·7/2 만 생산실적이 있고 7/3 은 아직 없다
        _make_production_db(production_path, {
            ("F10A", days[0]): 200_000,
            ("F10A", days[1]): 210_000,
        })

        # 구 화면에서 받아 둔 값이 남아 있는 상태로 적재 (생산 동기화 없이)
        eb.write_raw(
            {"남양주1": {d: {"total_power_kwh": 30_000, "mix_prod_kg": 111_111}
                       for d in days}},
            raw_path, sync_production=False, recalculate=False,
        )
        assert set(_mix_by_date(raw_path, "남양주1").values()) == {111_111}

        # ── 1. dry-run 은 파일을 바꾸지 않는다 ──
        stats, changes = eb.resync_production(
            raw_path, production_path=production_path,
            recalculate=False, dry_run=True,
        )
        assert stats["남양주1"] == {"updated": 2, "unchanged": 0, "missing": 1}, stats
        assert len(changes) == 2, changes
        assert set(_mix_by_date(raw_path, "남양주1").values()) == {111_111}, \
            "dry-run 인데 파일이 변경됨"
        print("  ✓ 재집계 dry-run — 변경 예정만 보고, 파일 무변경")

        # ── 2. 기간 필터 — 7/1 만 대상 ──
        eb.resync_production(
            raw_path, date_from=days[0], date_to=days[0],
            production_path=production_path, recalculate=False,
        )
        mix = _mix_by_date(raw_path, "남양주1")
        assert mix[days[0]] == 200_000, mix
        assert mix[days[1]] == 111_111, "기간 밖 행이 변경됨"
        print("  ✓ 재집계 기간 필터 — 지정 범위 밖은 손대지 않음")

        # ── 3. 전체 재집계 — 생산실적 없는 날은 기존 값 보존 ──
        stats, changes = eb.resync_production(
            raw_path, production_path=production_path, recalculate=False,
        )
        assert stats["남양주1"] == {"updated": 1, "unchanged": 1, "missing": 1}, stats
        mix = _mix_by_date(raw_path, "남양주1")
        assert mix[days[0]] == 200_000
        assert mix[days[1]] == 210_000
        assert mix[days[2]] == 111_111, "생산실적 없는 날의 기존 값이 지워짐"
        print("  ✓ 재집계 — 생산실적 기준 정정, 미수집일은 보존")

        # ── 4. 원단위 수식이 갱신 행에 채워져 있다 ──
        wb = load_workbook(raw_path, data_only=False)
        ws = wb["남양주1"]
        assert str(ws.cell(2, 17).value).startswith("="), "전력원단위 수식 없음"
        wb.close()

        # ── 5. 재실행 시 변경 없음 (idempotent) ──
        stats, changes = eb.resync_production(
            raw_path, production_path=production_path, recalculate=False,
        )
        assert changes == [], changes
        assert stats["남양주1"]["updated"] == 0, stats
        print("  ✓ 재집계 idempotent — 두 번 돌려도 변경 없음")


def test_resolve_sheet_names() -> None:
    """가공 CLI 의 --factories 는 공장명·공장코드를 모두 받는다."""
    assert eb.resolve_sheet_names(None) == list(eb.FACTORY_SHEETS)
    assert eb.resolve_sheet_names("경산") == ["경산"]
    assert eb.resolve_sheet_names("F50") == ["경산"]
    assert eb.resolve_sheet_names("경산,김해") == ["김해", "경산"]   # 순서 정규화
    assert eb.resolve_sheet_names("경산,F50") == ["경산"]            # 중복 제거
    for bad in ("경산공장", "F99", ","):
        try:
            eb.resolve_sheet_names(bad)
        except SystemExit:
            continue
        raise AssertionError(f"거부되지 않음: {bad!r}")
    print("  ✓ 가공 CLI --factories — 공장명/코드 혼용, 오타 거부")


def test_missing_production_blocks_write() -> None:
    """생산실적 없는 날짜는 빈 원단위를 만들지 말고 적재 전에 중단한다."""
    with tempfile.TemporaryDirectory() as tmp:
        raw_path = Path(tmp) / "RawDB_에너지.xlsx"
        production_path = Path(tmp) / "DB_생산실적.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "daily"
        ws.append(["date", "factory", "actual_qty"])
        ws.append([date(2026, 7, 1), "F10A", 100_000])
        wb.save(production_path)
        wb.close()

        try:
            eb.write_raw(
                {"남양주1": {date(2026, 7, 2): {"total_power_kwh": 1_000}}},
                raw_path, production_path=production_path, recalculate=False,
            )
        except ValueError as exc:
            assert "생산 RPA를 먼저 실행" in str(exc), exc
        else:
            raise AssertionError("생산실적 누락인데 RawDB 적재가 진행됨")
        assert not raw_path.exists(), "검증 실패 전에 RawDB 파일이 생성됨"
        print("  ✓ 생산량 누락 차단 — 빈 분모/원단위 적재 없음")


def test_bems_can_parse_raw() -> None:
    """BEMS 파서가 RawDB 를 **전치 분기 없이** 읽을 수 있어야 한다.

    BEMS `_parse_korean_excel` 은 A열에 항목명이 있으면 전치형으로 간주해 행 필터
    (`metric_mask`)를 적용한다. tidy 형태(A열=날짜)에서는 그 분기를 타지 않아야
    하며, 머리글이 한글 부분매칭 맵에 걸려 필수 컬럼이 모두 잡혀야 한다.
    이 조건이 깨지면 웹앱이 조용히 빈 데이터를 적재한다.
    """
    import pandas as pd

    # BEMS daily_energy_sync_service._KOR_SUBSTR_MAP / EXPECTED_COLUMNS 사본
    kor_substr_map = {
        "날짜": "date", "일자": "date",
        "냉동원단위": "freezing_per_ton_kwh",
        "공압기원단위": "air_compressor_per_ton_kwh",
        "냉동전력량": "freezing_power_kwh", "공압기": "air_compressor_kwh",
        "공업기": "air_compressor_kwh", "공기압축기": "air_compressor_kwh",
        "전력량": "total_power_kwh", "연료량": "fuel_nm3",
        "용수량": "water_ton", "폐수량": "wastewater_ton",
        "mix생산량": "mix_prod_kg", "믹스생산량": "mix_prod_kg",
        "전력원단위": "power_per_ton_kwh", "전력단위": "power_per_ton_kwh",
        "연료원단위": "fuel_per_ton_nm3", "연료단위": "fuel_per_ton_nm3",
        "용수원단위": "water_per_ton_ton", "용수단위": "water_per_ton_ton",
    }
    expected = ["date", "freezing_power_kwh", "air_compressor_kwh",
                "total_power_kwh", "fuel_nm3", "water_ton", "wastewater_ton",
                "mix_prod_kg", "freezing_per_ton_kwh",
                "air_compressor_per_ton_kwh", "power_per_ton_kwh",
                "fuel_per_ton_nm3", "water_per_ton_ton"]

    with tempfile.TemporaryDirectory() as tmp:
        raw_path = Path(tmp) / "RawDB_에너지.xlsx"
        eb.write_raw({"남양주1": rpa.parse_unit_input_clipboard(
            _unit_input_clipboard(), "2026-07")}, raw_path,
            sync_production=False, recalculate=False)

        df = pd.read_excel(raw_path, sheet_name="남양주1", engine="openpyxl")

        first_col = [str(v).replace(" ", "") for v in df.iloc[:, 0].dropna().values]
        is_transposed = any("냉동전력량" in v or "전력량" in v for v in first_col)
        assert not is_transposed, "A열이 항목명으로 인식됨 → 전치 분기를 타 행 필터가 적용된다"

        mapped, unmapped = [], []
        for col in df.columns:
            key = str(col).strip().lower().replace(" ", "")
            hit = next((v for k, v in kor_substr_map.items() if k in key), None)
            (mapped if hit else unmapped).append(hit or str(col).strip())

        missing = [c for c in expected if c not in mapped]
        assert not missing, f"BEMS 필수 컬럼 누락 → validate_columns 실패: {missing}"
        dup = {c for c in mapped if mapped.count(c) > 1}
        assert not dup, f"두 머리글이 같은 컬럼으로 매핑됨: {dup}"

        # 신규 6개 항목은 아직 BEMS 맵에 없어 무시된다 (적재 작업 시 맵 확장 필요)
        new_labels = {f.label for f in eb.FIELDS
                      if f.key in ("power_cost_krw", "power_price_krw_kwh",
                                   "fuel_cost_krw", "fuel_price_krw_nm3",
                                   "influent_cod_ppm", "effluent_cod_ppm")}
        assert new_labels <= set(unmapped), \
            f"신규 항목 매핑 상태 변화: unmapped={unmapped}"
        print("  ✓ BEMS 파서 호환 — tidy 형태로 전치 분기 없이 필수 컬럼 전부 인식")


class _FakeWindow:
    def set_focus(self):
        pass


class _StubRPA(rpa.MISUtilityRPA):
    """MIS 없이 collect_months 흐름만 검증하기 위한 스텁.

    화면 조작은 전부 no-op 이고, copy_grid_data 가 (년월, 공장)별 준비된
    클립보드 텍스트를 돌려준다.
    """

    def __init__(self, clipboards: dict, **kwargs):
        super().__init__(**kwargs)
        self.clipboards = clipboards
        self.main_window = _FakeWindow()
        self.visited: list[tuple[str, str]] = []
        self._current_ym = None
        self._current_org = None

    def navigate_to_screen(self, screen):
        pass

    def set_year_month(self, screen, year_month=None):
        self._current_ym = year_month or self.year_month

    def select_factory(self, screen, org_code):
        self._current_org = org_code

    def click_query(self, screen):
        pass

    def copy_grid_data(self, screen):
        key = (self._current_ym, self._current_org)
        self.visited.append(key)
        return self.clipboards.get(key, "")


def _month_clipboard(day_count: int, base: int) -> str:
    """day_count 일치 가짜 그리드. base 로 공장/월별 값을 구분한다."""
    rows = [[""] * (13 + _TRAILING_BLANKS)]
    for day in range(1, day_count + 1):
        usage = base + day
        price = 200.0
        rows.append(_cells([day, base, base, usage, round(usage * price, 2), price,
                            usage, round(usage * 900.0, 2), 900.0,
                            base, base, base, base]))
    return "\n".join(",".join(row) for row in rows)


def test_collect_months() -> None:
    """과거 월 백필 — 여러 달 순회, 월 단위 체크포인트, 실적 없는 달 처리."""
    clipboards = {
        ("2026-05", "F1A"): _month_clipboard(31, 100),
        ("2026-05", "F1B"): _month_clipboard(31, 200),
        ("2026-06", "F1A"): _month_clipboard(30, 300),
        # ("2026-06", "F1B") 없음 → 빈 클립보드 → 스킵
        ("2026-07", "F1A"): "\n".join([",".join([""] * 18),
                                       ",".join(["TOTAL"] + [" "] * 17)]),
        ("2026-07", "F1B"): _month_clipboard(28, 400),
    }
    stub = _StubRPA(clipboards, year_month="2026-07", dry_run=True,
                    org_codes=["F1A", "F1B"])

    checkpoints: list[tuple[str, int]] = []
    records = stub.collect_months(
        rpa.SCREEN_UNIT_INPUT, rpa.parse_unit_input_clipboard,
        ["2026-05", "2026-06", "2026-07"],
        on_month_done=lambda ym, rec: checkpoints.append(
            (ym, sum(len(v) for v in rec.values()))),
    )

    # 지정한 2개 사업장 × 3개월 = 6회만 순회 (org_codes 필터 동작)
    assert len(stub.visited) == 6, stub.visited
    assert {c for _, c in stub.visited} == {"F1A", "F1B"}

    # 월 단위 체크포인트가 매달 호출됐는지 (2026-06 은 F1A 30일만)
    assert checkpoints == [("2026-05", 62), ("2026-06", 30), ("2026-07", 28)], checkpoints

    # 여러 달이 사업장별로 병합됐는지
    assert records["남양주1"] and records["남양주2"]
    assert len(records["남양주1"]) == 31 + 30, len(records["남양주1"])   # 5월+6월
    assert len(records["남양주2"]) == 31 + 28, len(records["남양주2"])   # 5월+7월
    assert min(records["남양주1"]) == date(2026, 5, 1)
    assert max(records["남양주1"]) == date(2026, 6, 30)

    # 실적 없는 달(2026-07 F1A)이 예외로 중단되지 않고 스킵됐는지
    assert not any(d.month == 7 for d in records["남양주1"]), "빈 그리드가 적재됨"
    print("  ✓ 과거 월 백필 — 다중 월 순회·체크포인트·실적 없는 달 스킵")


def test_resolve_org_codes() -> None:
    """--factories 는 코드/공장명을 모두 받고, 오타는 조용히 통과하지 않아야 한다."""
    assert rpa.resolve_org_codes("논산,경산") == ["F40", "F50"]
    assert rpa.resolve_org_codes("F40,F50") == ["F40", "F50"]
    assert rpa.resolve_org_codes("경산, 논산") == ["F40", "F50"]   # 순서 정규화
    assert rpa.resolve_org_codes("f40") == ["F40"]                  # 대소문자 무시
    assert rpa.resolve_org_codes("논산,F40") == ["F40"]             # 중복 제거
    assert rpa.resolve_org_codes(None) == list(rpa.FACTORY_SHEET_MAP)

    for bad in ("논산공장", "F99", "F4O", ","):
        try:
            rpa.resolve_org_codes(bad)
        except SystemExit:
            continue
        raise AssertionError(f"거부되지 않음: {bad!r}")
    print("  ✓ --factories 해석 — 코드/공장명 혼용, 오타 즉시 거부")


def test_duplicate_grid_guard() -> None:
    """좌표가 어긋나 직전 공장 그리드를 재복사하면 적재를 거부해야 한다.

    2026-07-30 dry-run 에서 김해(y=200)와 논산(y=214)이 13개 항목·29일 전부
    동일한 그리드를 뽑았는데 값만 봐서는 알아채기 어려웠던 사례에 대한 회귀 테스트.
    """
    same = _month_clipboard(29, 500)
    stub = _StubRPA(
        {("2026-07", "F20"): same,          # 김해
         ("2026-07", "F30"): _month_clipboard(29, 600),
         ("2026-07", "F40"): same,          # 논산 — 김해 그리드 재복사
         ("2026-07", "F50"): _month_clipboard(29, 700)},
        year_month="2026-07", dry_run=True,
        org_codes=["F20", "F30", "F40", "F50"],
    )
    records = stub.collect_screen(rpa.SCREEN_UNIT_INPUT,
                                 rpa.parse_unit_input_clipboard)

    assert set(records) == {"김해", "광주", "경산"}, sorted(records)
    assert "논산" not in records, "중복 그리드가 적재됨"
    assert stub.duplicate_grids == [
        (rpa.SCREEN_UNIT_INPUT, "2026-07", "논산", "김해")
    ], stub.duplicate_grids
    assert stub.report_duplicate_grids() is True

    # 정상 케이스에서는 오탐이 없어야 한다
    ok = _StubRPA({("2026-07", "F20"): _month_clipboard(29, 500),
                   ("2026-07", "F40"): _month_clipboard(29, 600)},
                  year_month="2026-07", dry_run=True, org_codes=["F20", "F40"])
    ok.collect_screen(rpa.SCREEN_UNIT_INPUT, rpa.parse_unit_input_clipboard)
    assert ok.duplicate_grids == [], ok.duplicate_grids
    assert ok.report_duplicate_grids() is False
    print("  ✓ 중복 그리드 차단 — 공장 선택 실패 시 적재 거부 + 오탐 없음")


def main() -> int:
    print("에너지 파이프라인 검증")
    for fn in (test_parse_unit_input, test_cost_price_autodetect,
               test_write_and_read_raw, test_missing_production_blocks_write,
               test_resync_production, test_resolve_sheet_names,
               test_bems_can_parse_raw,
               test_collect_months, test_resolve_org_codes,
               test_duplicate_grid_guard):
        fn()
    print("\n전체 통과")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
