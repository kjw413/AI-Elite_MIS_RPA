from __future__ import annotations

import math
import sys
import tempfile
import unittest
from datetime import date
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook, load_workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

import build_energy_dataset as energy_cli  # noqa: E402
import energy_builder as eb  # noqa: E402


DAY = date(2026, 8, 20)


def _write_production_db(path: Path, rows: list[tuple]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "daily"
    ws.append([
        "date", "item_code", "item_name", "factory",
        "category1", "category2", "planned_qty", "actual_qty",
    ])
    for factory, day, item_code, quantity in rows:
        ws.append([
            day, item_code, str(item_code), factory,
            "냉장", "MY", 0, quantity,
        ])
    wb.save(path)
    wb.close()


def _write_production_raw(
    path: Path,
    sheets: dict[str, list[tuple[object, float]]],
) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    for sheet_name, rows in sheets.items():
        ws = wb.create_sheet(sheet_name)
        ws.append(["__PERIOD__", DAY, DAY])
        ws.append([
            "Item Code", "Item 명", "물품대", "누계 계획",
            "누계 실적", "누계 진척률", "20일",
        ])
        for item_code, quantity in rows:
            ws.append([
                item_code, str(item_code), 0, 0, quantity, 0, quantity,
            ])
    wb.save(path)
    wb.close()


def _expected_gwangju(rows: list[tuple[object, float]]) -> float:
    return sum(
        quantity * eb.GWANGJU_MIX_CONVERSION.get(eb._item_code(item_code), 1.0)
        for item_code, quantity in rows
    )


class GwangjuEnergyConversionTests(unittest.TestCase):
    def test_db_loader_converts_each_gwangju_item_once(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            production_path = Path(tmp) / "DB_생산실적.xlsx"
            gwangju_rows = [
                ("100001", 11.0),
                (129998, 2.0),
                ("129999", 3.0),
                ("260014.0", 5.0),
                ("260016", 7.0),
                ("260039", 11.0),
                ("260042", 13.0),
                ("260047", 17.0),
                ("260351", 19.0),
                ("260352", 23.0),
            ]
            db_rows = [
                ("F30", DAY, item_code, quantity)
                for item_code, quantity in gwangju_rows
            ]
            # 동일 광주 행은 한 번만 반영되어야 한다.
            db_rows.append(("F30", DAY, 129998, 2.0))
            # 비광주는 기존 actual_qty 단순 합산(동일 행 포함)을 유지한다.
            db_rows.extend([
                ("F20", DAY, 129998, 2.0),
                ("F20", DAY, 129998, 2.0),
            ])
            _write_production_db(production_path, db_rows)

            actuals = eb._load_production_actuals(production_path)

            self.assertTrue(math.isclose(
                actuals[("F30", "26-08-20")],
                _expected_gwangju(gwangju_rows),
                rel_tol=0.0,
                abs_tol=1e-9,
            ))
            self.assertEqual(actuals[("F20", "26-08-20")], 4.0)
            self.assertEqual(
                eb.GWANGJU_MIX_CONVERSION,
                {
                    "260014": 10.91954,
                    "260016": 1.0,
                    "260039": 1.0,
                    "260042": 4.0,
                    "260047": 1.0,
                    "260351": 1.0,
                    "260352": 1.0,
                    "129998": 10.91954,
                    "129999": 1.0,
                },
            )

    def test_raw_loader_converts_and_overrides_db_without_double_counting(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            production_path = Path(tmp) / "DB_생산실적.xlsx"
            production_raw_path = Path(tmp) / "RawDB_생산실적.xlsx"
            _write_production_db(production_path, [
                ("F30", DAY, "129998", 100.0),
                ("F20", DAY, "129998", 100.0),
            ])
            raw_gwangju = [
                ("129998", 4.0),
                ("260014", 5.0),
                ("100001", 6.0),
            ]
            _write_production_raw(production_raw_path, {
                "F30_냉장": raw_gwangju + [("129998", 4.0)],
                "F20_냉장": [("129998", 4.0)],
            })

            raw_actuals = eb._load_raw_production_actuals(production_raw_path)
            expected_gwangju = _expected_gwangju(raw_gwangju)
            self.assertTrue(math.isclose(
                raw_actuals[("F30", "26-08-20")], expected_gwangju,
                rel_tol=0.0, abs_tol=1e-9,
            ))
            self.assertEqual(raw_actuals[("F20", "26-08-20")], 4.0)

            merged = eb.merge_production_actuals(
                production_path, production_raw_path
            )
            self.assertTrue(math.isclose(
                merged[("F30", "26-08-20")], expected_gwangju,
                rel_tol=0.0, abs_tol=1e-9,
            ))
            self.assertEqual(merged[("F20", "26-08-20")], 4.0)

    def test_build_cli_writes_converted_mix_and_recalculable_unit_formula(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            collection_path = tmp_path / "RawDB_에너지.xlsx"
            output_path = tmp_path / "DB_에너지.xlsx"
            production_path = tmp_path / "DB_생산실적.xlsx"
            production_raw_path = tmp_path / "RawDB_생산실적.xlsx"

            eb.write_collected_raw({
                "광주": {DAY: {"total_power_kwh": 3_000.0}},
                "김해": {DAY: {"total_power_kwh": 2_000.0}},
            }, collection_path)
            _write_production_db(production_path, [
                ("F30", DAY, "129998", 999.0),
                ("F20", DAY, "129998", 999.0),
            ])
            raw_gwangju = [("129998", 10.0), ("260014", 20.0)]
            _write_production_raw(production_raw_path, {
                "F30_냉장": raw_gwangju,
                "F20_냉장": [("129998", 10.0)],
            })

            argv = [
                "build_energy_dataset.py",
                "--raw", str(collection_path),
                "--out", str(output_path),
                "--from", "2026-08",
                "--to", "2026-08",
                "--factories", "광주,김해",
            ]
            with (
                patch.object(eb, "DEFAULT_PRODUCTION_PATH", production_path),
                patch.object(eb, "DEFAULT_PRODUCTION_RAW_PATH", production_raw_path),
                patch.object(eb, "_recalculate_with_excel", return_value=True) as recalc,
                patch.object(sys, "argv", argv),
            ):
                self.assertEqual(energy_cli.main(), 0)

            expected_gwangju = _expected_gwangju(raw_gwangju)
            processed = eb.read_raw(output_path)
            self.assertTrue(math.isclose(
                processed["광주"][DAY]["mix_prod_kg"], expected_gwangju,
                rel_tol=0.0, abs_tol=1e-9,
            ))
            self.assertEqual(processed["김해"][DAY]["mix_prod_kg"], 10.0)
            recalc.assert_called_once_with(output_path)

            wb = load_workbook(output_path, data_only=False)
            try:
                ws = wb["광주"]
                row_idx = 2
                formula = ws.cell(
                    row_idx, eb._raw_column("power_per_ton_kwh")
                ).value
                self.assertEqual(formula, f"=D{row_idx}*1000/$N{row_idx}")
                self.assertTrue(wb.calculation.fullCalcOnLoad)
                expected_unit = 3_000.0 * 1_000.0 / expected_gwangju
                self.assertTrue(math.isclose(
                    expected_unit,
                    3_000.0 * 1_000.0
                    / ws.cell(row_idx, eb._raw_column("mix_prod_kg")).value,
                    rel_tol=0.0,
                    abs_tol=1e-9,
                ))
            finally:
                wb.close()


if __name__ == "__main__":
    unittest.main()
