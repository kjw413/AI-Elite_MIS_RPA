# 신구조(단일 Raw 파일 → 공장별 wide + 제품마스터 + 계획 + daily) 빌드 검증.
#
# MIS/E: 드라이브 없이 임시 fixture 로 production_dw_service.build_dataset 을 검증한다.
# 실행: .\.venv\Scripts\python.exe tests/test_production_dw_new_structure.py
from __future__ import annotations

import sys
import tempfile
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook

try:
    sys.stdout.reconfigure(encoding="utf-8")  # type: ignore[attr-defined]
    sys.stderr.reconfigure(encoding="utf-8")  # type: ignore[attr-defined]
except Exception:
    pass

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

import production_builder as svc  # noqa: E402

HEADER = ["Item Code", "Item 명", "물품대", "누계 계획", "누계 실적", "누계 진척률",
          "01일", "02일", "03일"]


def _write_raw(path: Path, sheets: dict[str, tuple[str, str, list[list]]]) -> None:
    """sheets: {시트명: (start, end, [데이터행...])}. 1행 마커 + 2행 헤더 + 데이터."""
    wb = Workbook()
    wb.remove(wb.active)
    for sn, (start, end, data_rows) in sheets.items():
        ws = wb.create_sheet(sn)
        ws.append([svc.PERIOD_MARKER, start, end])
        ws.append(HEADER)
        for r in data_rows:
            ws.append(r)
    wb.save(path)


def _read_sheet(path: Path, name: str) -> pd.DataFrame:
    return pd.read_excel(path, sheet_name=name, engine="openpyxl")


def main() -> int:
    tmp = Path(tempfile.mkdtemp(prefix="prod_dw_test_"))
    raw = tmp / "RawDB_생산실적.xlsx"
    out = tmp / "DB_생산실적.xlsx"
    failures: list[str] = []

    def check(cond: bool, msg: str):
        print(("  ✓ " if cond else "  ✗ ") + msg)
        if not cond:
            failures.append(msg)

    # ── 1차 빌드 (2026-05) ──
    # F10_냉장: 260014 바나나우유(냉장+MY→F10A), 270001 요플레(냉장+FM→F10B)
    # F20_냉동: 280001 투게더(냉동→IC→F20)
    # F30_냉장_FM: 290001 닥터캡슐(override FM→F30)
    print("[1차 빌드] 2026-05-01~03")
    _write_raw(raw, {
        "F10_냉장": ("2026-05-01", "2026-05-03", [
            ["260014", "바나나우유", 0, 300, 0, 0, 100, 100, 100],
            ["270001", "요플레",     0, 60,  0, 0, 20,  20,  20],
        ]),
        "F20_냉동": ("2026-05-01", "2026-05-03", [
            ["280001", "투게더", 0, 90, 0, 0, 30, 30, 30],
        ]),
        "F30_냉장_FM": ("2026-05-01", "2026-05-03", [
            ["290001", "닥터캡슐", 0, 15, 0, 0, 5, 5, 5],
        ]),
    })
    daily, saved = svc.build_dataset(raw_path=raw, output_path=out)

    wb = load_workbook(out)
    sheets = set(wb.sheetnames)
    # 품목군 시트 = RawDB 카테고리 시트와 동일 (F10_냉장, F20_냉동, F30_냉장_FM)
    check({"F10_냉장", "F20_냉동", "F30_냉장_FM", svc.MASTER_SHEET, svc.PLAN_SHEET, svc.DAILY_SHEET} <= sheets,
          f"품목군 필수 시트 존재 (현재: {sorted(sheets)})")
    check("남양주1" not in sheets and "남양주2" not in sheets, "구 공장명 시트(남양주1/2) 없음")
    check(wb[svc.DAILY_SHEET].sheet_state == "hidden", "daily 시트 숨김 처리")

    # F10 통합 검증: 남양주1(260014/MY)·남양주2(270001/FM)가 같은 F10_냉장 시트에 공존,
    # F10A/F10B 분석 구분은 숨김 daily 시트에서 검증(아래 daily factory 코드).
    f10n_cols = [str(c) for c in _read_sheet(out, "F10_냉장").columns]
    check("260014" in f10n_cols, "F10_냉장에 260014(바나나우유/MY) 컬럼")
    check("270001" in f10n_cols, "F10_냉장에 270001(요플레/FM) 컬럼")

    # category2 / 제품마스터
    master = _read_sheet(out, svc.MASTER_SHEET)
    m = {str(r["품목코드"]): r for _, r in master.iterrows()}
    check(m.get("260014") is not None and m["260014"]["제품유형"] == "MY", "마스터 260014 → MY")
    check(m.get("270001") is not None and m["270001"]["제품유형"] == "FM", "마스터 270001 → FM")
    check(m.get("280001") is not None and m["280001"]["제품유형"] == "IC", "마스터 280001 → IC")
    check(m.get("290001") is not None and m["290001"]["제품유형"] == "FM", "마스터 290001 → FM(override)")
    check(m.get("280001") is not None and m["280001"]["보관유형"] == "냉동", "마스터 280001 보관유형 냉동")

    # daily(호환) factory 코드
    facs = set(daily["factory"].unique())
    check({"F10A", "F10B", "F20", "F30"} <= facs, f"daily factory 코드 {sorted(facs)}")
    # 값 검증: 260014 2026-05-02 actual=100, F10A
    row = daily[(daily["item_code"] == "260014") & (daily["date"].astype(str) == "2026-05-02")]
    check(len(row) == 1 and abs(float(row.iloc[0]["actual_qty"]) - 100) < 1e-6
          and row.iloc[0]["factory"] == "F10A", "daily 260014 2026-05-02 = 100 @F10A")
    # planned 보존: 260014 planned=300
    check(len(row) == 1 and abs(float(row.iloc[0]["planned_qty"]) - 300) < 1e-6, "daily 260014 planned=300")

    # 계획 시트
    plan = _read_sheet(out, svc.PLAN_SHEET)
    pl = plan[(plan["품목코드"].astype(str) == "260014") & (plan["연월"].astype(str) == "2026-05")]
    check(len(pl) == 1 and pl.iloc[0]["공장"] == "남양주1" and abs(float(pl.iloc[0]["계획량"]) - 300) < 1e-6,
          "계획 260014 2026-05 = 300 @남양주1")

    base_may_rows = len(daily)
    backup = tmp / "DB_생산실적.bak.20260601_000000.xlsx"
    with pd.ExcelWriter(backup, engine="openpyxl") as writer:
        _read_sheet(out, svc.DAILY_SHEET).to_excel(writer, sheet_name=svc.DAILY_SHEET, index=False)
        _read_sheet(out, svc.MASTER_SHEET).to_excel(writer, sheet_name=svc.MASTER_SHEET, index=False)
        _read_sheet(out, svc.PLAN_SHEET).to_excel(writer, sheet_name=svc.PLAN_SHEET, index=False)
    out.write_bytes(b"PK\x03\x04partial xlsx without central directory")

    # ── 2차 빌드 (이력 병합): 2026-06 김해 새 데이터 + 2026-05 김해 정정 ──
    print("\n[2차 빌드] 이력 병합 (2026-06 신규 + 2026-05 정정)")
    _write_raw(raw, {
        "F20_냉동": ("2026-06-01", "2026-06-02", [
            ["280001", "투게더", 0, 50, 0, 0, 40, 40, 0],
        ]),
    })
    daily2, _ = svc.build_dataset(raw_path=raw, output_path=out)
    check(len(daily2) > base_may_rows, "손상된 현재 DB 대신 최신 정상 백업에서 이력 병합")

    # 2026-05 남양주1 데이터가 보존되는지 (2차는 김해만 건드림)
    has_may_ny1 = len(daily2[(daily2["factory"] == "F10A") & (daily2["date"].astype(str).str.startswith("2026-05"))]) > 0
    check(has_may_ny1, "2차 빌드 후에도 2026-05 남양주1 이력 보존")
    # 2026-06 김해 추가
    jun_kh = daily2[(daily2["factory"] == "F20") & (daily2["date"].astype(str) == "2026-06-01")]
    check(len(jun_kh) >= 1 and abs(float(jun_kh.iloc[0]["actual_qty"]) - 40) < 1e-6, "2026-06-01 김해 280001=40 추가")
    # 2026-05 김해 보존
    may_kh = daily2[(daily2["factory"] == "F20") & (daily2["date"].astype(str) == "2026-05-02")]
    check(len(may_kh) == 1 and abs(float(may_kh.iloc[0]["actual_qty"]) - 30) < 1e-6, "2026-05 김해 이력 보존(30)")

    print("\n" + ("=" * 50))
    if failures:
        print(f"❌ 실패 {len(failures)}건:")
        for f in failures:
            print(f"   - {f}")
        return 1
    print("✅ 전체 통과")
    print(f"(임시 산출물: {out})")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
