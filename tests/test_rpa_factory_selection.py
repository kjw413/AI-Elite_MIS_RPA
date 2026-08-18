from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook

import WIP_refactoring
import production_builder
from _common import resolve_factory_codes
from utility_daily_rpa import resolve_org_codes


class FactorySelectionTests(unittest.TestCase):
    def test_common_factory_selection_defaults_to_all(self) -> None:
        expected = ["F10", "F20", "F30", "F40", "F50"]
        self.assertEqual(resolve_factory_codes(None), expected)
        self.assertEqual(resolve_factory_codes(""), expected)
        self.assertEqual(resolve_factory_codes("전체"), expected)

    def test_common_factory_selection_accepts_name_code_and_csv(self) -> None:
        self.assertEqual(resolve_factory_codes("광주"), ["F30"])
        self.assertEqual(resolve_factory_codes("f30"), ["F30"])
        self.assertEqual(resolve_factory_codes("광주,김해"), ["F20", "F30"])
        with self.assertRaises(SystemExit):
            resolve_factory_codes("F3O")

    def test_utility_expands_namyangju_group(self) -> None:
        self.assertEqual(resolve_org_codes("남양주"), ["F1A", "F1B"])
        self.assertEqual(resolve_org_codes("F10"), ["F1A", "F1B"])
        self.assertEqual(resolve_org_codes("남양주1"), ["F1A"])
        self.assertEqual(resolve_org_codes("전체"), list(resolve_org_codes(None)))

    def test_partial_production_plan_merge_preserves_other_factory(self) -> None:
        existing = pd.DataFrame([
            {"연월": "2026-08", "공장": "김해", "품목코드": "200", "계획량": 20},
            {"연월": "2026-08", "공장": "광주", "품목코드": "300", "계획량": 30},
        ])
        new_rows = pd.DataFrame([
            {
                "date": "2026-08-01",
                "factory": "F30",
                "item_code": "301",
                "planned_qty": 31,
            }
        ])

        merged = production_builder._merge_plan(existing, new_rows)

        values = {
            (row["공장"], row["품목코드"]): row["계획량"]
            for _, row in merged.iterrows()
        }
        self.assertEqual(values[("김해", "200")], 20)
        self.assertNotIn(("광주", "300"), values)
        self.assertEqual(values[("광주", "301")], 31)

    def test_partial_wip_save_preserves_unselected_sheet(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "wip.xlsx"
            wb = Workbook()
            wb.active.title = "김해"
            wb["김해"]["A1"] = "keep"
            wb.create_sheet("광주")["A1"] = "old"
            wb.save(path)
            wb.close()

            WIP_refactoring.save_to_excel(
                {"광주": pd.DataFrame({"날짜": ["2026-08-01"], "수량": [10]})},
                str(path),
                preserve_existing=True,
            )

            saved = load_workbook(path, data_only=False)
            self.assertIn("김해", saved.sheetnames)
            self.assertEqual(saved["김해"]["A1"].value, "keep")
            self.assertEqual(saved["광주"]["B2"].value, 10)
            saved.close()


if __name__ == "__main__":
    unittest.main()
